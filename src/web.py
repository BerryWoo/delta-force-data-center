import asyncio
import base64
import hashlib
import html
import io
import json
import re
import threading
import time
import uuid
import webbrowser
from collections import defaultdict
from datetime import datetime, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from urllib.request import Request, urlopen

from . import config
from . import vault_store
from .auth import login_and_capture_cookies, load_cookies
from .client import WeGameClient
from .database import Database, sanitize_player_name

GAME_RESULT_MAP = {0: "撤离成功", 1: "撤离失败", 2: "行动超时", 3: "中途退出"}
ROLE_MAP = config.ROLE_MAP
ACTION_JOBS: dict[str, dict] = {}
ACTION_JOBS_LOCK = threading.Lock()
TEAM_REPORT_FILES: dict[str, dict] = {}
TEAM_REPORT_FILES_LOCK = threading.Lock()
AI_CONFIG_PATH = config.BASE_DIR / "config.ai.json"
AI_CONFIG_OBF_PATH = config.BASE_DIR / "config.ai.obf"
AI_CONFIG_OBFUSCATION_KEY = b"Delta Force Data Center local test ai config v1"
LOGO_PATH = config.INFO_DIR / "LOGO.png"
REPORTS_DIR = config.DATA_DIR / "reports"
TEAM_REPORT_INDEX_PATH = REPORTS_DIR / "index.json"
TEAM_REPORT_JOBS: dict[str, dict] = {}
TEAM_REPORT_QUEUE: list[str] = []
TEAM_REPORT_LOCK = threading.Lock()
TEAM_REPORT_WORKER_ACTIVE = False
TEAM_OTHER_TOKEN = "__OTHER_PLAYER__"
TEAM_PLAYER_ID_PREFIX = "id:"
TEAM_PLAYER_NAME_PREFIX = "name:"

COLLECTIBLE_OBJECT_LIST_URL = (
    "https://comm.ams.game.qq.com/ide/?instanceid=661959&sIdeFlow=xXpyy2"
    "&method=dfm/object.list&param={%22primary%22:%22assets%22}"
)


def _team_player_key(player_id: str | None, player_name: str | None = "") -> str:
    pid = str(player_id or "").strip()
    if pid:
        return TEAM_PLAYER_ID_PREFIX + pid
    return TEAM_PLAYER_NAME_PREFIX + sanitize_player_name(player_name)


def _parse_team_player_targets(players_param: str) -> tuple[list[dict], bool]:
    targets: list[dict] = []
    seen: set[str] = set()
    name_counts: dict[str, int] = defaultdict(int)
    use_other_slot = False
    for raw in str(players_param or "").split(","):
        token = raw.strip()
        if not token:
            continue
        if token == TEAM_OTHER_TOKEN:
            use_other_slot = True
            continue
        if token.startswith(TEAM_PLAYER_ID_PREFIX):
            player_id = token[len(TEAM_PLAYER_ID_PREFIX):].strip()
            if not player_id:
                continue
            key = TEAM_PLAYER_ID_PREFIX + player_id
            target = {"key": key, "kind": "id", "player_id": player_id, "player_name": ""}
            if key in seen:
                continue
            seen.add(key)
        else:
            if token.startswith(TEAM_PLAYER_NAME_PREFIX):
                token = token[len(TEAM_PLAYER_NAME_PREFIX):].strip()
            player_name = sanitize_player_name(token)
            if not player_name:
                continue
            base_key = TEAM_PLAYER_NAME_PREFIX + player_name
            name_counts[base_key] += 1
            key = base_key if name_counts[base_key] == 1 else f"{base_key}#{name_counts[base_key]}"
            target = {"key": key, "kind": "name", "player_id": "", "player_name": player_name}
        targets.append(target)
    return targets, use_other_slot


def _team_target_matches(target: dict, player_id: str | None, player_name: str | None) -> bool:
    if target.get("kind") == "id":
        return str(player_id or "").strip() == str(target.get("player_id") or "")
    return sanitize_player_name(player_name) == str(target.get("player_name") or "")


def _match_team_target(player: dict, targets: list[dict]) -> str:
    player_id = player.get("player_id")
    player_name = player.get("player_name")
    for target in targets:
        if _team_target_matches(target, player_id, player_name):
            return str(target.get("key") or "")
    return ""


def _resolve_team_targets_for_room(members: list[dict], targets: list[dict]) -> dict[str, dict]:
    resolved: dict[str, dict] = {}
    used_indexes: set[int] = set()
    for target in targets:
        target_key = str(target.get("key") or "")
        if not target_key:
            continue
        match_index = -1
        if target.get("kind") == "id":
            wanted_id = str(target.get("player_id") or "")
            for idx, member in enumerate(members):
                if idx in used_indexes:
                    continue
                if str(member.get("player_id") or "") == wanted_id:
                    match_index = idx
                    break
        else:
            wanted_name = str(target.get("player_name") or "")
            candidates = [
                (idx, member)
                for idx, member in enumerate(members)
                if idx not in used_indexes
                and sanitize_player_name(member.get("player_name")) == wanted_name
            ]
            candidates.sort(
                key=lambda item: (
                    -_safe_int(item[1].get("is_self")),
                    str(item[1].get("player_id") or ""),
                )
            )
            if candidates:
                match_index = candidates[0][0]
        if match_index >= 0:
            used_indexes.add(match_index)
            resolved[target_key] = members[match_index]
    return resolved


def _team_report_token_label(token: str) -> str:
    token = str(token or "").strip()
    if token == TEAM_OTHER_TOKEN:
        return "其他玩家"
    if "#" in token and token.startswith(TEAM_PLAYER_NAME_PREFIX):
        token = token.rsplit("#", 1)[0]
    if token.startswith(TEAM_PLAYER_ID_PREFIX):
        player_id = token[len(TEAM_PLAYER_ID_PREFIX):].strip()
        return f"ID {player_id[-6:]}" if player_id else ""
    if token.startswith(TEAM_PLAYER_NAME_PREFIX):
        token = token[len(TEAM_PLAYER_NAME_PREFIX):].strip()
    return sanitize_player_name(token)


class TeamReportCanceled(RuntimeError):
    pass


def _open_source_session_snapshot() -> dict:
    """Open-source builds have no software-account, seat, or membership gate."""
    return {
        "logged_in": True,
        "backend_ok": True,
        "backend_url": "local://open-source",
        "open_source": True,
        "user": None,
        "entitlements": {
            "feature_access": _feature_access_all(),
            "member_active": True,
        },
        "accounts": [],
    }


def _feature_access_all() -> dict:
    return {
        "records": True,
        "record_detail": True,
        "items": True,
        "fetch": True,
        "assets": True,
        "analysis": True,
        "trend_analysis": True,
        "team_analysis": True,
        "team_export": True,
        "ai_report": True,
    }


def _require_open_source_session() -> dict:
    return _open_source_session_snapshot()


def _require_open_source_feature(feature_key: str) -> dict:
    return _open_source_session_snapshot()


def _account_availability_status(account: dict | None, snapshot: dict | None = None) -> dict:
    return {"status": "available", "label": "可用"}

FAVICON_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 128 128">
  <defs>
    <linearGradient id="bgGlow" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0" stop-color="#fefefe"/>
      <stop offset="1" stop-color="#edf4f9"/>
    </linearGradient>
    <linearGradient id="greenGrad" x1="0.15" y1="0.1" x2="0.85" y2="0.9">
      <stop offset="0" stop-color="#58e93f"/>
      <stop offset="1" stop-color="#30ce1f"/>
    </linearGradient>
  </defs>
  <rect width="128" height="128" rx="18" fill="url(#bgGlow)"/>
  <rect x="4" y="7" width="120" height="114" rx="14" fill="none" stroke="#215f90" stroke-width="4.5"/>
  <line x1="4" y1="29" x2="124" y2="29" stroke="#215f90" stroke-width="3.8"/>
  <circle cx="13" cy="19" r="3" fill="#215f90"/>
  <circle cx="23" cy="19" r="3" fill="#215f90"/>
  <circle cx="33" cy="19" r="3" fill="#215f90"/>
  <path d="M64 14 30 72h26L45 88h59L75 33l-9 14z" fill="url(#greenGrad)"/>
  <path d="M66 28 49 60h12l-5 10 26-25H71z" fill="#ffffff"/>
  <path d="M63 39 58 48h5l-4 11 12-12h-6l5-8z" fill="url(#greenGrad)"/>
  <line x1="87" y1="37" x2="114" y2="37" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="90" y1="44" x2="114" y2="44" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="93" y1="51" x2="114" y2="51" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="93" y1="58" x2="114" y2="58" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="93" y1="65" x2="114" y2="65" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <path d="M28 109 48 74l16 29z" fill="#bdcad4"/>
  <path d="M33 109 55 87l9 5 11-8 8 4 18-16 20 37z" fill="none" stroke="#215f90" stroke-width="4.6" stroke-linecap="round" stroke-linejoin="round"/>
  <line x1="44" y1="104" x2="44" y2="95" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="54" y1="104" x2="54" y2="98" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="65" y1="104" x2="65" y2="92" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="75" y1="104" x2="75" y2="95" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="85" y1="104" x2="85" y2="88" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
  <line x1="96" y1="104" x2="96" y2="84" stroke="#215f90" stroke-width="3.4" stroke-linecap="round"/>
</svg>"""


def _safe_int(value, default=0):
    try:
        return int(value or 0)
    except Exception:
        return default


def _logs_indicate_login_expired(logs: list[str]) -> bool:
    text = "\n".join(str(item or "") for item in logs)
    return "8025004" in text or "登录信息过期" in text


def _is_non_normal_map(map_name: str) -> bool:
    text = str(map_name or "")
    return bool(text and "普通" not in text)


def _rule_tip(rule: dict) -> str:
    tip = str(rule.get("rule_text", "") or "")
    note = str(rule.get("note", "") or "").strip()
    if note:
        tip += f"\n备注：{note}"
    return tip


def _make_log_func(logs: list[str], log_sink=None):
    def _log(message: str):
        logs.append(message)
        if log_sink:
            log_sink(message)

    return _log


def _prune_action_jobs_locked():
    now = time.time()
    stale = [
        job_id
        for job_id, job in ACTION_JOBS.items()
        if job.get("status") != "running" and now - job.get("updated_at", now) > 3600
    ]
    for job_id in stale:
        ACTION_JOBS.pop(job_id, None)


def _create_action_job(action: str) -> str:
    job_id = uuid.uuid4().hex
    with ACTION_JOBS_LOCK:
        _prune_action_jobs_locked()
        ACTION_JOBS[job_id] = {
            "id": job_id,
            "action": action,
            "status": "running",
            "logs": [],
            "progress": None,
            "result": None,
            "error": "",
            "created_at": time.time(),
            "updated_at": time.time(),
        }
    return job_id


def _append_action_log(job_id: str, message: str):
    with ACTION_JOBS_LOCK:
        job = ACTION_JOBS.get(job_id)
        if not job:
            return
        job["logs"].append(message)
        if len(job["logs"]) > 500:
            job["logs"] = job["logs"][-500:]
        job["updated_at"] = time.time()


def _update_action_progress(job_id: str, progress: dict | None):
    with ACTION_JOBS_LOCK:
        job = ACTION_JOBS.get(job_id)
        if not job:
            return
        job["progress"] = progress
        job["updated_at"] = time.time()


def _finish_action_job(job_id: str, result: dict):
    with ACTION_JOBS_LOCK:
        job = ACTION_JOBS.get(job_id)
        if not job:
            return
        job["status"] = "done"
        job["result"] = result
        job["updated_at"] = time.time()


def _fail_action_job(job_id: str, error: str):
    with ACTION_JOBS_LOCK:
        job = ACTION_JOBS.get(job_id)
        if not job:
            return
        job["status"] = "error"
        job["error"] = error
        job["updated_at"] = time.time()


def _get_action_job(job_id: str) -> dict | None:
    with ACTION_JOBS_LOCK:
        job = ACTION_JOBS.get(job_id)
        if not job:
            return None
        return {
            "id": job["id"],
            "action": job["action"],
            "status": job["status"],
            "logs": list(job["logs"]),
            "progress": job.get("progress"),
            "result": job["result"],
            "error": job["error"],
        }


def _cleanup_credentials():
    vault_store.clear_sensitive_runtime_files()


def _run_login(db: Database) -> dict:
    logs: list[str] = []
    logs.append("准备登录流程")
    logs.append("即将弹出 WeGame 窗口，请完成扫码登录；如页面已进入助手站点，建议点击“查看最近战局”以便更快进入对局记录页并完成接口捕获")
    _cleanup_credentials()
    asyncio.run(login_and_capture_cookies())
    logs.append("登录流程完成，开始读取凭证")
    cookies = load_cookies()
    client = WeGameClient(cookies, db, logger=logs.append)
    try:
        logs.append("加载地图和干员元数据")
        client.fetch_maps()
        client.fetch_agents()
        openid = client.ensure_account()
    finally:
        client.close()
    if not openid:
        raise RuntimeError("登录完成，但未能获取账号信息")
    account = db.get_active_account()
    if not account:
        raise RuntimeError("登录完成，但数据库中没有可用账号")
    logs.append(f"当前账号: {account.get('player_name') or account.get('player_id')}")
    return {"account": account, "logs": logs}


def _run_fetch(
    db: Database, queue: str, count: int, log_sink=None, progress_sink=None
) -> dict:
    logs: list[str] = []
    log = _make_log_func(logs, log_sink)
    total_steps = 3

    def set_progress(
        step: int,
        step_label: str,
        current: int,
        total: int,
        detail: str = "",
    ):
        ratio = 1 if total <= 0 else max(0, min(current / total, 1))
        if progress_sink:
            progress_sink(
                {
                    "total": {
                        "label": "总进度",
                        "current": step,
                        "total": total_steps,
                        "text": f"{step}/{total_steps} {step_label}",
                        "percent": round(((step - 1) + ratio) / total_steps * 100, 1),
                    },
                    "sub": {
                        "label": step_label,
                        "current": current,
                        "total": total,
                        "text": detail or (f"{current}/{total}" if total > 0 else "等待中"),
                        "percent": 100 if total <= 0 else round(ratio * 100, 1),
                    },
                }
            )

    cookies = load_cookies()
    client = WeGameClient(cookies, db, logger=log)
    queue_name = "烽火地带" if queue == "sol" else "全面战场"
    battles: list[dict] = []
    player_id = db.get_active_player_id()
    try:
        before = db.get_record_count(player_id=player_id)
        log(f"开始抓取: {queue_name}，目标 {count} 条")
        set_progress(1, "抓取战绩列表", 0, count, f"0/{count} 条")
        log("加载地图和干员元数据")
        client.fetch_maps()
        client.fetch_agents()
        log("刷新当前登录账号信息")
        client.ensure_account()
        player_id = db.get_active_player_id()
        log("抓取战绩列表")
        battles = client.fetch_battle_list(
            queue=queue,
            target_count=count,
            progress_callback=lambda current, total: set_progress(
                1,
                "抓取战绩列表",
                current,
                total,
                f"{current}/{total} 条",
            ),
        )
        if not battles and _logs_indicate_login_expired(logs):
            raise RuntimeError("登录信息过期，请重新登录 WeGame")
        log(f"战绩列表抓取完成，返回 {len(battles)} 条")
        if battles:
            log("补全缺失的对局详情")
            detail_room_ids = [
                str(r["room_id"])
                for r in db.get_records_without_report(player_id=player_id)
            ]
            detail_total = len(detail_room_ids)
            set_progress(
                2,
                "抓取对局详情",
                0,
                detail_total,
                f"0/{detail_total} 条",
            )
            fetched_report = client.fetch_battle_details(
                room_ids=detail_room_ids,
                player_id=player_id,
                progress_callback=lambda current, total: set_progress(
                    2,
                    "抓取对局详情",
                    current,
                    total,
                    f"{current}/{total} 条",
                ),
            )
            log(f"对局详情补全完成: {fetched_report} 条")
            log("补全缺失的房间详情")
            room_ids = [
                str(r["room_id"])
                for r in db.get_records_without_detail(player_id=player_id)
                if str(r.get("queue") or "sol") == queue
            ]
            room_total = len(room_ids)
            set_progress(
                3,
                "抓取房间详情",
                0,
                room_total,
                f"0/{room_total} 条",
            )
            fetched_room = client.fetch_room_info(
                room_ids=room_ids,
                queue=queue,
                player_id=player_id,
                progress_callback=lambda current, total: set_progress(
                    3,
                    "抓取房间详情",
                    current,
                    total,
                    f"{current}/{total} 条",
                ),
            )
            log(f"房间详情补全完成: {fetched_room} 条")
        else:
            log("本轮没有返回新的战绩列表数据")
            set_progress(2, "抓取对局详情", 0, 0, "无新增战绩，已跳过")
            set_progress(3, "抓取房间详情", 0, 0, "无新增战绩，已跳过")
        client._save_raw_summary()
        after = db.get_record_count(player_id=player_id)
    finally:
        client.close()
    new_records = after - before
    duplicate_count = max(len(battles) - new_records, 0)
    set_progress(3, "抓取房间详情", 1, 1, "已完成")
    log(
        f"抓取结束: 新增 {new_records} 条，重复 {duplicate_count} 条，当前总战绩 {after} 条"
    )
    return {"new_records": new_records, "total_records": after, "logs": logs}


def _run_fetch_smart(
    db: Database, queue: str, count: int = 100, log_sink=None, progress_sink=None
) -> dict:
    logs: list[str] = []
    log = _make_log_func(logs, log_sink)
    total_steps = 3

    def set_progress(
        step: int,
        step_label: str,
        current: int,
        total: int,
        detail: str = "",
    ):
        ratio = 1 if total <= 0 else max(0, min(current / total, 1))
        if progress_sink:
            progress_sink(
                {
                    "total": {
                        "label": "总进度",
                        "current": step,
                        "total": total_steps,
                        "text": f"{step}/{total_steps} {step_label}",
                        "percent": round(((step - 1) + ratio) / total_steps * 100, 1),
                    },
                    "sub": {
                        "label": step_label,
                        "current": current,
                        "total": total,
                        "text": detail or (f"{current}/{total}" if total > 0 else "等待中"),
                        "percent": 100 if total <= 0 else round(ratio * 100, 1),
                    },
                }
            )

    cookies = load_cookies()
    client = WeGameClient(cookies, db, logger=log)
    queue_name = "烽火地带" if queue == "sol" else "全面战场"
    battles: list[dict] = []
    player_id = db.get_active_player_id()
    try:
        before = db.get_record_count(player_id=player_id)
        log(f"开始智能抓取: {queue_name}，上限 {count} 条")
        set_progress(1, "抓取战绩列表", 0, count, f"0/{count} 条，智能翻页中")
        log("加载地图和干员元数据")
        client.fetch_maps()
        client.fetch_agents()
        log("刷新当前登录账号信息")
        client.ensure_account()
        player_id = db.get_active_player_id()
        log("智能翻页抓取战绩列表，遇到重复或无更多数据后停止")
        battles = client.fetch_battle_list(
            queue=queue,
            target_count=count,
            stop_on_duplicate=True,
            progress_callback=lambda current, total: set_progress(
                1,
                "抓取战绩列表",
                current,
                count,
                f"累计 {current}/{count} 条，智能翻页中",
            ),
        )
        if not battles and _logs_indicate_login_expired(logs):
            raise RuntimeError("登录信息过期，请重新登录 WeGame")
        log(f"智能抓取完成，本轮返回 {len(battles)} 条战绩列表")
        if battles:
            log("补全缺失的对局详情")
            detail_room_ids = [
                str(r["room_id"])
                for r in db.get_records_without_report(player_id=player_id)
            ]
            detail_total = len(detail_room_ids)
            set_progress(2, "抓取对局详情", 0, detail_total, f"0/{detail_total} 条")
            fetched_report = client.fetch_battle_details(
                room_ids=detail_room_ids,
                player_id=player_id,
                progress_callback=lambda current, total: set_progress(
                    2,
                    "抓取对局详情",
                    current,
                    total,
                    f"{current}/{total} 条",
                ),
            )
            log(f"对局详情补全完成: {fetched_report} 条")
            room_ids = [
                str(r["room_id"])
                for r in db.get_records_without_detail(player_id=player_id)
                if str(r.get("queue") or "sol") == queue
            ]
            room_total = len(room_ids)
            set_progress(3, "抓取房间详情", 0, room_total, f"0/{room_total} 条")
            fetched_room = client.fetch_room_info(
                room_ids=room_ids,
                queue=queue,
                player_id=player_id,
                progress_callback=lambda current, total: set_progress(
                    3,
                    "抓取房间详情",
                    current,
                    total,
                    f"{current}/{total} 条",
                ),
            )
            log(f"房间详情补全完成: {fetched_room} 条")
        else:
            log("本轮没有新的战绩列表数据")
            set_progress(2, "抓取对局详情", 0, 0, "无新增战绩，已跳过")
            set_progress(3, "抓取房间详情", 0, 0, "无新增战绩，已跳过")
        client._save_raw_summary()
        after = db.get_record_count(player_id=player_id)
    finally:
        client.close()
    new_records = after - before
    duplicate_count = max(len(battles) - new_records, 0)
    set_progress(3, "抓取房间详情", 1, 1, "已完成")
    log(
        f"智能抓取结束: 新增 {new_records} 条，重复 {duplicate_count} 条，当前总战绩 {after} 条"
    )
    return {"new_records": new_records, "total_records": after, "logs": logs}


def _run_fetch_missing_details(db: Database, log_sink=None) -> dict:
    logs: list[str] = []
    log = _make_log_func(logs, log_sink)
    cookies = load_cookies()
    client = WeGameClient(cookies, db, logger=log)
    player_id = db.get_active_player_id()
    try:
        missing_report_rows = db.get_records_without_report(player_id=player_id)
        missing_room_rows = db.get_records_without_detail(player_id=player_id)
        report_room_ids = [str(r["room_id"]) for r in missing_report_rows]
        missing_room_ids = [str(r["room_id"]) for r in missing_room_rows]
        missing = len(set(report_room_ids) | set(missing_room_ids))
        log(f"需要补全的对局数: {missing}")
        if not missing:
            log("没有需要补全的对局")
            return {
                "missing": 0,
                "fetched_report": 0,
                "fetched_room": 0,
                "logs": logs,
            }

        fetched_report = 0
        if report_room_ids:
            log(f"开始补全 BattleReport 详情: {len(report_room_ids)} 条")
            fetched_report = client.fetch_battle_details(report_room_ids)
            log(f"BattleReport 补全完成: {fetched_report} 条")
        else:
            log("BattleReport 详情已完整")

        queue_groups: dict[str, list[str]] = defaultdict(list)
        for row in missing_room_rows:
            queue_groups[str(row["queue"] or "sol")].append(str(row["room_id"]))
        fetched_room = 0
        if queue_groups:
            for queue_key, queue_room_ids in queue_groups.items():
                queue_name = "烽火地带" if queue_key == "sol" else "全面战场"
                log(f"开始补全房间详情: {queue_name} {len(queue_room_ids)} 条")
                queue_fetched = client.fetch_room_info(queue_room_ids, queue=queue_key)
                fetched_room += queue_fetched
                log(f"房间详情补全完成: {queue_name} {queue_fetched} 条")
        else:
            log("房间详情已完整")
        log(
            f"补全结束: 缺失 {missing} 条，获取 {fetched_report} 条详情 / {fetched_room} 条房间详情"
        )
        return {
            "missing": missing,
            "fetched_report": fetched_report,
            "fetched_room": fetched_room,
            "logs": logs,
        }
    finally:
        client.close()


def _refresh_collection(db: Database, log_sink=None) -> dict:
    logs: list[str] = []
    log = _make_log_func(logs, log_sink)
    cookies = load_cookies()
    if not cookies:
        raise RuntimeError("请先登录 WeGame")
    client = WeGameClient(cookies, db, logger=log)
    try:
        log("开始刷新账号资产")
        client.fetch_maps()
        client.fetch_agents()
        client.ensure_account()
        summary = client.fetch_collectibles()
        log(
            f"资产刷新完成：共 {summary.get('total_entries', 0)} 条，"
            f"典藏枪 {summary.get('collectible_guns', 0)} 条"
        )
        return {"summary": summary, "logs": logs}
    finally:
        client.close()


def _extract_collectible_catalog_count(data: dict) -> int:
    cur = data
    for key in ("jData", "data", "data"):
        if isinstance(cur, dict) and key in cur:
            cur = cur[key]
    if isinstance(cur, dict) and isinstance(cur.get("list"), list):
        return len(cur["list"])
    return 0


def _refresh_collectible_catalog() -> dict:
    req = Request(
        COLLECTIBLE_OBJECT_LIST_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            )
        },
    )
    with urlopen(req, timeout=30) as resp:
        raw = resp.read()
    remote = json.loads(raw.decode("utf-8"))
    remote_count = _extract_collectible_catalog_count(remote)
    if remote_count <= 0:
        raise RuntimeError("物品列表接口返回空数据")

    target = config.COLLECTIBLE_OBJECT_OVERRIDE_JSON
    target.parent.mkdir(parents=True, exist_ok=True)
    current_path = target if target.exists() else config.COLLECTIBLE_OBJECT_JSON
    same_as_current = False
    current_count = 0
    if current_path.exists():
        try:
            current = json.loads(current_path.read_text(encoding="utf-8"))
            current_count = _extract_collectible_catalog_count(current)
            same_as_current = current == remote
        except Exception:
            same_as_current = False
    if not same_as_current:
        target.write_text(
            json.dumps(remote, ensure_ascii=False, indent=4),
            encoding="utf-8",
        )
    return {
        "updated": not same_as_current,
        "item_count": remote_count,
        "previous_count": current_count,
        "path": str(target),
    }


def _clear_data(db: Database) -> dict:
    player_id = db.get_active_player_id()
    record_count = db.get_record_count(player_id=player_id)
    collection_count = db.get_collection_summary(player_id=player_id).get(
        "total_entries", 0
    )
    try:
        if player_id:
            item_count = db.conn.execute(
                """SELECT COUNT(*)
                   FROM battles_items bi
                   JOIN Record r ON r.room_id = bi.room_id
                   WHERE r.player_id = ?""",
                (player_id,),
            ).fetchone()[0]
        else:
            item_count = db.conn.execute(
                "SELECT COUNT(*) FROM battles_items"
            ).fetchone()[0]
    except Exception:
        item_count = 0
    db.clear_records(player_id=player_id)
    return {
        "cleared_records": record_count,
        "cleared_items": item_count,
        "cleared_collections": collection_count,
        "logs": [
            f"已清空当前账号数据: 战绩 {record_count} 条，物品 {item_count} 件，资产 {collection_count} 条"
        ],
    }


def _delete_account(db: Database, player_id: str) -> dict:
    account = db.get_account(player_id)
    if not account:
        raise RuntimeError("账号不存在或已被删除")
    record_count = db.get_record_count(player_id=player_id)
    try:
        item_count = db.conn.execute(
            """SELECT COUNT(*)
               FROM battles_items bi
               JOIN Record r ON r.room_id = bi.room_id
               WHERE r.player_id = ?""",
            (player_id,),
        ).fetchone()[0]
    except Exception:
        item_count = 0
    db.clear_records(player_id=player_id)
    removed = db.delete_account(player_id)
    active_account = db.get_active_account() or {}
    removed_name = (
        (removed or {}).get("player_name", "") or (removed or {}).get("player_id", "")
    )
    next_name = active_account.get("player_name", "") or active_account.get(
        "player_id", ""
    )
    logs = [
        f"已删除账号 {removed_name or player_id}: 战绩 {record_count} 条，物品 {item_count} 件"
    ]
    if next_name:
        logs.append(f"当前激活账号已切换为 {next_name}")
    else:
        logs.append("当前数据库已无可用账号")
    return {
        "deleted_account": removed_name or player_id,
        "deleted_player_id": player_id,
        "cleared_records": record_count,
        "cleared_items": item_count,
        "active_account": next_name,
        "active_player_id": active_account.get("player_id", ""),
        "logs": logs,
    }


def _run_action_in_background(job_id: str, runner):
    def _worker():
        db = Database()
        db.connect()
        try:
            result = runner(
                db,
                lambda message: _append_action_log(job_id, message),
                lambda progress: _update_action_progress(job_id, progress),
            )
            _finish_action_job(job_id, result)
        except Exception as e:
            _fail_action_job(job_id, str(e))
        finally:
            try:
                db.close()
            finally:
                APIHandler.action_lock.release()

    threading.Thread(target=_worker, daemon=True).start()


def _xor_bytes(data: bytes, key: bytes) -> bytes:
    result = bytearray(len(data))
    offset = 0
    counter = 0
    while offset < len(data):
        block = hashlib.sha256(key + counter.to_bytes(4, "big")).digest()
        for item in block:
            if offset >= len(data):
                break
            result[offset] = data[offset] ^ item
            offset += 1
        counter += 1
    return bytes(result)


def _load_obfuscated_ai_config(path) -> dict:
    try:
        wrapper = json.loads(path.read_text(encoding="utf-8-sig"))
        payload = str(wrapper.get("payload", "") or "")
        raw = base64.b64decode(payload.encode("ascii"))
        decoded = _xor_bytes(raw, AI_CONFIG_OBFUSCATION_KEY)
        data = json.loads(decoded.decode("utf-8"))
    except Exception as e:
        raise RuntimeError(f"AI 混淆配置文件读取失败: {e}") from e
    if not isinstance(data, dict):
        raise RuntimeError("AI 混淆配置文件格式错误，应为 JSON 对象")
    return data


def _load_ai_config() -> dict:
    if AI_CONFIG_PATH.exists():
        try:
            data = json.loads(AI_CONFIG_PATH.read_text(encoding="utf-8-sig"))
        except json.JSONDecodeError as e:
            raise RuntimeError(f"AI 配置文件格式错误: {e}") from e
    elif AI_CONFIG_OBF_PATH.exists():
        data = _load_obfuscated_ai_config(AI_CONFIG_OBF_PATH)
    else:
        raise RuntimeError(
            "AI 配置文件不存在，请在项目根目录创建 config.ai.json 并填写配置"
        )
    if not isinstance(data, dict):
        raise RuntimeError("AI 配置文件格式错误，应为 JSON 对象")

    base_url = str(data.get("base_url", "") or "").strip().rstrip("/")
    api_key = str(data.get("api_key", "") or "").strip()
    model = str(data.get("model", "") or "").strip()
    if not base_url or not api_key or not model or "请填写" in api_key:
        raise RuntimeError("AI 报告服务未配置，请先配置 API Key")
    timeout_seconds = _safe_int(data.get("timeout_seconds"), 120)
    if timeout_seconds <= 0:
        timeout_seconds = 120
    try:
        temperature = float(data.get("temperature", 0.3))
    except Exception:
        temperature = 0.3
    return {
        "base_url": base_url,
        "api_key": api_key,
        "model": model,
        "timeout_seconds": timeout_seconds,
        "temperature": temperature,
    }


def _read_ai_config_for_editor() -> dict:
    if AI_CONFIG_PATH.exists():
        try:
            data = json.loads(AI_CONFIG_PATH.read_text(encoding="utf-8-sig"))
        except json.JSONDecodeError as e:
            raise RuntimeError(f"AI 配置文件格式错误: {e}") from e
    elif AI_CONFIG_OBF_PATH.exists():
        data = _load_obfuscated_ai_config(AI_CONFIG_OBF_PATH)
    else:
        data = {}
    if not isinstance(data, dict):
        data = {}
    try:
        temperature = float(data.get("temperature", 0.3))
    except Exception:
        temperature = 0.3
    return {
        "base_url": str(data.get("base_url", "https://api.openai.com/v1") or "").strip(),
        "api_key": str(data.get("api_key", "") or "").strip(),
        "model": str(data.get("model", "gpt-4.1-mini") or "").strip(),
        "timeout_seconds": _safe_int(data.get("timeout_seconds"), 120) or 120,
        "temperature": temperature,
    }


def _save_ai_config(data: dict) -> dict:
    base_url = str(data.get("base_url", "") or "").strip().rstrip("/")
    api_key = str(data.get("api_key", "") or "").strip()
    model = str(data.get("model", "") or "").strip()
    if not base_url:
        raise RuntimeError("请填写 AI API 端点")
    if not model:
        raise RuntimeError("请填写模型名称")
    timeout_seconds = _safe_int(data.get("timeout_seconds"), 120)
    if timeout_seconds <= 0:
        timeout_seconds = 120
    try:
        temperature = float(data.get("temperature", 0.3))
    except Exception:
        temperature = 0.3
    temperature = max(0.0, min(2.0, temperature))
    normalized = {
        "base_url": base_url,
        "api_key": api_key,
        "model": model,
        "timeout_seconds": timeout_seconds,
        "temperature": temperature,
    }
    AI_CONFIG_PATH.write_text(
        json.dumps(normalized, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return normalized


def _report_progress_payload(step: int, label: str, detail: str = "") -> dict:
    total_steps = 4
    step = max(1, min(step, total_steps))
    percent = round(step / total_steps * 100, 1)
    return {
        "total": {
            "label": "报告生成",
            "current": step,
            "total": total_steps,
            "text": f"{step}/{total_steps} {label}",
            "percent": percent,
        },
        "sub": {
            "label": label,
            "current": 1,
            "total": 1,
            "text": detail or label,
            "percent": 100,
        },
    }


def _build_team_report_context(
    analysis: dict, players_param: str, start_dt: str = "", end_dt: str = ""
) -> dict:
    players = analysis.get("players") or []
    games = analysis.get("games") or []
    summary = analysis.get("summary") or {}
    high_value_items = analysis.get("high_value_items") or []
    event_times = sorted(
        str(game.get("event_time", "") or "").strip()
        for game in games
        if str(game.get("event_time", "") or "").strip()
    )
    actual_start = event_times[0] if event_times else (start_dt or "未限制")
    actual_end = event_times[-1] if event_times else (end_dt or "未限制")
    from collections import Counter

    result_counter = Counter(
        GAME_RESULT_MAP.get(game.get("game_result"), game.get("game_result"))
        for game in games
    )
    map_counter = Counter(str(game.get("map_name", "") or "未知地图") for game in games)
    total_games = len(games)

    def _player_record_payload(record: dict | None, fallback_name: str) -> dict | None:
        if not record:
            return None
        role_id = str(record.get("armed_force_id", "") or "")
        return {
            "player_id": str(record.get("player_id", "") or ""),
            "player_name": sanitize_player_name(record.get("player_name") or fallback_name),
            "role_name": ROLE_MAP.get(role_id, role_id) or "",
            "result": GAME_RESULT_MAP.get(record.get("game_result"), record.get("game_result")),
            "kill_player": _safe_int(record.get("kill_player")),
            "rescue": _safe_int(record.get("rescue")),
            "original_equipment_price": _safe_int(record.get("original_equipment_price")),
            "gained_price": _safe_int(record.get("gained_price")),
            "profit_loss": _safe_int(record.get("profit_loss")),
        }

    games_brief = []
    for game in games:
        records = game.get("player_records") or []
        record_payloads = []
        for idx, player in enumerate(players):
            record = records[idx] if idx < len(records) else None
            payload = _player_record_payload(record, player.get("player_name", ""))
            if payload:
                record_payloads.append(payload)
        games_brief.append(
            {
                "event_time": game.get("event_time", ""),
                "map_name": game.get("map_name", ""),
                "result": GAME_RESULT_MAP.get(game.get("game_result"), game.get("game_result")),
                "duration_s": _safe_int(game.get("duration_s")),
                "profit_loss": sum(_safe_int(item.get("profit_loss")) for item in record_payloads),
                "player_records": record_payloads,
            }
        )

    return {
        "report_meta": {
            "title": "组队分析报告",
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "start": actual_start,
            "end": actual_end,
            "filter_start": start_dt or "未限制",
            "filter_end": end_dt or "未限制",
            "players": [sanitize_player_name(p.get("player_name", "")) for p in players],
            "selected_players": [
                sanitize_player_name(p.get("player_name", ""))
                for p in players
                if p.get("player_name")
            ],
            "game_count": len(games),
            "included_game_count": len(games),
            "is_truncated": False,
        },
        "summary": {
            "total_games": _safe_int(summary.get("total_games")),
            "total_escaped": _safe_int(summary.get("total_escaped")),
            "evac_rate": summary.get("evac_rate", ""),
            "team_kd": summary.get("team_kd", 0),
            "total_kills": _safe_int(summary.get("total_kills")),
            "total_profit": _safe_int(summary.get("total_profit")),
            "avg_duration": summary.get("avg_duration", 0),
            "common_map": summary.get("common_map", ""),
            "kill_king": sanitize_player_name(summary.get("kill_king", "")),
            "kill_king_kd": summary.get("kill_king_kd", 0),
            "profit_king": sanitize_player_name(summary.get("profit_king", "")),
            "profit_king_val": _safe_int(summary.get("profit_king_val")),
        },
        "players": [
            {
                "player_id": str(p.get("player_id", "") or ""),
                "player_name": sanitize_player_name(p.get("player_name", "")),
                "games": _safe_int(p.get("games")),
                "evac_rate": p.get("evac_rate", ""),
                "kd": round(float(p.get("kd") or 0), 2),
                "avg_pk": round(float(p.get("avg_pk") or 0), 2),
                "avg_eq": round(float(p.get("avg_eq") or 0), 0),
                "total_out": _safe_int(p.get("total_out")),
                "total_profit": _safe_int(p.get("total_profit")),
                "avg_rate": p.get("avg_rate", ""),
                "common_role": p.get("common_role", ""),
            }
            for p in players
        ],
        "result_distribution": [
            {
                "result": str(name or "未知结果"),
                "count": count,
                "ratio": count / total_games if total_games else 0,
            }
            for name, count in result_counter.most_common()
        ],
        "map_distribution": [
            {
                "map_name": name,
                "count": count,
                "ratio": count / total_games if total_games else 0,
            }
            for name, count in map_counter.most_common(5)
        ],
        "high_value_items": [
            {
                "event_time": str(item.get("event_time", "") or ""),
                "room_id": str(item.get("room_id", "") or ""),
                "map_name": str(item.get("map_name", "") or ""),
                "player_name": sanitize_player_name(item.get("player_name", "")),
                "role_name": str(item.get("role_name", "") or ""),
                "item_name": str(item.get("item_name", "") or "-"),
                "num": _safe_int(item.get("num"), 1),
                "price": _safe_int(item.get("price")),
                "total_price": _safe_int(item.get("total_price"))
                or _safe_int(item.get("price")) * max(_safe_int(item.get("num"), 1), 1),
            }
            for item in high_value_items
            if _safe_int(item.get("price")) > 1_000_000
        ],
        "games_brief": games_brief,
    }


def _extract_ai_message_content(data: dict) -> str:
    try:
        return str(data["choices"][0]["message"]["content"] or "").strip()
    except Exception as e:
        raise RuntimeError("AI API 返回格式异常，未找到 choices[0].message.content") from e


def _extract_stream_delta(data: dict) -> str:
    choices = data.get("choices") if isinstance(data, dict) else None
    if not choices:
        return ""
    delta = choices[0].get("delta") or {}
    content = delta.get("content")
    if content is None:
        content = choices[0].get("message", {}).get("content")
    return str(content or "")


def _call_ai_team_report(ai_cfg: dict, report_context: dict, cancel_checker=None) -> str:
    try:
        import httpx
    except ImportError as e:
        raise RuntimeError("AI 调用依赖 httpx 未安装，请执行 `python -m pip install httpx` 后重试") from e

    base_url = ai_cfg["base_url"]
    endpoint = (
        base_url
        if base_url.endswith("/chat/completions")
        else base_url.rstrip("/") + "/chat/completions"
    )
    system_prompt = (
        "你是三角洲行动撤离模式固定队复盘助手。"
        "你只根据用户提供的结构化数据进行分析，不允许编造不存在的数据。"
        "你只输出 JSON，不输出额外说明。"
        "你只负责 PDF 第二部分“AI 分析”和第三部分“AI 结论”，"
        "第一部分客观数据由系统生成，你不能重复生成客观数据表。"
    )
    user_prompt = (
        "请基于以下组队分析客观数据生成“AI 分析”和“AI 结论”两个字段。\n\n"
        "写作目标：\n"
        "1. 这是三角洲行动撤离模式固定队/组队样本复盘，不是通用 FPS 战报。\n"
        "2. 默认优先按猛攻型固定队口径解释；如果样本明显更像搜刮队/运营队，可明确说明后按对应口径解释。\n"
        "3. 只优化主观分析，不要重复第一部分客观数据，不要生成客观表格。\n"
        "4. 所有判断都必须基于输入数据，不要编造地图理解、干员技能、队内分工、战术细节。\n\n"
        "核心分析原则：\n"
        "1. 不要默认高战备、高失败率、高波动就是负面；对猛攻队要先判断这套打法值不值。\n"
        "2. KD 很重要，但不能单独解读，必须与撤离和经济一起解释。\n"
        "3. 必须解释“击杀是否转化为撤离，撤离是否转化为收益”。\n"
        "4. 优先使用以下数据：profit_loss、original_equipment_price、gained_price、high_value_items、kill_player、rescue、duration_s。\n"
        "5. 输入中每名队员也有独立 duration_s，必须使用它分析生存顺序。\n"
        "6. 由于可能存在救援/返场，不要写“首死”，应写“首个掉点倾向”；不要绝对归责。\n"
        "7. 如果样本基本为单图，必须明确说明无法做地图维度比较；不要硬写地图倾向。\n"
        "8. 如果没有明确证据，不要过度推演干员职责；只允许保守描述常用干员组合。\n"
        "9. 不要写“优化落点”，应写“优化开局路线/前压节奏/第一波接敌决策”。\n"
        "10. 如果样本不足，必须明确说明样本不足。\n"
        "11. conclusion_markdown 控制在 500 个汉字以内。\n"
        "12. 必须严格返回 JSON 对象，不要使用 Markdown 代码块，不要输出 JSON 以外的文字。\n\n"
        "analysis_markdown 固定结构：\n"
        "## 队伍概况\n"
        "要求：先用一句话定义本轮更像什么队伍，必须同时包含撤离与盈亏/带出，不能一上来先写 KD。\n"
        "## 生存与撤离分析\n"
        "要求：说明失败主要集中在哪个阶段、是否存在大量速败局、成功局与失败局时长差异、问题更像打不过还是打完带不走。\n"
        "## 经济收益分析\n"
        "要求：说明平均带入是否过高、成功局是否覆盖失败局成本、高价值物品是偶发还是稳定收益来源、整体是高风险高回报还是高风险低回报。\n"
        "## 队员表现分析\n"
        "要求：从经营和执行角度分析谁是稳定收益点、谁是高波动收益点、谁会在失败局放大亏损、谁能把成功局转成高收益；不要写情绪化评价。\n"
        "## 生存顺序分析\n"
        "要求：必须使用每名队员的 duration_s，分析谁有首个掉点倾向、谁更常活到最后、哪些失败是整队同步崩盘、哪些更像单点掉人后连锁崩盘。\n"
        "## 代表性对局\n"
        "要求：固定写最快失败局、最高收益局、最大亏损局或最高价值物品局；每局至少说明时间、结果、时长、总盈亏、关键队员或关键问题。\n\n"
        "conclusion_markdown 固定结构：\n"
        "## 综合判断\n"
        "要求：一句话概括这套打法当前是赚、亏、勉强打平，还是高度依赖少数高光局。\n"
        "## 调整建议\n"
        "要求：给出 2-3 条可执行建议，允许写降装、控成本、调整前压节奏、优化分工、连续速败后的止损策略；不允许写空泛建议。\n"
        "## 风险提示\n"
        "要求：指出样本是否过少、是否单图、是否单核依赖、是否高收益局掩盖整体亏损，用于控制结论强度。\n\n"
        "禁止事项：\n"
        "1. 不要重复第一部分客观数据表。\n"
        "2. 不要把猛攻队简单批成“打法不健康”。\n"
        "3. 不要使用“绝对核心”“拖累队伍”“必须替换”等过猛措辞。\n"
        "4. 不要输出空泛建议，如“提升意识”“多练枪”“更稳一点”“注意配合”。\n\n"
        "返回格式：\n"
        "{\"analysis_markdown\":\"## 队伍概况\\n...\",\"conclusion_markdown\":\"## 综合判断\\n...\"}\n\n"
        "组队分析 JSON：\n"
        f"{json.dumps(report_context, ensure_ascii=False)}"
    )
    payload = {
        "model": ai_cfg["model"],
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": ai_cfg["temperature"],
        "stream": True,
    }
    headers = {
        "Authorization": f"Bearer {ai_cfg['api_key']}",
        "Content-Type": "application/json",
    }
    def check_cancel():
        _check_team_report_cancel(cancel_checker)

    try:
        check_cancel()
        with httpx.Client(timeout=ai_cfg["timeout_seconds"]) as client:
            with client.stream("POST", endpoint, headers=headers, json=payload) as resp:
                if resp.status_code >= 400:
                    text = resp.read().decode("utf-8", errors="replace")[:500]
                    text = text.replace(ai_cfg["api_key"], "***")
                    raise RuntimeError(f"AI API 返回错误 {resp.status_code}: {text}")

                content_type = resp.headers.get("content-type", "").lower()
                if "text/event-stream" not in content_type:
                    body = resp.read()
                    check_cancel()
                    data = json.loads(body.decode("utf-8", errors="replace"))
                    content = _extract_ai_message_content(data)
                    if not content:
                        raise RuntimeError("AI API 返回内容为空")
                    return content

                parts: list[str] = []
                for line in resp.iter_lines():
                    check_cancel()
                    line = str(line or "").strip()
                    if not line or line.startswith(":"):
                        continue
                    if not line.startswith("data:"):
                        continue
                    payload_line = line[5:].strip()
                    if payload_line == "[DONE]":
                        break
                    try:
                        parts.append(_extract_stream_delta(json.loads(payload_line)))
                    except json.JSONDecodeError:
                        continue
                check_cancel()
                content = "".join(parts).strip()
                if not content:
                    raise RuntimeError("AI API 返回内容为空")
                return content
    except httpx.TimeoutException as e:
        raise RuntimeError("AI API 请求超时，请检查模型服务或调大 timeout_seconds") from e
    except httpx.RequestError as e:
        raise RuntimeError(f"AI API 网络请求失败: {e}") from e


def _parse_ai_team_report_sections(ai_text: str) -> dict:
    text = str(ai_text or "").strip()
    if not text:
        raise RuntimeError("AI API 返回内容为空")

    fenced = re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", text, flags=re.I)
    if fenced:
        text = fenced.group(1).strip()

    def load_json(candidate: str):
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            return None

    data = load_json(text)
    if data is None:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            data = load_json(text[start : end + 1])

    if not isinstance(data, dict):
        raise RuntimeError("AI 返回格式异常：未能解析 analysis_markdown 和 conclusion_markdown JSON")

    analysis_markdown = str(data.get("analysis_markdown", "") or "").strip()
    conclusion_markdown = str(data.get("conclusion_markdown", "") or "").strip()
    if not analysis_markdown or not conclusion_markdown:
        raise RuntimeError("AI 返回格式异常：缺少 analysis_markdown 或 conclusion_markdown")
    if len(conclusion_markdown) > 500:
        conclusion_markdown = conclusion_markdown[:497].rstrip() + "..."
    return {
        "analysis_markdown": analysis_markdown,
        "conclusion_markdown": conclusion_markdown,
    }


def _generate_team_report_sections(
    ai_cfg: dict, report_context: dict, cancel_checker=None, progress_callback=None, log_sink=None
) -> dict:
    if str(ai_cfg.get("mode", "") or "").strip().lower() == "cloud":
        ai_cfg = _load_ai_config()
        ai_cfg["mode"] = "local"
    ai_text = _call_ai_team_report(ai_cfg, report_context, cancel_checker)
    return _parse_ai_team_report_sections(ai_text)


def _markdown_to_pdf_bytes(markdown_text: str, report_context: dict) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.fonts import addMapping
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as e:
        raise RuntimeError("PDF 生成依赖 reportlab 未安装，请执行 `python -m pip install reportlab` 后重试") from e

    font_name = "DeltaKpiCJK"
    bold_font_name = "DeltaKpiCJKBold"
    try:
        pdfmetrics.registerFont(TTFont(font_name, r"C:\Windows\Fonts\msyh.ttc"))
        pdfmetrics.registerFont(TTFont(bold_font_name, r"C:\Windows\Fonts\msyhbd.ttc"))
        addMapping(font_name, 0, 0, font_name)
        addMapping(font_name, 1, 0, bold_font_name)
        addMapping(font_name, 0, 1, font_name)
        addMapping(font_name, 1, 1, bold_font_name)
    except Exception:
        font_name = "STSong-Light"
        bold_font_name = "STSong-Light"
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        except Exception:
            font_name = "Helvetica"
            bold_font_name = "Helvetica-Bold"

    def clean_pdf_text(text: str) -> str:
        text = str(text or "").replace("\ufeff", "").replace("\u200b", "").replace("\ufffd", "")
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

    def inline_text(text: str, auto_bold_label: bool = False) -> str:
        text = clean_pdf_text(text).replace("`", "")
        escaped = html.escape(text)
        escaped = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", escaped)
        escaped = re.sub(r"^(\s*\d+[.、]\s*)", r"<b>\1</b>", escaped)
        escaped = re.sub(r"^(\s*[一二三四五六七八九十]+[、.]\s*)", r"<b>\1</b>", escaped)
        if auto_bold_label and "<b>" not in escaped:
            escaped = re.sub(r"^([^：:]{1,24})([：:])", r"<b>\1\2</b>", escaped)
        return escaped

    def split_table_row(line: str) -> list[str]:
        text = clean_pdf_text(line).strip()
        if text.startswith("|"):
            text = text[1:]
        if text.endswith("|"):
            text = text[:-1]
        return [cell.strip() for cell in text.split("|")]

    def is_table_separator(line: str) -> bool:
        cells = split_table_row(line)
        return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell or "") for cell in cells)

    def is_table_start(lines: list[str], idx: int) -> bool:
        return (
            idx + 1 < len(lines)
            and "|" in lines[idx]
            and "|" in lines[idx + 1]
            and is_table_separator(lines[idx + 1])
        )

    def append_markdown_table(story: list, rows: list[list[str]]):
        if not rows:
            return
        max_cols = max(len(row) for row in rows)
        norm_rows = [row + [""] * (max_cols - len(row)) for row in rows]
        table_style = ParagraphStyle(
            "DeltaTableCell",
            parent=styles["body"],
            fontName=font_name,
            fontSize=8.5,
            leading=12,
            spaceAfter=0,
        )
        header_style = ParagraphStyle(
            "DeltaTableHead",
            parent=table_style,
            fontName=bold_font_name,
            textColor=colors.HexColor("#17364c"),
            leading=12,
        )
        table_data = []
        for row_idx, row in enumerate(norm_rows):
            style = header_style if row_idx == 0 else table_style
            table_data.append(
                [Paragraph(inline_text(cell, True), style) for cell in row]
            )
        usable_width = A4[0] - 36 * mm
        col_widths = [usable_width / max_cols] * max_cols
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dfeaf2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#17364c")),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c9d4df")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, 8))

    def is_markdown_block_start(line: str) -> bool:
        text = clean_pdf_text(line).strip()
        return (
            not text
            or text.startswith("|")
            or text.startswith(("- ", "* "))
            or bool(re.match(r"^\s*#{1,6}\s*", text))
            or bool(re.match(r"^\s*\d+[.、]\s+", text))
            or bool(re.match(r"^\s*[一二三四五六七八九十]+[、.]\s+", text))
        )

    def normalize_markdown_lines(text: str) -> list[str]:
        raw_lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
        out: list[str] = []
        buffer: list[str] = []

        def flush():
            if buffer:
                out.append(" ".join(part.strip() for part in buffer if part.strip()))
                buffer.clear()

        for raw in raw_lines:
            line = clean_pdf_text(raw).strip()
            if not line:
                flush()
                out.append("")
                continue
            if is_markdown_block_start(line):
                flush()
                out.append(line)
                continue
            buffer.append(line)
        flush()
        return out

    meta = report_context.get("report_meta", {})
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title="组队分析报告",
    )
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "DeltaTitle",
            parent=base["Title"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=20,
            leading=26,
            textColor=colors.HexColor("#17364c"),
            spaceAfter=12,
            wordWrap="CJK",
        ),
        "h1": ParagraphStyle(
            "DeltaH1",
            parent=base["Heading1"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=16,
            leading=22,
            textColor=colors.HexColor("#1c4b63"),
            spaceBefore=12,
            spaceAfter=8,
            wordWrap="CJK",
        ),
        "h2": ParagraphStyle(
            "DeltaH2",
            parent=base["Heading2"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=13,
            leading=18,
            textColor=colors.HexColor("#1c4b63"),
            spaceBefore=10,
            spaceAfter=6,
            wordWrap="CJK",
        ),
        "body": ParagraphStyle(
            "DeltaBody",
            parent=base["BodyText"],
            fontName=font_name,
            boldFontName=bold_font_name,
            fontSize=10.5,
            leading=16,
            textColor=colors.HexColor("#203040"),
            spaceAfter=6,
            wordWrap="CJK",
        ),
        "meta": ParagraphStyle(
            "DeltaMeta",
            parent=base["BodyText"],
            fontName=font_name,
            boldFontName=bold_font_name,
            fontSize=9,
            leading=14,
            textColor=colors.HexColor("#607585"),
            spaceAfter=4,
            wordWrap="CJK",
        ),
    }
    story = [
        Paragraph("组队分析报告", styles["title"]),
        Paragraph(f"生成时间：{inline_text(meta.get('generated_at', ''))}", styles["meta"]),
        Paragraph(
            "时间范围："
            + inline_text(meta.get("start", "未限制"))
            + " 至 "
            + inline_text(meta.get("end", "未限制")),
            styles["meta"],
        ),
        Paragraph(
            "玩家范围：" + inline_text("、".join(meta.get("players") or []) or "-"),
            styles["meta"],
        ),
        Spacer(1, 8),
    ]

    lines = normalize_markdown_lines(str(markdown_text or ""))
    idx = 0
    while idx < len(lines):
        raw_line = lines[idx]
        line = raw_line.strip()
        if is_table_start(lines, idx):
            table_rows = [split_table_row(lines[idx])]
            idx += 2
            while idx < len(lines) and "|" in lines[idx] and lines[idx].strip():
                table_rows.append(split_table_row(lines[idx]))
                idx += 1
            append_markdown_table(story, table_rows)
            continue
        if not line:
            story.append(Spacer(1, 6))
            idx += 1
            continue
        heading_match = re.match(r"^(#{1,6})\s*(.+?)\s*$", line)
        if heading_match:
            level = len(heading_match.group(1))
            heading = heading_match.group(2).strip()
            if heading and heading != "组队分析报告":
                story.append(Paragraph(inline_text(heading), styles["h1" if level <= 2 else "h2"]))
        elif line.startswith(("- ", "* ")):
            story.append(Paragraph("• " + inline_text(line[2:], True), styles["body"]))
        elif re.match(r"^\d+\.\s+", line):
            story.append(Paragraph(inline_text(line, True), styles["body"]))
        else:
            story.append(Paragraph(inline_text(line, True), styles["body"]))
        idx += 1
    doc.build(story)
    return out.getvalue()


def _build_team_report_pdf_bytes(
    report_context: dict, analysis_markdown: str, conclusion_markdown: str
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.fonts import addMapping
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as e:
        raise RuntimeError("PDF 生成依赖 reportlab 未安装，请执行 `python -m pip install reportlab` 后重试") from e

    font_name = "DeltaKpiCJK"
    bold_font_name = "DeltaKpiCJKBold"
    try:
        pdfmetrics.registerFont(TTFont(font_name, r"C:\Windows\Fonts\msyh.ttc"))
        pdfmetrics.registerFont(TTFont(bold_font_name, r"C:\Windows\Fonts\msyhbd.ttc"))
        addMapping(font_name, 0, 0, font_name)
        addMapping(font_name, 1, 0, bold_font_name)
        addMapping(font_name, 0, 1, font_name)
        addMapping(font_name, 1, 1, bold_font_name)
    except Exception:
        font_name = "STSong-Light"
        bold_font_name = "STSong-Light"
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        except Exception:
            font_name = "Helvetica"
            bold_font_name = "Helvetica-Bold"

    def clean_pdf_text(text: str) -> str:
        text = str(text or "").replace("\ufeff", "").replace("\u200b", "").replace("\ufffd", "")
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

    def inline_text(text: str, auto_bold_label: bool = False) -> str:
        text = clean_pdf_text(text).replace("`", "")
        escaped = html.escape(text)
        escaped = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", escaped)
        escaped = re.sub(r"^(\s*\d+[.、]\s*)", r"<b>\1</b>", escaped)
        escaped = re.sub(r"^(\s*[一二三四五六七八九十]+[、.]\s*)", r"<b>\1</b>", escaped)
        if auto_bold_label and "<b>" not in escaped:
            escaped = re.sub(r"^([^：:]{1,24})([：:])", r"<b>\1\2</b>", escaped)
        return escaped

    def fmt_price(value) -> str:
        return f"{_safe_int(value):,}"

    def fmt_signed_price(value) -> str:
        number = _safe_int(value)
        return f"{number:,}"

    def fmt_duration(value) -> str:
        seconds = _safe_int(value)
        minutes, rest = divmod(max(0, seconds), 60)
        return f"{minutes}分{rest}秒"

    def fmt_float(value, digits: int = 2) -> str:
        try:
            return f"{float(value or 0):.{digits}f}"
        except Exception:
            return f"{0:.{digits}f}"

    def fmt_percent_value(value) -> str:
        if isinstance(value, str) and value.strip().endswith("%"):
            return value.strip()
        try:
            return f"{float(value) * 100:.1f}%"
        except Exception:
            return "0.0%"

    def percent_to_float(value) -> float:
        if isinstance(value, str) and value.strip().endswith("%"):
            try:
                return float(value.strip().rstrip("%")) / 100
            except Exception:
                return 0
        try:
            number = float(value or 0)
            return number if number <= 1 else number / 100
        except Exception:
            return 0

    out = io.BytesIO()
    page_size = A4
    doc = SimpleDocTemplate(
        out,
        pagesize=page_size,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="组队分析报告",
    )
    usable_width = page_size[0] - doc.leftMargin - doc.rightMargin
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "DeltaReportTitle",
            parent=base["Title"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=20,
            leading=26,
            textColor=colors.HexColor("#17364c"),
            spaceAfter=8,
            wordWrap="CJK",
        ),
        "section": ParagraphStyle(
            "DeltaReportSection",
            parent=base["Heading1"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=15,
            leading=20,
            textColor=colors.HexColor("#1c4b63"),
            spaceBefore=12,
            spaceAfter=6,
            wordWrap="CJK",
        ),
        "subsection": ParagraphStyle(
            "DeltaReportSubsection",
            parent=base["Heading2"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=12,
            leading=16,
            textColor=colors.HexColor("#1f5973"),
            spaceBefore=8,
            spaceAfter=5,
            wordWrap="CJK",
        ),
        "body": ParagraphStyle(
            "DeltaReportBody",
            parent=base["BodyText"],
            fontName=font_name,
            boldFontName=bold_font_name,
            fontSize=9.5,
            leading=14,
            textColor=colors.HexColor("#203040"),
            spaceAfter=5,
            wordWrap="CJK",
        ),
        "meta": ParagraphStyle(
            "DeltaReportMeta",
            parent=base["BodyText"],
            fontName=font_name,
            boldFontName=bold_font_name,
            fontSize=8.5,
            leading=12,
            textColor=colors.HexColor("#607585"),
            spaceAfter=3,
            wordWrap="CJK",
        ),
        "cell": ParagraphStyle(
            "DeltaReportCell",
            parent=base["BodyText"],
            fontName=font_name,
            boldFontName=bold_font_name,
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#203040"),
            wordWrap="CJK",
        ),
        "cell_bold": ParagraphStyle(
            "DeltaReportCellBold",
            parent=base["BodyText"],
            fontName=bold_font_name,
            boldFontName=bold_font_name,
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#17364c"),
            wordWrap="CJK",
        ),
        "small": ParagraphStyle(
            "DeltaReportSmall",
            parent=base["BodyText"],
            fontName=font_name,
            boldFontName=bold_font_name,
            fontSize=7,
            leading=9,
            textColor=colors.HexColor("#203040"),
            wordWrap="CJK",
        ),
    }

    def para(text, style_name: str = "cell", auto_bold_label: bool = False):
        return Paragraph(inline_text(text, auto_bold_label), styles[style_name])

    def bar_flow(value, max_value, color_hex: str, width=46 * mm):
        try:
            ratio = abs(float(value or 0)) / abs(float(max_value or 1))
        except Exception:
            ratio = 0
        ratio = max(0.02 if value else 0, min(1, ratio))
        fill_width = width * ratio
        rest_width = max(0.1 * mm, width - fill_width)
        table = Table([["", ""]], colWidths=[fill_width, rest_width], rowHeights=[3.8 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(color_hex)),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#edf3f7")),
                    ("BOX", (0, 0), (-1, -1), 0.2, colors.HexColor("#d6e0e8")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        return table

    def add_table(
        story: list,
        rows: list[list],
        col_widths: list | None = None,
        right_cols: set[int] | None = None,
        small: bool = False,
        extra_style: list | None = None,
    ):
        if not rows:
            return
        right_cols = right_cols or set()
        max_cols = max(len(row) for row in rows)
        if not col_widths:
            col_widths = [usable_width / max_cols] * max_cols
        table_rows = []
        for r_idx, row in enumerate(rows):
            out_row = []
            for cell in row + [""] * (max_cols - len(row)):
                if hasattr(cell, "wrap"):
                    out_row.append(cell)
                else:
                    style_name = "cell_bold" if r_idx == 0 else ("small" if small else "cell")
                    out_row.append(para(cell, style_name, False))
            table_rows.append(out_row)
        table = Table(table_rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        table_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dfeaf2")),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#c9d4df")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ]
        for col in right_cols:
            table_style.append(("ALIGN", (col, 1), (col, -1), "RIGHT"))
        if extra_style:
            table_style.extend(extra_style)
        table.setStyle(TableStyle(table_style))
        story.append(table)
        story.append(Spacer(1, 7))

    def append_markdown(story: list, markdown_text: str):
        for raw in str(markdown_text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            line = clean_pdf_text(raw).strip()
            if not line:
                story.append(Spacer(1, 3))
                continue
            heading = re.match(r"^(#{1,6})\s*(.+?)\s*$", line)
            if heading:
                story.append(Paragraph(inline_text(heading.group(2)), styles["subsection"]))
            elif line.startswith(("- ", "* ")):
                story.append(Paragraph("• " + inline_text(line[2:], True), styles["body"]))
            else:
                story.append(Paragraph(inline_text(line, True), styles["body"]))

    meta = report_context.get("report_meta", {})
    summary = report_context.get("summary", {})
    players = report_context.get("players") or []
    result_distribution = report_context.get("result_distribution") or []
    map_distribution = report_context.get("map_distribution") or []
    high_value_items = report_context.get("high_value_items") or []
    story = [
        Paragraph("组队分析报告", styles["title"]),
        Paragraph("一、客观数据", styles["section"]),
        Paragraph("1. 报告范围", styles["subsection"]),
    ]

    add_table(
        story,
        [
            ["字段", "内容", "字段", "内容"],
            ["生成时间", meta.get("generated_at", "-"), "玩家范围", "、".join(meta.get("players") or []) or "-"],
            ["对局时间范围", f"{meta.get('start', '未限制')} 至 {meta.get('end', '未限制')}", "样本场次", f"{meta.get('game_count', 0)} 局"],
            ["AI 分析样本", f"使用 {meta.get('included_game_count', meta.get('game_count', 0))} 局", "筛选时间范围", f"{meta.get('filter_start', '未限制')} 至 {meta.get('filter_end', '未限制')}"],
        ],
        col_widths=[24 * mm, 60 * mm, 24 * mm, usable_width - 108 * mm],
    )

    story.append(Paragraph("2. 核心摘要", styles["subsection"]))
    add_table(
        story,
        [
            ["指标", "数值", "指标", "数值"],
            ["总场次", summary.get("total_games", 0), "撤离成功", summary.get("total_escaped", 0)],
            ["撤离率", summary.get("evac_rate", "0%"), "队伍 KD", summary.get("team_kd", 0)],
            ["总盈亏", fmt_signed_price(summary.get("total_profit")), "平均时长", fmt_duration(summary.get("avg_duration"))],
            ["常用地图", summary.get("common_map", "-"), "击杀王", f"{summary.get('kill_king', '-') or '-'} / KD {fmt_float(summary.get('kill_king_kd'))}"],
            ["收益王", summary.get("profit_king", "-") or "-", "收益王盈亏", fmt_signed_price(summary.get("profit_king_val"))],
        ],
        col_widths=[24 * mm, 60 * mm, 24 * mm, usable_width - 108 * mm],
        right_cols={1, 3},
    )

    story.append(Paragraph("3. 玩家汇总", styles["subsection"]))
    add_table(
        story,
        [
            ["玩家", "局数", "撤离率", "KD", "常用干员", "平均带入", "总带出", "总盈亏", "收益率"],
            *[
                [
                    p.get("player_name", "-"),
                    p.get("games", 0),
                    p.get("evac_rate", "0%"),
                    fmt_float(p.get("kd")),
                    p.get("common_role", "-"),
                    fmt_price(p.get("avg_eq")),
                    fmt_price(p.get("total_out")),
                    fmt_signed_price(p.get("total_profit")),
                    p.get("avg_rate", "0%"),
                ]
                for p in players
            ],
        ],
        col_widths=[28 * mm, 12 * mm, 16 * mm, 12 * mm, 22 * mm, 22 * mm, 24 * mm, 24 * mm, usable_width - 160 * mm],
        right_cols={1, 2, 3, 5, 6, 7, 8},
    )

    story.append(Paragraph("4. 结果分布", styles["subsection"]))
    result_rows = [["结果", "场次", "占比", "比例条"]]
    for item in result_distribution:
        ratio = float(item.get("ratio") or 0)
        color = "#2ebf7f" if "成功" in item.get("result", "") else "#d86c68"
        if "超时" in item.get("result", "") or "退出" in item.get("result", ""):
            color = "#e6b85c"
        result_rows.append([
            item.get("result", "-"),
            item.get("count", 0),
            fmt_percent_value(ratio),
            bar_flow(ratio, 1, color, 44 * mm),
        ])
    add_table(story, result_rows, col_widths=[38 * mm, 20 * mm, 20 * mm, usable_width - 78 * mm], right_cols={1, 2})

    story.append(Paragraph("5. 地图分布", styles["subsection"]))
    map_rows = [["地图", "场次", "占比", "比例条"]]
    for item in map_distribution:
        ratio = float(item.get("ratio") or 0)
        map_rows.append([
            item.get("map_name", "-"),
            item.get("count", 0),
            fmt_percent_value(ratio),
            bar_flow(ratio, 1, "#3d91d1", 44 * mm),
        ])
    add_table(story, map_rows, col_widths=[50 * mm, 20 * mm, 20 * mm, usable_width - 90 * mm], right_cols={1, 2})

    story.append(Paragraph("6. 带出高价值物品", styles["subsection"]))
    story.append(Paragraph("筛选规则：单价 > 100万", styles["meta"]))
    item_rows = [["时间", "地图", "队员", "干员", "物品", "数量", "单价", "小计"]]
    if high_value_items:
        for item in high_value_items:
            item_rows.append(
                [
                    str(item.get("event_time", "") or "-")[:16],
                    item.get("map_name", "-") or "-",
                    item.get("player_name", "-") or "-",
                    item.get("role_name", "-") or "-",
                    item.get("item_name", "-") or "-",
                    item.get("num", 1),
                    fmt_price(item.get("price")),
                    fmt_price(item.get("total_price") or (_safe_int(item.get("price")) * max(_safe_int(item.get("num"), 1), 1))),
                ]
            )
    else:
        item_rows.append(["-", "-", "-", "-", "本次报告范围内未记录单价超过100万的高价值物品", "-", "-", "-"])
    add_table(
        story,
        item_rows,
        col_widths=[27 * mm, 22 * mm, 24 * mm, 18 * mm, 35 * mm, 10 * mm, 22 * mm, usable_width - 158 * mm],
        right_cols={5, 6, 7},
        small=True,
    )

    story.append(Paragraph("7. 玩家表现对比", styles["subsection"]))
    max_kd = max([float(p.get("kd") or 0) for p in players] or [1])
    max_profit = max([abs(float(p.get("total_profit") or 0)) for p in players] or [1])
    comparison_rows = [["玩家", "KD", "KD 对比", "总盈亏", "盈亏对比", "撤离率", "撤离率对比"]]
    for p in players:
        profit = float(p.get("total_profit") or 0)
        evac_rate = percent_to_float(p.get("evac_rate"))
        comparison_rows.append(
            [
                p.get("player_name", "-"),
                fmt_float(p.get("kd")),
                bar_flow(float(p.get("kd") or 0), max_kd, "#3d91d1", 30 * mm),
                fmt_signed_price(profit),
                bar_flow(profit, max_profit, "#2ebf7f" if profit >= 0 else "#d86c68", 30 * mm),
                p.get("evac_rate", "0%"),
                bar_flow(evac_rate, 1, "#2ebf7f", 30 * mm),
            ]
        )
    add_table(
        story,
        comparison_rows,
        col_widths=[28 * mm, 12 * mm, 34 * mm, 22 * mm, 34 * mm, 18 * mm, usable_width - 148 * mm],
        right_cols={1, 3, 5},
    )

    story.append(Paragraph("二、AI 分析", styles["section"]))
    append_markdown(story, analysis_markdown)
    story.append(Paragraph("三、AI 结论", styles["section"]))
    append_markdown(story, conclusion_markdown)

    doc.build(story)
    return out.getvalue()


def _ensure_reports_dir():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _read_team_report_index() -> dict:
    _ensure_reports_dir()
    data = vault_store.load_team_report_index()
    reports = data.get("reports", data) if isinstance(data, dict) else data
    if not isinstance(reports, list):
        return {}
    out = {}
    for item in reports:
        if isinstance(item, dict) and item.get("id"):
            out[str(item["id"])] = item
    return out


def _write_team_report_index(index: dict):
    _ensure_reports_dir()
    reports = sorted(index.values(), key=lambda item: item.get("created_at", 0))
    vault_store.save_team_report_index({"reports": reports})


def _upsert_team_report_index(entry: dict):
    index = _read_team_report_index()
    index[str(entry["id"])] = entry
    _write_team_report_index(index)


def _delete_team_report_index_entry(report_id: str):
    index = _read_team_report_index()
    if str(report_id) in index:
        index.pop(str(report_id), None)
        _write_team_report_index(index)


def _save_team_report_file(job_id: str, filename: str, content: bytes) -> str:
    _ensure_reports_dir()
    safe_name = re.sub(r'[\\/:*?"<>|]+', "_", filename)
    path = REPORTS_DIR / safe_name
    if path.exists():
        path = REPORTS_DIR / f"{job_id}_{safe_name}"
        safe_name = path.name
    path.write_bytes(content)
    return safe_name


def _get_team_report_file_info(report_id: str) -> dict | None:
    with TEAM_REPORT_LOCK:
        job = TEAM_REPORT_JOBS.get(report_id)
        if job and job.get("filename"):
            path = REPORTS_DIR / str(job["filename"])
            if path.exists():
                return {"filename": job["filename"], "path": path}
    entry = _read_team_report_index().get(report_id)
    if entry and entry.get("filename"):
        path = REPORTS_DIR / str(entry["filename"])
        if path.exists():
            return {"filename": entry["filename"], "path": path}
    return None


def _check_team_report_cancel(cancel_checker=None):
    if cancel_checker and cancel_checker():
        raise TeamReportCanceled("报告生成已取消")


def _run_team_report_job(
    build_analysis,
    job_id: str,
    ai_cfg: dict,
    players_param: str,
    start_dt: str,
    end_dt: str,
    log_sink=None,
    progress_sink=None,
    cancel_checker=None,
) -> dict:
    def log(message: str):
        if log_sink:
            log_sink(message)

    def progress(step: int, label: str, detail: str = ""):
        if progress_sink:
            progress_sink(_report_progress_payload(step, label, detail))

    if str(ai_cfg.get("mode", "") or "").strip().lower() == "cloud":
        ai_cfg = _load_ai_config()
        ai_cfg["mode"] = "local"

    _check_team_report_cancel(cancel_checker)
    log("准备组队分析数据")
    progress(1, "准备组队分析数据")
    analysis = build_analysis(players_param, start_dt, end_dt)
    if analysis.get("error"):
        msg = str(analysis.get("error") or "")
        if msg in {"暂无数据", "无数据"}:
            msg = "当前筛选范围无组队分析数据"
        raise RuntimeError(msg)
    if not analysis.get("games"):
        raise RuntimeError("当前筛选范围无组队分析数据")
    players_label = "、".join(
        sanitize_player_name(p.get("player_name", ""))
        for p in analysis.get("players", [])
        if p.get("player_name")
    )
    if players_label:
        _set_team_report_job(job_id, players_label=players_label)

    _check_team_report_cancel(cancel_checker)
    log("裁剪报告上下文")
    progress(2, "裁剪报告上下文")
    report_context = _build_team_report_context(analysis, players_param, start_dt, end_dt)

    _check_team_report_cancel(cancel_checker)
    log("正在调用 AI 生成报告")
    progress(3, "正在调用 AI 生成分析和结论", ai_cfg["model"])
    ai_sections = _generate_team_report_sections(
        ai_cfg,
        report_context,
        cancel_checker,
        progress_callback=progress,
        log_sink=log,
    )

    _check_team_report_cancel(cancel_checker)
    log("正在生成 PDF")
    progress(4, "正在生成 PDF", "组装客观数据、AI 分析和 AI 结论")
    pdf_bytes = _build_team_report_pdf_bytes(
        report_context,
        ai_sections["analysis_markdown"],
        ai_sections["conclusion_markdown"],
    )
    filename = f"team_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename = _save_team_report_file(job_id, filename, pdf_bytes)
    log(f"PDF 报告已生成: {filename}")
    meta = report_context.get("report_meta", {})
    return {
        "download_url": f"/api/team-report-download?job_id={job_id}",
        "filename": filename,
        "actual_start": meta.get("start", ""),
        "actual_end": meta.get("end", ""),
    }


def _team_report_public(job: dict) -> dict:
    progress = job.get("progress")
    percent = 0
    if isinstance(progress, dict):
        percent = _safe_int((progress.get("total") or {}).get("percent"))
    if job.get("status") == "done":
        percent = 100
    return {
        "id": job.get("id", ""),
        "status": job.get("status", ""),
        "created_at": job.get("created_at", 0),
        "updated_at": job.get("updated_at", 0),
        "started_at": job.get("started_at", 0),
        "ai_started_at": job.get("ai_started_at", 0),
        "ai_timeout_seconds": job.get("ai_timeout_seconds", 0),
        "completed_at": job.get("completed_at", 0),
        "players": job.get("players_label", ""),
        "start": job.get("start", ""),
        "end": job.get("end", ""),
        "logs": list(job.get("logs", []))[-8:],
        "progress": progress,
        "percent": percent,
        "filename": job.get("filename", ""),
        "download_url": job.get("download_url", ""),
        "error": job.get("error", ""),
        "cancel_requested": bool(job.get("cancel_requested")),
    }


def _set_team_report_job(job_id: str, **updates):
    with TEAM_REPORT_LOCK:
        job = TEAM_REPORT_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = time.time()


def _append_team_report_log(job_id: str, message: str):
    with TEAM_REPORT_LOCK:
        job = TEAM_REPORT_JOBS.get(job_id)
        if not job:
            return
        job.setdefault("logs", []).append(message)
        job["logs"] = job["logs"][-100:]
        job["updated_at"] = time.time()


def _update_team_report_progress(job_id: str, progress: dict):
    updates = {"progress": progress}
    total = progress.get("total") if isinstance(progress, dict) else {}
    label = str((total or {}).get("label", "") or "")
    if isinstance(total, dict) and total.get("current") == 3 and "排队" not in label:
        with TEAM_REPORT_LOCK:
            job = TEAM_REPORT_JOBS.get(job_id)
            if job and not job.get("ai_started_at"):
                job["ai_started_at"] = time.time()
                job["updated_at"] = time.time()
        _set_team_report_job(job_id, **updates)
        return
    _set_team_report_job(job_id, **updates)


def _is_team_report_cancel_requested(job_id: str) -> bool:
    with TEAM_REPORT_LOCK:
        job = TEAM_REPORT_JOBS.get(job_id)
        return bool(job and job.get("cancel_requested"))


def _run_team_report_queue():
    global TEAM_REPORT_WORKER_ACTIVE
    while True:
        with TEAM_REPORT_LOCK:
            job_id = ""
            while TEAM_REPORT_QUEUE:
                candidate = TEAM_REPORT_QUEUE.pop(0)
                job = TEAM_REPORT_JOBS.get(candidate)
                if job and job.get("status") == "pending":
                    job_id = candidate
                    job["status"] = "running"
                    job["started_at"] = time.time()
                    job["updated_at"] = time.time()
                    break
            if not job_id:
                TEAM_REPORT_WORKER_ACTIVE = False
                return
            job = dict(TEAM_REPORT_JOBS[job_id])

        db = None
        try:
            db = Database()
            db.connect()
            handler = object.__new__(APIHandler)
            handler.db = db
            result = _run_team_report_job(
                lambda players, start, end: APIHandler._build_team_analysis_data(
                    handler, players, start, end
                ),
                job_id,
                job["ai_cfg"],
                job["players_param"],
                job["start"],
                job["end"],
                log_sink=lambda message, jid=job_id: _append_team_report_log(jid, message),
                progress_sink=lambda progress, jid=job_id: _update_team_report_progress(
                    jid, progress
                ),
                cancel_checker=lambda jid=job_id: _is_team_report_cancel_requested(jid),
            )
            completed_at = time.time()
            _set_team_report_job(
                job_id,
                status="done",
                completed_at=completed_at,
                filename=result.get("filename", ""),
                download_url=result.get("download_url", ""),
                start=result.get("actual_start", job.get("start", "")),
                end=result.get("actual_end", job.get("end", "")),
                progress=_report_progress_payload(4, "已完成", result.get("filename", "")),
            )
            with TEAM_REPORT_LOCK:
                public = _team_report_public(TEAM_REPORT_JOBS[job_id])
            _upsert_team_report_index(public)
        except TeamReportCanceled as e:
            _set_team_report_job(
                job_id,
                status="canceled",
                completed_at=time.time(),
                error=str(e),
                progress=_report_progress_payload(4, "已取消", "报告生成已取消"),
            )
        except Exception as e:
            _set_team_report_job(
                job_id,
                status="failed",
                completed_at=time.time(),
                error=str(e),
            )
        finally:
            if db:
                db.close()


def _create_team_report_job(
    ai_cfg: dict, players_param: str, start_dt: str = "", end_dt: str = "", players_label: str = ""
) -> dict:
    global TEAM_REPORT_WORKER_ACTIVE
    job_id = "team-report-" + uuid.uuid4().hex[:12]
    created_at = time.time()
    players_label = sanitize_player_name(players_label) or "、".join(
        label
        for label in (_team_report_token_label(p) for p in players_param.split(","))
        if label
    )
    job = {
        "id": job_id,
        "status": "pending",
        "players_param": players_param,
        "players_label": players_label,
        "start": start_dt,
        "end": end_dt,
        "filter_start": start_dt,
        "filter_end": end_dt,
        "ai_timeout_seconds": _safe_int(ai_cfg.get("timeout_seconds"), 120),
        "ai_started_at": 0,
        "ai_cfg": ai_cfg,
        "logs": ["已加入报告生成队列"],
        "progress": {
            "total": {
                "label": "报告生成",
                "current": 0,
                "total": 4,
                "text": "排队等待",
                "percent": 0,
            },
            "sub": {
                "label": "排队等待",
                "current": 0,
                "total": 1,
                "text": "等待前序报告完成",
                "percent": 0,
            },
        },
        "error": "",
        "filename": "",
        "download_url": "",
        "cancel_requested": False,
        "created_at": created_at,
        "updated_at": created_at,
        "started_at": 0,
        "completed_at": 0,
    }
    with TEAM_REPORT_LOCK:
        TEAM_REPORT_JOBS[job_id] = job
        TEAM_REPORT_QUEUE.append(job_id)
        should_start = not TEAM_REPORT_WORKER_ACTIVE
        if should_start:
            TEAM_REPORT_WORKER_ACTIVE = True
    if should_start:
        threading.Thread(target=_run_team_report_queue, daemon=True).start()
    return _team_report_public(job)


def _list_team_reports() -> list[dict]:
    with TEAM_REPORT_LOCK:
        current = {job_id: _team_report_public(job) for job_id, job in TEAM_REPORT_JOBS.items()}
    index = _read_team_report_index()
    for report_id, report in index.items():
        if report_id not in current:
            current[report_id] = report
    return sorted(current.values(), key=lambda item: item.get("created_at", 0), reverse=True)


def _cancel_team_report(job_id: str) -> dict:
    with TEAM_REPORT_LOCK:
        job = TEAM_REPORT_JOBS.get(job_id)
        if not job:
            return {"ok": False, "error": "报告任务不存在或已过期"}
        if job["status"] == "pending":
            job["status"] = "canceled"
            job["completed_at"] = time.time()
            job["error"] = "报告生成已取消"
            job["progress"] = _report_progress_payload(4, "已取消", "报告生成已取消")
            job["updated_at"] = time.time()
            while job_id in TEAM_REPORT_QUEUE:
                TEAM_REPORT_QUEUE.remove(job_id)
            return {"ok": True}
        if job["status"] == "running":
            job["cancel_requested"] = True
            job["status"] = "canceling"
            job["progress"] = _report_progress_payload(3, "取消中", "等待当前 AI 请求结束后停止")
            job["updated_at"] = time.time()
            return {"ok": True, "message": "已请求取消，当前 AI 请求结束后停止"}
        return {"ok": False, "error": "该报告当前状态不支持取消"}


def _delete_team_report(job_id: str) -> dict:
    with TEAM_REPORT_LOCK:
        job = TEAM_REPORT_JOBS.get(job_id)
        if job and job.get("status") in {"pending", "running", "canceling"}:
            return {"ok": False, "error": "报告仍在生成中，请先取消后再删除"}
        filename = job.get("filename") if job else ""
        if job:
            TEAM_REPORT_JOBS.pop(job_id, None)
    if not filename:
        entry = _read_team_report_index().get(job_id)
        filename = entry.get("filename", "") if entry else ""
    if filename:
        path = REPORTS_DIR / str(filename)
        if path.exists():
            path.unlink()
    _delete_team_report_index_entry(job_id)
    return {"ok": True}

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Delta Force Data Center</title>
<link rel="icon" href="/logo.png" type="image/png">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<style>
:root{color-scheme:dark}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","Segoe UI",sans-serif;background:#0f1923;color:#e0e0e0;min-height:100vh}
.header{background:#1a2634;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;border-bottom:2px solid #2a3a4a}
.brand{display:flex;align-items:center;gap:12px;min-width:0}
.brand-logo{width:42px;height:42px;border-radius:12px;object-fit:cover;box-shadow:0 8px 22px rgba(0,0,0,.28);background:#0f1923;flex-shrink:0}
.brand h1{font-size:20px;color:#00d4aa;margin:0;line-height:1.1}
.brand-subtitle{font-size:11px;color:#718295;letter-spacing:.18em;text-transform:uppercase;margin-top:4px;white-space:nowrap}
.header-right{display:flex;align-items:center;justify-content:flex-end;gap:12px;flex-wrap:wrap}
.header-right .btn{margin-right:0;white-space:nowrap}
.account-pill{display:flex;align-items:center;min-height:38px;padding:0 14px;border-radius:999px;background:#12202d;border:1px solid #2d4358;color:#dbe7f2;font-size:13px;max-width:260px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.btn-muted{background:#223344;color:#dde7ef;border:1px solid #35506a}
.btn-muted:hover{background:#2b4055}
.btn-feedback{background:#153b3f;color:#9bf0d6;border:1px solid #24686d;text-decoration:none;display:inline-flex;align-items:center;justify-content:center}
.btn-feedback:hover{background:#1b4a4f;color:#bfffee}
.btn-danger{background:#4a1d1f;color:#ffd9dc;border:1px solid #7c3138}
.btn-danger:hover{background:#61262a}
.container{padding:16px}
.tabs{display:flex;gap:4px;margin-bottom:16px}
.tab{padding:8px 20px;background:#1a2634;border:1px solid #2a3a4a;border-radius:6px 6px 0 0;cursor:pointer;color:#8899aa;font-size:14px}
.tab.active{background:#162029;border-color:#00d4aa;color:#00d4aa;border-bottom-color:#162029}
.panel{display:none;background:#162029;border:1px solid #2a3a4a;border-radius:0 6px 6px 6px;padding:16px;overflow-x:auto}
#panel-team{overflow-x:visible}
#panel-trends{overflow:visible}
.panel.active{display:block}
#panel-items table{table-layout:fixed}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#1a2c3a;color:#00d4aa;padding:10px 8px;text-align:left;white-space:nowrap;position:sticky;top:0}
td{padding:8px;border-bottom:1px solid #1e2e3e;white-space:nowrap}
tr:hover{background:#1a2c3a}
#panel-records th,#panel-records td{text-align:center}
#panel-records th.money,#panel-records td.money{text-align:right;padding-right:12px}
#panel-records th.tags-cell,#panel-records td.tags-cell{text-align:left}
#panel-records td.tags-cell{position:relative;white-space:normal;min-width:220px;overflow:visible}
#panel-records td.tags-cell:hover{z-index:70}
#panel-records tbody tr{cursor:pointer}
#panel-records tbody tr:hover{background:#203243}
#panel-items th,#panel-items td{text-align:center}
#panel-items th.money,#panel-items td.money{text-align:right;padding-right:12px}
#panel-items tbody tr{cursor:pointer}
#panel-items tbody tr:hover{background:#203243}
.tag-list{display:flex;flex-wrap:wrap;gap:6px}
.battle-tag{position:relative;display:inline-flex;align-items:center;padding:3px 9px;border-radius:999px;font-size:12px;font-weight:700;line-height:1.2;border:1px solid transparent;cursor:help}
.battle-tag:hover{filter:brightness(1.04);z-index:80}
.battle-tag::after,.battle-tag::before{content:none}
.battle-tag.pos{background:#123326;color:#64e0a7;border-color:#1f6247}
.battle-tag.neg{background:#3b1f1f;color:#ff9a8d;border-color:#8f453d}
.battle-tag.mid{background:#3b3420;color:#ffd46a;border-color:#7e6c2a}
.battle-tag-tooltip{display:none;position:fixed;left:0;top:0;z-index:2200;min-width:220px;max-width:320px;padding:8px 10px;border-radius:8px;background:#0e1720;border:1px solid #304456;color:#dbe7f2;font-size:12px;font-weight:400;line-height:1.5;white-space:pre-line;box-shadow:0 12px 26px rgba(0,0,0,.35);pointer-events:none}
.battle-tag-tooltip.show{display:block}
.detail-player-head{display:flex;align-items:flex-end;justify-content:space-between;gap:16px;margin:16px 0 8px;padding-bottom:6px;border-bottom:1px solid #2a3a4a}
.detail-player-head .section-title{margin:0;padding:0;border-bottom:none;flex:0 0 auto;font-weight:800}
.detail-player-meta{display:flex;align-items:center;justify-content:flex-end;gap:10px;flex:1;min-width:0}
.detail-player-meta .meta-label{font-size:13px;color:#8ca2b4;font-weight:400;white-space:nowrap}
.detail-player-meta .tag-list{justify-content:flex-end}
.detail-player-meta .battle-tag{font-size:11px;padding:2px 8px}
.detail-player-meta.empty .meta-label{color:#6f8698}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:12px;font-weight:bold}
.b0{background:#0a4d2e;color:#00ff88}.b1{background:#4d1a0a;color:#ff6b4a}.b2{background:#4d3a0a;color:#ffcc4a}.b3{background:#333;color:#888}
.asset-tag-list{display:flex;flex-wrap:wrap;gap:6px}
.asset-tag{display:inline-flex;align-items:center;padding:3px 9px;border-radius:999px;font-size:12px;font-weight:700;line-height:1.2;border:1px solid transparent}
.asset-tag-collab{background:#123326;color:#64e0a7;border-color:#1f6247}
.asset-tag-hidden{background:#3b3420;color:#ffd46a;border-color:#7e6c2a}
.grade{color:#ffcc00;font-weight:bold}
.btn{padding:8px 20px;border:none;border-radius:6px;cursor:pointer;font-size:14px;font-weight:bold;margin-right:8px}
.btn-go{background:#00d4aa;color:#0f1923}.btn-go:hover{background:#00eebb}
.btn-fetch{background:#2a5a8a;color:#fff}.btn-fetch:hover{background:#3a6a9a}
.btn.is-active{background:#00d4aa;color:#0f1923;border-color:#00d4aa;box-shadow:0 0 0 1px rgba(0,212,170,.25) inset}
.theme-toggle{background:#223344;color:#dde7ef;border:1px solid #35506a;padding:7px 14px}
.theme-toggle:hover{background:#2b4055}
.btn:disabled{opacity:.5;cursor:not-allowed}
input[type=number],select{background:#1a2634;border:1px solid #2a3a4a;color:#e0e0e0;padding:6px 10px;border-radius:4px;font-size:14px}
.toolbar{display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.toolbar label{color:#8899aa;font-size:13px}
.scroll-table{max-height:calc(100vh - 220px);overflow-y:auto}
.msg{padding:10px 12px;border-radius:6px;margin-bottom:12px;font-size:13px;display:none;align-items:flex-start;justify-content:space-between;gap:12px}
.msg.show{display:flex}
.msg .msg-text{flex:1;line-height:1.6}
.msg .msg-close{background:none;border:none;color:inherit;opacity:.72;font-size:18px;line-height:1;cursor:pointer;padding:0 2px;flex-shrink:0}
.msg .msg-close:hover{opacity:1}
.msg.info{background:#1a3a5a;color:#4a9eff}.msg.ok{background:#0a3a2a;color:#00d4aa}.msg.err{background:#3a1a1a;color:#ff4a4a}
.summary{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin-bottom:16px}
.summary .card{background:#1a2634;border:1px solid #2a3a4a;border-radius:8px;padding:14px;text-align:center;display:flex;flex-direction:column;justify-content:center;min-height:70px}
.summary .card .val{font-size:22px;color:#00d4aa;font-weight:bold}
.summary .card .lbl{font-size:12px;color:#8899aa;margin-top:4px}
.summary .card .sub{font-size:11px;color:#7a9ab0;margin-bottom:2px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.summary .summary-separator{display:flex;align-items:center;justify-content:center;min-height:70px}
.summary .summary-separator-line{width:1px;height:100%;min-height:76px;background:linear-gradient(180deg,rgba(42,58,74,0),rgba(42,58,74,.9),rgba(42,58,74,0))}
#assetSummaryCards{display:flex;align-items:stretch;gap:16px;flex-wrap:wrap}
#assetSummaryCards .asset-summary-group{display:grid;gap:12px}
#assetSummaryCards .asset-summary-group.left{grid-template-columns:repeat(3,minmax(110px,1fr));flex:0.78 1 0}
#assetSummaryCards .asset-summary-group.right{grid-template-columns:repeat(4,minmax(155px,1fr));flex:1.22 1 0}
#assetSummaryCards .summary-separator{flex:0 0 auto;padding:0 4px}
.summary .card.asset-summary-action{cursor:pointer;transition:transform .16s ease,box-shadow .16s ease,border-color .16s ease}
.summary .card.asset-summary-action:hover{transform:translateY(-1px);border-color:#4f7ea6}
.summary .card.asset-summary-action:active{transform:translateY(0)}
.asset-tab-name{font-weight:700}
.asset-tab-count{font-weight:400;opacity:.78;margin-left:2px}
.team-summary .card{background:#edf3f7;border:1px solid #c9d4df;box-shadow:0 6px 18px rgba(15,25,35,.12)}
.team-summary .card .val{color:#1c4b63}
.team-summary .card .lbl{color:#5b7283}
.team-summary .card .sub{color:#6d8495}
.pager{display:flex;align-items:center;gap:8px;margin-top:12px;color:#8899aa;font-size:13px}
a.room-link{color:#4a9eff;cursor:pointer;text-decoration:none}
a.room-link:hover{text-decoration:underline}

/* modal */
.modal-overlay{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.7);z-index:1000;justify-content:center;align-items:flex-start;padding:40px 20px;overflow-y:auto}
.modal-overlay.show{display:flex}
.modal{background:#162029;border:1px solid #2a3a4a;border-radius:10px;width:100%;max-width:900px;padding:0;overflow:hidden}
.modal-head{background:#1a2634;padding:14px 20px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #2a3a4a}
.modal-head h3{color:#00d4aa;font-size:16px}
.modal-close{background:none;border:none;color:#8899aa;font-size:22px;cursor:pointer;padding:0 4px}
.modal-close:hover{color:#ff6b4a}
.modal-body{padding:20px;max-height:calc(100vh - 140px);overflow-y:auto}
.detail-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;margin-bottom:20px}
.detail-item{background:#1a2634;border:1px solid #2a3a4a;border-radius:6px;padding:10px}
.detail-item .d-label{font-size:11px;color:#8899aa;margin-bottom:4px}
.detail-item .d-val{font-size:15px;color:#e0e0e0;font-weight:bold}
.section-title{color:#00d4aa;font-size:14px;margin:16px 0 8px;padding-bottom:6px;border-bottom:1px solid #2a3a4a}
.item-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:8px}
.item-card{background:#1a2634;border:1px solid #2a3a4a;border-radius:6px;padding:8px 10px;display:flex;align-items:center;gap:8px;font-size:12px}
.item-card img{width:36px;height:36px;border-radius:4px}
.item-card .item-info{flex:1}
.item-card .item-name{color:#e0e0e0;font-weight:500}
.item-card .item-meta{color:#8899aa;font-size:11px}
.player-row{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:8px;padding:10px;background:#1a2634;border:1px solid #2a3a4a;border-radius:6px;margin-bottom:8px}
.player-row.is-self{border-color:#00d4aa}
.detail-player-block{margin-bottom:18px}
.detail-player-block:last-child{margin-bottom:0}
.detail-player-block>.section-title{font-weight:700;font-size:15px;letter-spacing:.01em}
.detail-player-items{margin-top:10px;padding:12px;background:#101a24;border:1px solid #243444;border-radius:10px}
.detail-player-items-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:10px}
.detail-player-items-title{font-size:12px;color:#c7d6e2;font-weight:600;letter-spacing:.02em}
.detail-player-items-count{font-size:12px;color:#7f95a6}
.detail-empty-items{padding:12px;border:1px dashed #304456;border-radius:8px;background:#15212c;color:#7f95a6;font-size:12px;text-align:center}
@media (max-width:980px){.player-row{grid-template-columns:repeat(3,minmax(0,1fr))}}
@media (max-width:700px){.detail-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.player-row{grid-template-columns:repeat(2,minmax(0,1fr))}.item-grid{grid-template-columns:1fr}}

.account-overlay{position:fixed;inset:0;background:rgba(8,12,18,.76);display:none;align-items:center;justify-content:center;padding:24px;z-index:1400}
.account-overlay.show{display:flex}
.account-modal{width:min(720px,100%);background:#162029;border:1px solid #2a3a4a;border-radius:16px;box-shadow:0 24px 70px rgba(0,0,0,.45);overflow:hidden}
.account-modal-head{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;padding:22px 24px 14px;border-bottom:1px solid #243546}
.account-modal-head h3{font-size:20px;color:#eef6fc}
.account-modal-head p{margin-top:6px;color:#8ca2b4;font-size:13px}
.account-modal-body{padding:20px 24px 24px}
.account-list{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}
.account-card{padding:16px;border-radius:12px;background:#1a2634;border:1px solid #2a3a4a;cursor:pointer;transition:border-color .2s,transform .2s,box-shadow .2s}
.account-card:hover{transform:translateY(-1px);border-color:#4d88b8;box-shadow:0 12px 26px rgba(0,0,0,.2)}
.account-card.active{border-color:#00d4aa;box-shadow:0 0 0 1px rgba(0,212,170,.25)}
.account-card-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px}
.account-card-tools{display:flex;align-items:center;gap:8px;flex-shrink:0}
.account-card.manage-mode{cursor:default}
.account-card.manage-mode:hover{transform:none}
.account-delete-btn{padding:6px 10px;border-radius:8px;font-size:12px;margin:0}
.account-card .name{font-size:16px;color:#eef6fc;font-weight:bold}
.account-card .meta{margin-top:8px;font-size:12px;color:#8ca2b4;line-height:1.6}
.account-empty{padding:18px 20px;border-radius:12px;background:#12202d;border:1px dashed #35506a;color:#9ab0c0;font-size:13px;line-height:1.7}
.account-actions{display:flex;flex-direction:column;align-items:flex-start;gap:8px;margin:0 0 16px}
.account-actions .left{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.account-actions .right{font-size:12px;color:#7f95a6;line-height:1.6}
.close-link{background:none;border:none;color:#7f95a6;font-size:14px;cursor:pointer}
.close-link:hover{color:#e7f0f7}
.fetch-hint{font-size:12px;color:#8aa1b3;margin:10px 0 12px;line-height:1.75;padding:12px 14px;border-radius:12px;background:linear-gradient(180deg,#142231 0%,#101c28 100%);border:1px solid #2a3a4a}
.fetch-hint p{margin:0 0 4px}
.fetch-hint p:last-child{margin-bottom:0}
.fetch-progress{display:none;background:#12202d;border:1px solid #2a3a4a;border-radius:10px;padding:12px 14px;margin-bottom:12px}
.fetch-progress.show{display:block}
.fetch-progress .row{margin-bottom:10px}
.fetch-progress .row:last-child{margin-bottom:0}
.fetch-progress .meta{display:flex;justify-content:space-between;gap:12px;font-size:12px;color:#9ab0c0;margin-bottom:6px}
.fetch-progress .bar{height:10px;border-radius:999px;background:#0d1721;border:1px solid #22384b;overflow:hidden}
.fetch-progress .fill{height:100%;width:0;background:linear-gradient(90deg,#1f7fb1 0%,#00d4aa 100%);transition:width .25s ease}
.action-overlay{position:fixed;inset:0;background:rgba(7,12,18,.72);display:none;align-items:center;justify-content:center;z-index:1500;padding:24px}
.action-overlay.show{display:flex}
#actionOverlay{z-index:1600}
.action-card{width:min(500px,100%);background:#162029;border:1px solid #2a3a4a;border-radius:16px;box-shadow:0 24px 70px rgba(0,0,0,.45);padding:22px 24px}
.action-card-head{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;margin-bottom:10px}
.action-card h3{font-size:20px;color:#eef6fc;margin:0}
.action-card-body p{color:#a9bbc8;line-height:1.7;font-size:14px;margin:0 0 10px}
.action-card-body p:last-child{margin-bottom:0}
.action-card-body .dialog-note{font-size:12px;color:#7f95a6}
.dialog-input{width:100%;background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:8px;padding:10px 12px;font-size:14px}
.action-card-actions{display:flex;align-items:center;gap:10px;justify-content:flex-end;flex-wrap:wrap;margin-top:18px}
.danger-keyword{color:#ff6b6b;font-weight:800}
.report-card{width:min(860px,100%)}
.report-list{display:flex;flex-direction:column;gap:10px;max-height:60vh;overflow:auto}
.report-item{border:1px solid #2a3a4a;border-radius:12px;background:#12202d;padding:12px}
.report-item-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:8px}
.report-item-title{font-size:14px;color:#eef6fc;font-weight:700}
.report-item-meta{font-size:12px;color:#7f95a6;line-height:1.7;margin-top:4px}
.report-item-actions{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}
.report-select{display:flex;align-items:flex-start;gap:10px}
.report-select input{margin-top:4px}
.report-batch-bar{display:none;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:10px}
.report-batch-bar.show{display:flex}
.report-status{display:inline-flex;align-items:center;border-radius:999px;padding:3px 8px;font-size:12px;font-weight:700}
.report-status.pending{background:#2b3543;color:#b8c7d5}
.report-status.running{background:#12334a;color:#6bd0ff}
.report-status.canceling{background:#46371d;color:#ffd27a}
.report-status.done{background:#103726;color:#67e49c}
.report-status.failed{background:#462226;color:#ff9a8f}
.report-status.canceled{background:#36313d;color:#c6b8ff}
.report-progress{height:8px;border-radius:999px;background:#0d1721;border:1px solid #22384b;overflow:hidden;margin-top:8px}
.report-progress .fill{height:100%;width:0;background:linear-gradient(90deg,#1f7fb1 0%,#00d4aa 100%)}
.report-empty{border:1px dashed #35506a;border-radius:12px;background:#12202d;color:#9ab0c0;padding:18px;text-align:center;font-size:13px}

/* analysis tab */
.analysis-section{margin-bottom:24px}
.analysis-section h3{color:#00d4aa;font-size:15px;margin-bottom:12px;padding-bottom:6px;border-bottom:1px solid #2a3a4a}
.analysis-summary{grid-template-columns:repeat(5,minmax(0,1fr));gap:14px}
.analysis-summary .card{align-items:center;text-align:center;min-height:92px;padding:18px 16px}
.analysis-summary .card .lbl{font-size:11px;letter-spacing:.02em;margin:10px 0 0;color:#7f96a8}
.analysis-summary .card .val{font-size:28px;line-height:1.14}
.analysis-summary .card .val.profit-pos{color:#57e58d}
.analysis-summary .card .val.profit-neg{color:#ff8c7a}
.analysis-summary .card .sub{margin-top:10px;margin-bottom:0}
@media (max-width:1400px){.analysis-summary{grid-template-columns:repeat(4,minmax(0,1fr))}}
@media (max-width:1120px){.analysis-summary{grid-template-columns:repeat(3,minmax(0,1fr))}}
@media (max-width:820px){.analysis-summary{grid-template-columns:repeat(2,minmax(0,1fr))}}
@media (max-width:520px){.analysis-summary{grid-template-columns:1fr}}
.analysis-section-head{display:flex;align-items:flex-end;justify-content:space-between;gap:12px;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #2a3a4a}
.analysis-section-head h3{margin:0;padding:0;border-bottom:none}
.analysis-note{font-size:11px;color:#7a9ab0;white-space:nowrap}
.trend-sticky{position:sticky;top:0;z-index:40;background:#162029;padding-bottom:12px;margin-bottom:16px}
.trend-summary .card{cursor:pointer;transition:transform .16s ease,border-color .16s ease,box-shadow .16s ease;text-decoration:none;color:inherit}
.trend-summary .card:hover{transform:translateY(-1px);border-color:#4f9fe2;box-shadow:0 10px 24px rgba(0,0,0,.16)}
.trend-meta{font-size:12px;color:#7f95a6;margin:-4px 0 12px;line-height:1.6}
.trend-chart{min-height:280px;scroll-margin-top:560px}
.trend-chart-body{overflow-x:auto;padding-bottom:4px}
.trend-svg{min-width:760px;width:100%;height:260px;display:block}
.trend-axis{stroke:#2d4358;stroke-width:1}
.trend-grid{stroke:#24384a;stroke-width:1;opacity:.72}
.trend-line{fill:none;stroke:#00d4aa;stroke-width:3;stroke-linecap:round;stroke-linejoin:round}
.trend-area{fill:rgba(0,212,170,.10)}
.trend-dot{fill:#00d4aa;stroke:#162029;stroke-width:2}
.trend-axis-label,.trend-value-label{fill:#8fa0ad;font-size:11px}
.trend-dot-label{fill:#dbe7f2;font-size:10px;font-weight:700;paint-order:stroke;stroke:#162029;stroke-width:3px;stroke-linejoin:round}
.trend-chart-stats{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:0 0 12px}
.trend-chart-stat{background:#12202d;border:1px solid #2a3a4a;border-radius:8px;padding:9px 10px;min-height:58px}
.trend-chart-stat .k{font-size:11px;color:#7f95a6;margin-bottom:4px}
.trend-chart-stat .v{font-size:16px;color:#dbe7f2;font-weight:800;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.trend-data-strip{display:flex;gap:8px;overflow-x:auto;padding:8px 0 2px;margin-top:8px}
.trend-data-chip{flex:0 0 auto;min-width:92px;background:#12202d;border:1px solid #2a3a4a;border-radius:8px;padding:7px 9px}
.trend-data-chip .date{font-size:10px;color:#7f95a6;margin-bottom:4px}
.trend-data-chip .value{font-size:13px;color:#dbe7f2;font-weight:800}
.trend-empty{padding:24px;border:1px dashed #35506a;border-radius:12px;background:#12202d;color:#9ab0c0;text-align:center}
@media (max-width:760px){.trend-chart-stats{grid-template-columns:repeat(2,minmax(0,1fr))}}
.stat-bar{display:flex;align-items:center;margin-bottom:6px;font-size:13px}
.stat-bar .stat-label{width:120px;color:#8899aa}
.stat-bar .stat-track{flex:1;background:#1a2634;border-radius:4px;height:20px;overflow:hidden;margin:0 8px}
.stat-bar .stat-fill{height:100%;border-radius:4px;transition:width .3s}
.stat-bar .stat-val{width:80px;text-align:right;font-weight:bold}
.top-list{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:8px}
.top-item{background:#1a2634;border:1px solid #2a3a4a;border-radius:6px;padding:10px;display:flex;justify-content:space-between;align-items:center}
.top-item .rank{color:#ffcc00;font-weight:bold;margin-right:8px}
.heat-table{width:100%;border-collapse:collapse;font-size:12px;margin-top:8px}
.heat-table th,.heat-table td{padding:8px 6px;text-align:center;border:1px solid #1e2e3e}
.heat-table th{background:#1a2c3a;color:#00d4aa;font-size:11px;position:static}
.heat-table th.num,.heat-table td.num{text-align:right;padding-right:12px}
.heat-cell{min-width:60px;font-weight:bold}
.chart-row{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px}
.chart-box{flex:1;min-width:300px;background:#1a2634;border:1px solid #2a3a4a;border-radius:8px;padding:14px}
.chart-box h4{color:#8899aa;font-size:12px;margin-bottom:8px}
.bar-chart{display:flex;flex-direction:column;gap:4px}
.bar-row{display:flex;align-items:center;font-size:12px}
.bar-row .bar-label{width:90px;color:#8899aa;text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.bar-row .bar-track{position:relative;flex:1;background:#0f1923;border-radius:3px;height:16px;overflow:hidden}
.bar-row .bar-fill{height:100%;border-radius:3px}
.bar-row .bar-text{position:absolute;left:6px;right:6px;top:0;bottom:0;display:flex;align-items:center;font-size:10px;font-weight:bold;color:#0f1923;pointer-events:none;white-space:nowrap}
.bar-row .bar-text.outside{color:#dbe7f2;text-shadow:0 1px 2px rgba(0,0,0,.4)}
.player-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:12px}
.player-item{display:flex;align-items:center;padding:8px 10px;background:#1e2e3e;border-radius:6px;cursor:pointer;border:2px solid transparent;transition:all .15s}
.player-item:hover{border-color:#2a5a8a;background:#243444}
.player-item.selected{border-color:#00d4aa;background:#0d2d2d}
.player-item.special{border-style:dashed}
.player-item .p-main{flex:1;min-width:0}
.player-item .p-name{font-size:13px;color:#ddeeff;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.player-item .p-id{font-size:10px;color:#6f8799;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;margin-top:2px}
.player-item .p-count{color:#8899aa;font-size:11px;margin-left:6px}
.player-item .p-check{width:16px;height:16px;border:2px solid #4a6a8a;border-radius:4px;margin-right:8px;display:flex;align-items:center;justify-content:center;font-size:11px;color:#00d4aa;font-weight:bold;flex-shrink:0}
.player-item.selected .p-check{border-color:#00d4aa;background:#00d4aa;color:#0f1923}
.team-hint{font-size:12px;color:#556677;margin-bottom:8px}
.team-btns{display:flex;gap:8px;margin:12px 0;flex-wrap:wrap}
.matrix-wrap{overflow:auto;overscroll-behavior:contain;border-radius:10px;border:1px solid #c9d4df;background:#f7fafc;box-shadow:0 8px 24px rgba(15,25,35,.18)}
.matrix-table{width:max-content;border-collapse:separate;border-spacing:0;font-size:12px;table-layout:fixed;background:#fdfefe;color:#203040}
.matrix-table th,.matrix-table td{padding:6px 8px;text-align:center;border-bottom:1px solid #d8e1e8;border-right:1px solid #d8e1e8;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;background:#fdfefe}
.matrix-table th{background:#dfeaf2;color:#29465f;font-size:11px;font-weight:700}
.matrix-table .col-tag{background:#edf3f7;color:#476173;font-weight:600;text-align:left;padding-left:10px;position:sticky;left:0;z-index:5;border-right:none}
.matrix-table .col-sum{background:#edf3f7;color:#17364c;font-weight:bold;text-align:right;padding-right:12px;position:sticky;left:90px;z-index:5;border-right:3px solid #88a9c2!important}
.matrix-table .col-detail{background:#edf3f7;color:#476173;font-weight:600;text-align:left;padding-left:10px;position:sticky;left:190px;z-index:5;min-width:100px;width:100px;max-width:100px}
.matrix-table thead th{position:sticky;top:0;z-index:6}
.matrix-table thead .col-tag,.matrix-table thead .col-sum,.matrix-table thead .col-detail{z-index:7}
.matrix-table .cell-ok{background:#d4f5e0;color:#1a6b3a}
.matrix-table .cell-fail{background:#f5d4d4;color:#8b2020}
.matrix-table .cell-timeout{background:#f5ecd4;color:#7a6020}
.matrix-table .cell-leave{background:#ddd;color:#555}
.matrix-table .heat-cell{font-weight:600;color:#1a1a2a}
.matrix-table tr.player-head td:first-child{background:#cfe2ee;padding:7px 10px;font-weight:bold;font-size:13px;color:#1c4b63;text-align:left;border-bottom:2px solid #9bb6c8;border-right:3px solid #88a9c2;position:sticky;left:0;z-index:10;width:190px;max-width:190px}
.matrix-table tr.player-head .pb-stats{color:#5d7486;font-size:11px;font-weight:normal;margin-left:16px}
.matrix-table tr.role-row td{font-weight:700}
.matrix-table tr.global-row td.col-tag,.matrix-table tr.global-row td.col-sum,.matrix-table tr.global-row td.col-detail{background:#e4edf3}
.matrix-table th.col-time{min-width:120px}
.matrix-table th.col-time.time-link{cursor:pointer;transition:background .15s,color .15s,box-shadow .15s}
.matrix-table th.col-time.time-link:hover{background:#d0e2ee;color:#16384d;box-shadow:inset 0 -2px 0 #4a7aaa}
.matrix-note{display:flex;align-items:flex-start;gap:10px;margin:14px 0 12px;padding:12px 14px;border-radius:12px;background:linear-gradient(180deg,#edf4f8 0%,#e7f0f5 100%);border:1px solid #c9d8e2;color:#476173;font-size:12px;line-height:1.7}
.matrix-note .note-icon{display:inline-flex;align-items:center;justify-content:center;width:22px;height:22px;border-radius:50%;background:#d8e8f3;color:#1c4b63;font-weight:700;flex-shrink:0}
.matrix-note strong{color:#17364c}
.g{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:2px}
.g0{background:#ff4444}.g1{background:#ff8800}.g2{background:#ffcc00}.g3{background:#88cc44}.g4{background:#44ff88}.g5{background:#00d4aa}

body:not(.light-theme){
 color-scheme:dark;
 background:
  radial-gradient(circle at top left, rgba(0,212,170,.09), transparent 20%),
  radial-gradient(circle at 86% 10%, rgba(74,158,255,.13), transparent 18%),
  radial-gradient(circle at 18% 84%, rgba(122,111,240,.10), transparent 18%),
  linear-gradient(180deg,#0b1520 0%,#0f1923 100%);
}
body:not(.light-theme)::before{
 content:"";
 position:fixed;
 inset:0;
 pointer-events:none;
 z-index:0;
 background:
  linear-gradient(rgba(74,158,255,.05) 1px,transparent 1px),
  linear-gradient(90deg,rgba(74,158,255,.05) 1px,transparent 1px);
 background-size:52px 52px;
 mask-image:radial-gradient(circle at center, black 42%, transparent 92%);
 opacity:.55;
}
body:not(.light-theme) .header,
body:not(.light-theme) .container{position:relative;z-index:1}
body:not(.light-theme) .header{
 position:relative;
  width:min(1480px,calc(100% - 28px));
  margin:18px auto 0;
  padding:14px 18px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:999px;
 background:rgba(12,22,33,.84);
 box-shadow:0 18px 40px rgba(0,0,0,.28);
 backdrop-filter:blur(18px);
}
body:not(.light-theme) .brand{gap:14px}
body:not(.light-theme) .brand h1{
 color:#f3f8fc;
 font-size:22px;
 font-weight:800;
 letter-spacing:-.03em;
}
body:not(.light-theme) .brand-subtitle{
 margin-top:4px;
 color:#7f95a6;
 font-size:11px;
 letter-spacing:.12em;
 text-transform:uppercase;
}
body:not(.light-theme) .container{
 width:min(1480px,calc(100% - 28px));
 margin:18px auto 0;
 padding:0 0 56px;
}
body:not(.light-theme) .msg{
 margin-bottom:16px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:18px;
 box-shadow:0 14px 30px rgba(0,0,0,.2);
}
body:not(.light-theme) .tabs{
 gap:10px;
 margin-bottom:18px;
 padding:10px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:24px;
 background:linear-gradient(180deg,rgba(12,22,33,.96),rgba(15,25,35,.92));
 box-shadow:0 18px 42px rgba(0,0,0,.2);
 overflow:auto;
}
body:not(.light-theme) .tab{
 flex:0 0 auto;
 padding:12px 18px;
 border-radius:999px;
 border:1px solid transparent;
 background:transparent;
 color:#8da1b3;
 font-size:14px;
 font-weight:700;
 transition:all .18s ease;
}
body:not(.light-theme) .tab:hover{color:#b8d8f6;background:rgba(74,158,255,.08)}
body:not(.light-theme) .tab.active{
 background:linear-gradient(180deg,#182737 0%,#142131 100%);
 border-color:rgba(74,158,255,.18);
 color:#eef6fc;
 box-shadow:0 12px 28px rgba(0,0,0,.18);
}
body:not(.light-theme) .panel{
 padding:28px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:32px;
 background:
  radial-gradient(circle at top left, rgba(0,212,170,.06), transparent 20%),
  linear-gradient(180deg, rgba(15,25,35,.98), rgba(13,21,31,.94));
 box-shadow:0 18px 42px rgba(0,0,0,.22);
}
body:not(.light-theme) .toolbar{
 gap:10px 12px;
 margin-bottom:18px;
 padding:16px 18px;
 border:1px solid rgba(41,61,79,.92);
 border-radius:24px;
 background:rgba(12,22,33,.78);
 box-shadow:0 14px 32px rgba(0,0,0,.18);
}
body:not(.light-theme) .toolbar label{
 color:#8da1b3;
 font-size:12px;
 font-weight:700;
}
body:not(.light-theme) .btn{
 margin-right:0;
 padding:10px 16px;
 border-radius:14px;
 border:1px solid transparent;
 font-weight:700;
 box-shadow:0 1px 2px rgba(0,0,0,.08);
}
body:not(.light-theme) .btn-go{
 background:linear-gradient(135deg,#1f7fb1 0%,#00d4aa 100%);
 color:#08131d;
}
body:not(.light-theme) .btn-go:hover{background:linear-gradient(135deg,#2a8fc3 0%,#14e0b9 100%)}
body:not(.light-theme) .btn-fetch{
 background:#173246;
 color:#cfe7fb;
 border-color:#284b63;
}
body:not(.light-theme) .btn-fetch:hover{background:#1d3b52}
body:not(.light-theme) .btn-muted{
 background:#162635;
 color:#d7e6f2;
 border-color:#2a4256;
}
body:not(.light-theme) .btn-muted:hover{background:#1a2c3d}
body:not(.light-theme) .btn-danger{
 background:#3a1a1a;
 color:#ffd7d2;
 border-color:#6c2f33;
}
body:not(.light-theme) .btn-danger:hover{background:#472021}
body:not(.light-theme) .btn-feedback{
 background:#123238;
 color:#a5f2dc;
 border-color:#255f64;
}
body:not(.light-theme) .btn-feedback:hover{background:#184047}
body:not(.light-theme) .theme-toggle{
 background:#162635;
 color:#d7e6f2;
 border-color:#2a4256;
}
body:not(.light-theme) input[type=number],
body:not(.light-theme) input[type=time],
body:not(.light-theme) select{
 min-height:38px;
 padding:8px 12px;
 border-radius:12px;
 background:#101b27!important;
 border:1px solid #294153!important;
 color:#eaf2f8!important;
}
body:not(.light-theme) .scroll-table{
 border:1px solid rgba(41,61,79,.96);
 border-radius:24px;
 background:#0f1923;
 box-shadow:0 14px 34px rgba(0,0,0,.18);
}
body:not(.light-theme) table:not(.matrix-table){background:#0f1923;border-collapse:separate;border-spacing:0}
body:not(.light-theme) table:not(.matrix-table) th{
 background:linear-gradient(180deg,#162736 0%,#132231 100%);
 color:#9fe8d7;
 padding:12px 10px;
 border-bottom:1px solid #294153;
}
body:not(.light-theme) table:not(.matrix-table) td{
 padding:10px;
 border-bottom:1px solid #1b2c3b;
 color:#dce7f1;
 background:#0f1923;
}
body:not(.light-theme) table:not(.matrix-table) tbody tr:nth-child(even) td{background:#12202d}
body:not(.light-theme) table:not(.matrix-table) tbody tr:hover td{background:#162635}
body:not(.light-theme) .summary{
 gap:14px;
 margin-bottom:18px;
}
body:not(.light-theme) .summary:not(.team-summary) .card,
body:not(.light-theme) .analysis-summary .card{
 position:relative;
 overflow:hidden;
 min-height:92px;
 padding:18px 16px 16px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:22px;
 background:linear-gradient(180deg,#132130 0%,#0f1923 100%);
 box-shadow:0 14px 34px rgba(0,0,0,.18);
}
body:not(.light-theme) .summary:not(.team-summary) .card::before,
body:not(.light-theme) .analysis-summary .card::before{
 content:"";
 position:absolute;
 left:18px;
 right:18px;
 top:0;
 height:4px;
 border-radius:999px;
 background:linear-gradient(90deg,#1f7fb1 0%,#00d4aa 100%);
}
body:not(.light-theme) .team-summary .card{
 position:relative;
 overflow:hidden;
 min-height:92px;
 padding:18px 16px 16px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:22px;
 background:linear-gradient(180deg,#132130 0%,#0f1923 100%);
 box-shadow:0 14px 34px rgba(0,0,0,.18);
}
body:not(.light-theme) .team-summary .card::before{
 content:"";
 position:absolute;
 left:18px;
 right:18px;
 top:0;
 height:4px;
 border-radius:999px;
 background:linear-gradient(90deg,#1f7fb1 0%,#00d4aa 100%);
}
body:not(.light-theme) .summary:not(.team-summary) .card .val,
body:not(.light-theme) .analysis-summary .card .val,
body:not(.light-theme) .team-summary .card .val{color:#eef6fc}
body:not(.light-theme) .summary:not(.team-summary) .card .lbl,
body:not(.light-theme) .analysis-summary .card .lbl,
body:not(.light-theme) .team-summary .card .lbl{color:#8da1b3}
body:not(.light-theme) .summary:not(.team-summary) .card .sub,
body:not(.light-theme) .analysis-summary .card .sub,
body:not(.light-theme) .team-summary .card .sub{color:#7b91a3}
body:not(.light-theme) .analysis-section{
 position:relative;
 margin-bottom:20px;
 padding:20px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:26px;
 background:linear-gradient(180deg,#132130 0%,#0f1923 100%);
 box-shadow:0 14px 34px rgba(0,0,0,.18);
}
body:not(.light-theme) .analysis-section::before{
 content:"";
 position:absolute;
 left:20px;
 right:20px;
 top:0;
 height:4px;
 border-radius:999px;
 background:linear-gradient(90deg,#1f7fb1 0%,#4a9eff 45%,#00d4aa 100%);
}
body:not(.light-theme) .analysis-section-head{padding-bottom:10px;border-bottom:1px solid #24384a}
body:not(.light-theme) .analysis-note{color:#7b91a3}
body:not(.light-theme) .chart-box,
body:not(.light-theme) .top-item,
body:not(.light-theme) .detail-item,
body:not(.light-theme) .item-card,
body:not(.light-theme) .player-row,
body:not(.light-theme) .detail-player-items{
 border-radius:18px;
 border:1px solid rgba(41,61,79,.96);
 background:linear-gradient(180deg,#132130 0%,#0f1923 100%);
 box-shadow:0 12px 28px rgba(0,0,0,.16);
}
body:not(.light-theme) .chart-box h4{color:#8da1b3}
body:not(.light-theme) .bar-row .bar-label{color:#8da1b3}
body:not(.light-theme) .bar-row .bar-track,
body:not(.light-theme) .stat-bar .stat-track{
 background:#0b1520;
 border:1px solid #24384a;
 border-radius:999px;
}
body:not(.light-theme) .bar-row .bar-fill,
body:not(.light-theme) .stat-bar .stat-fill{border-radius:999px}
body:not(.light-theme) .player-grid{gap:10px;margin-bottom:16px}
body:not(.light-theme) .player-item{
 padding:12px 14px;
 border-width:1px;
 border-radius:16px;
 background:linear-gradient(180deg,#152535 0%,#11202e 100%);
 border-color:#294153;
 box-shadow:0 10px 24px rgba(0,0,0,.14);
}
body:not(.light-theme) .player-item:hover{
 background:#1a2c3d;
 border-color:#417194;
 transform:translateY(-1px);
}
body:not(.light-theme) .player-item.selected{
 background:linear-gradient(180deg,#133343 0%,#102d30 100%);
 border-color:#00d4aa;
 box-shadow:0 0 0 3px rgba(0,212,170,.10);
}
body:not(.light-theme) .player-item .p-check{
 width:18px;
 height:18px;
 border-radius:6px;
 border-width:1px;
}
body:not(.light-theme) .team-hint{margin-bottom:12px;color:#8da1b3;line-height:1.7}
body:not(.light-theme) .team-btns{gap:10px;margin:16px 0 18px}
body:not(.light-theme) #teamResultContainer:not(:empty){
 padding:18px;
 border:1px solid rgba(41,61,79,.96);
 border-radius:24px;
 background:rgba(12,22,33,.70);
 box-shadow:0 14px 34px rgba(0,0,0,.18);
}
body:not(.light-theme) .matrix-note{
 border-radius:18px;
 background:linear-gradient(180deg,#132130 0%,#0f1923 100%);
 border:1px solid rgba(41,61,79,.96);
 color:#8da1b3;
 box-shadow:0 10px 24px rgba(0,0,0,.16);
}
body:not(.light-theme) .matrix-note .note-icon{
 background:#173246;
 color:#9fe8d7;
}
body:not(.light-theme) .matrix-note strong{color:#eef6fc}
body:not(.light-theme) .fetch-hint,
body:not(.light-theme) .fetch-progress,
body:not(.light-theme) .account-modal,
body:not(.light-theme) .action-card,
body:not(.light-theme) .modal{border-radius:24px}
body:not(.light-theme) .fetch-hint,
body:not(.light-theme) .fetch-progress{
 border:1px solid rgba(41,61,79,.96);
 box-shadow:0 12px 28px rgba(0,0,0,.16);
}
body:not(.light-theme) #fetchLog{
 border:1px solid #294153;
 border-radius:20px!important;
 box-shadow:0 12px 28px rgba(0,0,0,.16);
}
body:not(.light-theme) .account-card,
body:not(.light-theme) .report-item{
 border-radius:18px;
 background:linear-gradient(180deg,#132130 0%,#0f1923 100%);
}
@media (max-width:1100px){
 body:not(.light-theme) .header{
  position:relative;
  top:auto;
  border-radius:28px;
 }
 body:not(.light-theme) .panel{padding:22px}
}
@media (max-width:760px){
 body:not(.light-theme) .container,
 body:not(.light-theme) .header{width:calc(100% - 20px)}
 body:not(.light-theme) .header{
  flex-direction:column;
  align-items:flex-start;
  gap:12px;
  padding:14px;
 }
 body:not(.light-theme) .header-right{width:100%;justify-content:flex-start}
 body:not(.light-theme) .tabs{padding:8px;border-radius:20px}
 body:not(.light-theme) .tab{padding:10px 14px}
 body:not(.light-theme) .panel{padding:16px;border-radius:24px}
 body:not(.light-theme) .toolbar{
  padding:14px;
  border-radius:20px;
 }
 body:not(.light-theme) .summary{grid-template-columns:repeat(auto-fit,minmax(120px,1fr))}
 body:not(.light-theme) .scroll-table{border-radius:18px}
 body:not(.light-theme) #teamResultContainer:not(:empty){padding:12px;border-radius:20px}
}

body.light-theme{color-scheme:light;background:linear-gradient(180deg,#eef3f7 0%,#e8f0f5 100%);color:#243645}
body.light-theme .header{background:#ffffff;border-bottom-color:#d4dde5;box-shadow:0 2px 12px rgba(15,25,35,.06)}
body.light-theme .brand h1{color:#1c4b63}
body.light-theme .brand-logo{box-shadow:0 8px 22px rgba(42,71,94,.14);background:#ffffff}
body.light-theme .brand-subtitle{color:#7a8f9f}
body.light-theme .header .stats{color:#607585}
body.light-theme .tab{background:#e3ebf1;border-color:#cad6e0;color:#607585}
body.light-theme .tab.active{background:#ffffff;border-color:#6e97b2;color:#1c4b63;border-bottom-color:#ffffff;box-shadow:0 -1px 0 #ffffff inset}
body.light-theme .panel{background:rgba(255,255,255,.72);border-color:#d4dde5;box-shadow:0 12px 30px rgba(15,25,35,.08);backdrop-filter:blur(8px)}
body.light-theme .scroll-table{background:#ffffff;border:1px solid #dbe5ec;border-radius:12px;overflow:auto;box-shadow:0 8px 22px rgba(15,25,35,.05)}
body.light-theme table:not(.matrix-table){background:#ffffff}
body.light-theme table:not(.matrix-table) th{background:linear-gradient(180deg,#dbe8f0 0%,#edf4f8 100%);color:#23485d;border-bottom:2px solid #c9d8e2}
body.light-theme table:not(.matrix-table) td{border-bottom-color:#e4ebf0;color:#243645}
body.light-theme table:not(.matrix-table) tbody tr:nth-child(even) td{background:#f8fbfd}
body.light-theme table:not(.matrix-table) tbody tr:hover td{background:#edf5fa}
body.light-theme .btn{box-shadow:0 1px 2px rgba(15,25,35,.08)}
body.light-theme .btn-go{background:#1c8fb1;color:#ffffff}
body.light-theme .btn-go:hover{background:#167a98}
body.light-theme .btn-fetch{background:#6e89a0;color:#ffffff}
body.light-theme .btn-fetch:hover{background:#5f7a91}
body.light-theme input[type=number],
body.light-theme input[type=time],
body.light-theme select{background:#ffffff!important;border:1px solid #ccd8e2!important;color:#243645!important}
body.light-theme .toolbar label{color:#607585}
body.light-theme .msg.info{background:#e4f0fa;color:#2e74a6}
body.light-theme .msg.ok{background:#e3f6ec;color:#188d59}
body.light-theme .msg.err{background:#fde9e8;color:#c05353}
body.light-theme .summary:not(.team-summary) .card{background:linear-gradient(180deg,#ffffff 0%,#fbfdff 100%);border-color:#d7e1e8;box-shadow:0 10px 24px rgba(15,25,35,.06)}
body.light-theme .summary:not(.team-summary) .card .val{color:#1c4b63}
body.light-theme .summary:not(.team-summary) .card .lbl{color:#607585}
body.light-theme .summary:not(.team-summary) .card .sub{color:#6d8495}
body.light-theme .summary .summary-separator-line{background:linear-gradient(180deg,rgba(201,214,226,0),rgba(201,214,226,.95),rgba(201,214,226,0))}
@media (max-width: 1400px){
 #assetSummaryCards .asset-summary-group.left{grid-template-columns:repeat(3,minmax(100px,1fr));flex:0.72 1 0}
 #assetSummaryCards .asset-summary-group.right{grid-template-columns:repeat(4,minmax(135px,1fr));flex:1.28 1 0}
}
@media (max-width: 1100px){
 #assetSummaryCards{display:grid;grid-template-columns:1fr;gap:12px}
 #assetSummaryCards .asset-summary-group.left,
 #assetSummaryCards .asset-summary-group.right{grid-template-columns:repeat(auto-fit,minmax(140px,1fr))}
 #assetSummaryCards .summary-separator{display:none}
}
body.light-theme .analysis-summary .card{background:linear-gradient(180deg,#ffffff 0%,#f9fcff 100%);border-color:#d7e1e8}
body.light-theme .analysis-summary .card .lbl{color:#6a8091}
body.light-theme .analysis-summary .card .val{color:#204f67}
body.light-theme .analysis-summary .card .val.profit-pos{color:#1d9960}
body.light-theme .analysis-summary .card .val.profit-neg{color:#c45145}
body.light-theme .modal{background:#ffffff;border-color:#d4dde5}
body.light-theme .modal-head{background:#f2f7fa;border-bottom-color:#d4dde5}
body.light-theme .modal-head h3{color:#1c4b63}
body.light-theme .modal-close{color:#607585}
body.light-theme .detail-item,
body.light-theme .item-card,
body.light-theme .player-row,
body.light-theme .detail-player-items,
body.light-theme .chart-box,
body.light-theme .top-item{background:#ffffff;border-color:#d4dde5;box-shadow:0 8px 20px rgba(15,25,35,.05)}
body.light-theme .analysis-section{background:linear-gradient(180deg,#ffffff 0%,#fbfdff 100%);border:1px solid #d7e1e8;border-radius:14px;padding:16px 16px 14px;box-shadow:0 10px 24px rgba(15,25,35,.05)}
body.light-theme .trend-sticky{background:#ffffff}
body.light-theme .trend-grid{stroke:#e4ebf0}
body.light-theme .trend-axis{stroke:#cbd8e2}
body.light-theme .trend-axis-label,
body.light-theme .trend-value-label{fill:#64748b}
body.light-theme .trend-dot-label{fill:#203040;stroke:#ffffff}
body.light-theme .trend-dot{stroke:#ffffff}
body.light-theme .trend-chart-stat,
body.light-theme .trend-data-chip{background:#f8fbfd;border-color:#d5e1ea}
body.light-theme .trend-chart-stat .k,
body.light-theme .trend-data-chip .date{color:#64748b}
body.light-theme .trend-chart-stat .v,
body.light-theme .trend-data-chip .value{color:#203040}
body.light-theme .trend-empty{background:#f8fbfd;border-color:#d5e1ea;color:#64748b}
body.light-theme .detail-item .d-label,
body.light-theme .item-card .item-meta,
body.light-theme .analysis-section h3,
body.light-theme .chart-box h4,
body.light-theme .bar-row .bar-label,
body.light-theme .stat-bar .stat-label,
body.light-theme .top-item .rank,
body.light-theme .player-item .p-count,
body.light-theme .pager{color:#607585}
body.light-theme .detail-item .d-val,
body.light-theme .item-card .item-name{color:#243645}
body.light-theme .section-title{color:#1c4b63;border-bottom-color:#d4dde5}
body.light-theme .detail-player-head{border-bottom-color:#d4dde5}
body.light-theme .detail-player-meta .meta-label{color:#6a8091}
body.light-theme .detail-player-meta.empty .meta-label{color:#7a8f9f}
body.light-theme .detail-player-items-title{color:#375467}
body.light-theme .detail-player-items-count,
body.light-theme .detail-empty-items{color:#607585}
body.light-theme .detail-empty-items{background:#f8fbfd;border-color:#d7e1e8}
body.light-theme .analysis-section h3{color:#1c4b63;border-bottom-color:#dde7ee;margin-bottom:14px}
body.light-theme .analysis-section-head{border-bottom-color:#dde7ee}
body.light-theme .analysis-note{color:#6c8394}
body.light-theme .bar-row .bar-track,
body.light-theme .stat-bar .stat-track{background:#ecf2f6;border:1px solid #d7e1e8}
body.light-theme .bar-row .bar-text.outside{color:#23485d;text-shadow:none}
body.light-theme .chart-box{background:linear-gradient(180deg,#fcfeff 0%,#f7fbfd 100%);border-color:#dde6ed}
body.light-theme .heat-table{background:#ffffff;border:1px solid #dbe5ec;border-collapse:separate;border-spacing:0;border-radius:10px;overflow:hidden;box-shadow:0 4px 14px rgba(15,25,35,.04)}
body.light-theme .heat-table th,
body.light-theme .heat-table td{border-color:#e1e9ef;border-right:1px solid #e1e9ef}
body.light-theme .heat-table th{background:linear-gradient(180deg,#dbe8f0 0%,#eef5f9 100%);color:#23485d}
body.light-theme .heat-table th:last-child,
body.light-theme .heat-table td:last-child{border-right:none}
body.light-theme .heat-table tbody tr:last-child td{border-bottom:none}
body.light-theme .badge{border:1px solid transparent;box-shadow:none}
body.light-theme .b0{background:#e6f7ee;color:#1e8e5a;border-color:#bfe5cf}
body.light-theme .b1{background:#fdebe7;color:#b85a4a;border-color:#f1c8bf}
body.light-theme .b2{background:#fff5de;color:#a97514;border-color:#edd9a5}
body.light-theme .b3{background:#eef2f5;color:#748391;border-color:#d7e0e7}
body.light-theme .asset-tag-collab{background:#e9f8f0;color:#1f7a53;border-color:#c4e8d3}
body.light-theme .asset-tag-hidden{background:#fff6df;color:#9b6d16;border-color:#efd89e}
body.light-theme .battle-tag.pos{background:#e9f8f0;color:#1f7a53;border-color:#c4e8d3}
body.light-theme .battle-tag.neg{background:#feeeea;color:#b85a4a;border-color:#f0c9c0}
body.light-theme .battle-tag.mid{background:#fff6df;color:#9b6d16;border-color:#efd89e}
body.light-theme .battle-tag-tooltip{background:#ffffff;border-color:#d7e1e8;color:#243645;box-shadow:0 14px 28px rgba(42,71,94,.16)}
body.light-theme .player-item{background:#eef4f8}
body.light-theme .player-item:hover{background:#e2edf4;border-color:#8eabc0}
body.light-theme .player-item.selected{background:#d8edf0;border-color:#1c8fb1}
body.light-theme .player-item.special{border-style:dashed}
body.light-theme .player-item .p-name{color:#243645}
body.light-theme .player-item .p-id{color:#6d8495}
body.light-theme .player-item .p-check{border-color:#8eabc0;color:#1c8fb1}
body.light-theme .player-item.selected .p-check{border-color:#1c8fb1;background:#1c8fb1;color:#ffffff}
body.light-theme a.room-link{color:#1c76b3}
body.light-theme .account-pill{background:#f6fbff;border-color:#d3e0ea;color:#244258}
body.light-theme .btn-muted{background:#edf4f8;color:#1c4b63;border-color:#ccd8e2}
body.light-theme .btn-muted:hover{background:#e3edf4}
body.light-theme .btn-feedback{background:#e7f8f3;color:#15705e;border-color:#b9ddd4}
body.light-theme .btn-feedback:hover{background:#d8f1ea;color:#0f5b4c}
body.light-theme .btn-danger{background:#fff2f0;color:#9a4438;border-color:#efc2ba}
body.light-theme .btn-danger:hover{background:#ffe7e3}
body.light-theme .account-overlay{background:rgba(222,233,241,.72)}
body.light-theme .account-modal{background:#ffffff;border-color:#d5e0e8;box-shadow:0 24px 70px rgba(42,71,94,.18)}
body.light-theme .account-modal-head{border-bottom-color:#e3ebf1}
body.light-theme .account-modal-head h3{color:#18384c}
body.light-theme .account-modal-head p{color:#688092}
body.light-theme .account-card{background:#f9fcfe;border-color:#d8e4ec}
body.light-theme .account-card:hover{border-color:#5d93ba;box-shadow:0 14px 28px rgba(42,71,94,.12)}
body.light-theme .account-card.active{border-color:#1c8fb1;box-shadow:0 0 0 1px rgba(28,143,177,.2)}
body.light-theme .account-card .name{color:#18384c}
body.light-theme .account-card .meta{color:#6b8293}
body.light-theme .account-empty{background:#f4f9fc;border-color:#d6e4ee;color:#5f788a}
body.light-theme .account-actions .right{color:#6f8899}
body.light-theme .close-link{color:#6f8899}
body.light-theme .close-link:hover{color:#18384c}
body.light-theme .fetch-hint{background:linear-gradient(180deg,#ffffff 0%,#f4f9fc 100%);border-color:#d7e1e8;color:#5d7486;box-shadow:0 8px 20px rgba(42,71,94,.08)}
body.light-theme .fetch-progress{background:#f6fbff;border-color:#d5e0e8}
body.light-theme .fetch-progress .meta{color:#5f788a}
body.light-theme .fetch-progress .bar{background:#eaf1f5;border-color:#d4e0e8}
body.light-theme .fetch-progress .fill{background:linear-gradient(90deg,#4d89c4 0%,#1cb58e 100%)}
body.light-theme .matrix-table th.col-time.time-link:hover{background:#d9eaf4;color:#173e55;box-shadow:inset 0 -2px 0 #6a92af}
body.light-theme .matrix-note{background:linear-gradient(180deg,#ffffff 0%,#f4f9fc 100%);border-color:#d7e1e8;color:#5d7486;box-shadow:0 8px 20px rgba(42,71,94,.08)}
body.light-theme .matrix-note .note-icon{background:#e5f0f7;color:#1c4b63}
body.light-theme .matrix-note strong{color:#1c4b63}
body.light-theme .action-overlay{background:rgba(222,233,241,.72)}
body.light-theme .action-card{background:#ffffff;border-color:#d5e0e8;box-shadow:0 24px 70px rgba(42,71,94,.18)}
body.light-theme .action-card h3{color:#18384c}
body.light-theme .action-card-body p{color:#688092}
body.light-theme .action-card-body .dialog-note{color:#6f8899}
body.light-theme .dialog-input{background:#f6fafc;color:#18384c;border-color:#d5e0e8}
body.light-theme .danger-keyword{color:#cc433e}
body.light-theme .report-item{background:#f7fbfd;border-color:#d8e4ec}
body.light-theme .report-item-title{color:#18384c}
body.light-theme .report-item-meta{color:#6b8293}
body.light-theme .report-progress{background:#eaf1f5;border-color:#d4e0e8}
body.light-theme .report-progress .fill{background:linear-gradient(90deg,#4d89c4 0%,#1cb58e 100%)}
body.light-theme .report-empty{background:#f4f9fc;border-color:#d6e4ee;color:#5f788a}
body.light-theme .report-select input{accent-color:#2d8fbd}
body.light-theme .theme-toggle{background:#edf4f8;color:#1c4b63;border:1px solid #ccd8e2}
body.light-theme .theme-toggle:hover{background:#e3edf4}

/* light UI refresh inspired by the website download page */
body.light-theme{
 color-scheme:light;
 color:#181d26;
 background:
  radial-gradient(circle at top left, rgba(85,182,255,.20), transparent 22%),
  radial-gradient(circle at 88% 10%, rgba(122,111,240,.12), transparent 18%),
  radial-gradient(circle at 20% 86%, rgba(40,176,111,.11), transparent 18%),
  linear-gradient(180deg,#f9fbfe 0%,#f3f7fb 100%);
}
body.light-theme::before{
 content:"";
 position:fixed;
 inset:0;
 pointer-events:none;
 z-index:0;
 background:
  linear-gradient(rgba(27,97,201,.035) 1px,transparent 1px),
  linear-gradient(90deg,rgba(27,97,201,.035) 1px,transparent 1px);
 background-size:52px 52px;
 mask-image:radial-gradient(circle at center, black 42%, transparent 92%);
 opacity:.72;
}
body.light-theme .header,
body.light-theme .container{position:relative;z-index:1}
body.light-theme .header{
 position:relative;
  width:min(1480px,calc(100% - 28px));
  margin:18px auto 0;
  padding:14px 18px;
 border:1px solid rgba(224,226,230,.96);
 border-radius:999px;
 background:rgba(255,255,255,.84);
 box-shadow:0 16px 38px rgba(24,29,38,.06);
 backdrop-filter:blur(18px);
}
body.light-theme .brand{gap:14px}
body.light-theme .brand h1{
 color:#181d26;
 font-size:22px;
 font-weight:800;
 letter-spacing:-.03em;
}
body.light-theme .brand-subtitle{
 margin-top:4px;
 color:#64748b;
 font-size:11px;
 letter-spacing:.12em;
 text-transform:uppercase;
}
body.light-theme .container{
 width:min(1480px,calc(100% - 28px));
 margin:18px auto 0;
 padding:0 0 56px;
}
body.light-theme .msg{
 margin-bottom:16px;
 border:1px solid rgba(212,218,226,.92);
 border-radius:18px;
 box-shadow:0 12px 28px rgba(24,29,38,.05);
}
body.light-theme .tabs{
 gap:10px;
 margin-bottom:18px;
 padding:10px;
 border:1px solid rgba(224,226,230,.96);
 border-radius:24px;
 background:linear-gradient(180deg,rgba(255,255,255,.98),rgba(246,250,255,.94));
 box-shadow:0 18px 42px rgba(27,56,97,.06);
 overflow:auto;
}
body.light-theme .tab{
 flex:0 0 auto;
 padding:12px 18px;
 border-radius:999px;
 border:1px solid transparent;
 background:transparent;
 color:#64748b;
 font-size:14px;
 font-weight:700;
 transition:all .18s ease;
}
body.light-theme .tab:hover{color:#1b61c9;background:rgba(27,97,201,.06)}
body.light-theme .tab.active{
 background:#ffffff;
 border-color:rgba(27,97,201,.14);
 color:#181d26;
 border-bottom-color:rgba(27,97,201,.14);
 box-shadow:0 12px 28px rgba(27,56,97,.08);
}
body.light-theme .panel{
 padding:28px;
 border:1px solid rgba(224,226,230,.96);
 border-radius:32px;
 background:
  radial-gradient(circle at top left, rgba(85,182,255,.08), transparent 20%),
  linear-gradient(180deg, rgba(255,255,255,.98), rgba(246,250,255,.94));
 box-shadow:0 18px 42px rgba(27,56,97,.06);
}
body.light-theme .toolbar{
 gap:10px 12px;
 margin-bottom:18px;
 padding:16px 18px;
 border:1px solid rgba(219,227,236,.92);
 border-radius:24px;
 background:rgba(255,255,255,.84);
 box-shadow:0 14px 32px rgba(27,56,97,.05);
}
body.light-theme .toolbar label{
 color:#64748b;
 font-size:12px;
 font-weight:700;
}
body.light-theme .btn{
 margin-right:0;
 padding:10px 16px;
 border-radius:14px;
 border:1px solid transparent;
 font-weight:700;
 box-shadow:0 1px 2px rgba(15,25,35,.04);
}
body.light-theme .btn-go{
 background:linear-gradient(135deg,#1b61c9 0%,#254fad 100%);
 color:#ffffff;
}
body.light-theme .btn-go:hover{background:linear-gradient(135deg,#215fbc 0%,#1f478f 100%)}
body.light-theme .btn-fetch{
 background:#edf4ff;
 color:#1b61c9;
 border-color:#d4e2f7;
}
body.light-theme .btn-fetch:hover{background:#e4efff}
body.light-theme .btn-muted{
 background:#f6f9fc;
 color:#334155;
 border-color:#d8e2eb;
}
body.light-theme .btn-muted:hover{background:#edf4f8}
body.light-theme .btn-danger{
 background:#fff2f0;
 color:#9a4438;
 border-color:#efc2ba;
}
body.light-theme .btn-danger:hover{background:#ffe7e3}
body.light-theme .btn-feedback{
 background:#e9f8f1;
 color:#15705e;
 border-color:#bfe4d5;
}
body.light-theme .btn-feedback:hover{background:#dcf3e8}
body.light-theme .theme-toggle{
 background:#f7f9fc;
 color:#334155;
 border-color:#d8e2eb;
}
body.light-theme input[type=number],
body.light-theme input[type=time],
body.light-theme select{
 min-height:38px;
 padding:8px 12px;
 border-radius:12px;
 background:#ffffff!important;
 border:1px solid #d8e2eb!important;
 color:#181d26!important;
 box-shadow:inset 0 1px 0 rgba(255,255,255,.9);
}
body.light-theme .scroll-table{
 border:1px solid rgba(219,227,236,.96);
 border-radius:24px;
 background:#ffffff;
 box-shadow:0 14px 34px rgba(27,56,97,.05);
}
body.light-theme table:not(.matrix-table){
 background:#ffffff;
 border-collapse:separate;
 border-spacing:0;
}
body.light-theme table:not(.matrix-table) th{
 background:linear-gradient(180deg,#edf4fb 0%,#e6eef6 100%);
 color:#23485d;
 padding:12px 10px;
 border-bottom:1px solid #d9e4ec;
}
body.light-theme table:not(.matrix-table) td{
 padding:10px;
 border-bottom:1px solid #e9eef3;
 color:#243645;
 background:#ffffff;
}
body.light-theme table:not(.matrix-table) tbody tr:nth-child(even) td{background:#fbfdff}
body.light-theme table:not(.matrix-table) tbody tr:hover td{background:#f4f9fd}
body.light-theme .summary{
 gap:14px;
 margin-bottom:18px;
}
body.light-theme .summary:not(.team-summary) .card,
body.light-theme .analysis-summary .card{
 position:relative;
 overflow:hidden;
 min-height:92px;
 padding:18px 16px 16px;
 border:1px solid rgba(219,227,236,.96);
 border-radius:22px;
 background:linear-gradient(180deg,#ffffff 0%,#fbfdff 100%);
 box-shadow:0 14px 34px rgba(27,56,97,.05);
}
body.light-theme .summary:not(.team-summary) .card::before,
body.light-theme .analysis-summary .card::before{
 content:"";
 position:absolute;
 left:18px;
 right:18px;
 top:0;
 height:4px;
 border-radius:999px;
 background:linear-gradient(90deg,#1b61c9 0%,#28b06f 100%);
}
body.light-theme .summary:not(.team-summary) .card .val,
body.light-theme .analysis-summary .card .val{color:#163c57}
body.light-theme .summary:not(.team-summary) .card .lbl,
body.light-theme .analysis-summary .card .lbl{color:#64748b}
body.light-theme .summary:not(.team-summary) .card .sub,
body.light-theme .analysis-summary .card .sub{color:#7b8c9d}
body.light-theme .analysis-section{
 position:relative;
 margin-bottom:20px;
 padding:20px;
 border:1px solid rgba(219,227,236,.96);
 border-radius:26px;
 background:linear-gradient(180deg,#ffffff 0%,#fbfdff 100%);
 box-shadow:0 14px 34px rgba(27,56,97,.05);
}
body.light-theme .analysis-section::before{
 content:"";
 position:absolute;
 left:20px;
 right:20px;
 top:0;
 height:4px;
 border-radius:999px;
 background:linear-gradient(90deg,#1b61c9 0%,#55b6ff 45%,#28b06f 100%);
}
body.light-theme .analysis-section-head{padding-bottom:10px;border-bottom:1px solid #e5ecf1}
body.light-theme .analysis-note{color:#7b8c9d}
body.light-theme .chart-box,
body.light-theme .top-item,
body.light-theme .detail-item,
body.light-theme .item-card,
body.light-theme .player-row,
body.light-theme .detail-player-items{
 border-radius:18px;
 border:1px solid rgba(219,227,236,.96);
 background:linear-gradient(180deg,#ffffff 0%,#fbfdff 100%);
 box-shadow:0 12px 28px rgba(27,56,97,.04);
}
body.light-theme .chart-box h4{color:#64748b}
body.light-theme .bar-row .bar-label{color:#64748b}
body.light-theme .bar-row .bar-track,
body.light-theme .stat-bar .stat-track{
 background:#eef4f8;
 border:1px solid #dde7ee;
 border-radius:999px;
}
body.light-theme .bar-row .bar-fill,
body.light-theme .stat-bar .stat-fill{border-radius:999px}
body.light-theme .player-grid{gap:10px;margin-bottom:16px}
body.light-theme .player-item{
 padding:12px 14px;
 border-width:1px;
 border-radius:16px;
 background:linear-gradient(180deg,#ffffff 0%,#f7fbff 100%);
 border-color:#d9e4ec;
 box-shadow:0 10px 24px rgba(27,56,97,.04);
}
body.light-theme .player-item:hover{
 background:#f2f8fc;
 border-color:#9dc3dd;
 transform:translateY(-1px);
}
body.light-theme .player-item.selected{
 background:linear-gradient(180deg,#ecf6ff 0%,#e8f6f1 100%);
 border-color:#1b61c9;
 box-shadow:0 0 0 3px rgba(27,97,201,.08);
}
body.light-theme .player-item .p-check{
 width:18px;
 height:18px;
 border-radius:6px;
 border-width:1px;
}
body.light-theme .team-hint{
 margin-bottom:12px;
 color:#64748b;
 line-height:1.7;
}
body.light-theme .team-btns{
 gap:10px;
 margin:16px 0 18px;
}
body.light-theme #teamResultContainer:not(:empty){
 padding:18px;
 border:1px solid rgba(219,227,236,.96);
 border-radius:24px;
 background:rgba(255,255,255,.72);
 box-shadow:0 14px 34px rgba(27,56,97,.05);
}
body.light-theme .matrix-wrap{
 border-radius:20px;
 border:1px solid #dbe5ec;
 background:#ffffff;
 box-shadow:0 14px 34px rgba(27,56,97,.06);
}
body.light-theme .matrix-table{
 background:#ffffff;
 color:#243645;
}
body.light-theme .matrix-table th,
body.light-theme .matrix-table td{border-color:#e4ebf0}
body.light-theme .matrix-table th{
 background:linear-gradient(180deg,#edf4fb 0%,#e6eef6 100%);
 color:#23485d;
}
body.light-theme .matrix-table .col-tag,
body.light-theme .matrix-table .col-detail{
 background:#eef4f8;
 color:#476173;
}
body.light-theme .matrix-table .col-sum{
 background:#ffffff;
 color:#17364c;
 border-right:3px solid #9bb8cc!important;
}
body.light-theme .matrix-table tr.player-head td:first-child{
 background:#d8e8f3;
 border-right:3px solid #9bb8cc;
}
body.light-theme .matrix-note{
 border-radius:18px;
 background:linear-gradient(180deg,#ffffff 0%,#f4f9fc 100%);
 box-shadow:0 10px 24px rgba(27,56,97,.05);
}
body.light-theme .fetch-hint,
body.light-theme .fetch-progress,
body.light-theme .account-modal,
body.light-theme .action-card,
body.light-theme .modal{
 border-radius:24px;
}
body.light-theme .fetch-hint,
body.light-theme .fetch-progress{
 border:1px solid rgba(219,227,236,.96);
 box-shadow:0 12px 28px rgba(27,56,97,.05);
}
body.light-theme #fetchLog{
 background:#ffffff!important;
 color:#5f7486!important;
 border:1px solid #d8e2eb;
 border-radius:20px!important;
 box-shadow:0 12px 28px rgba(27,56,97,.05);
}
body.light-theme .account-pill{
 background:#f7fbff;
 border-color:#d7e4ed;
 color:#244258;
}
body.light-theme .account-card,
body.light-theme .report-item{
 border-radius:18px;
 background:linear-gradient(180deg,#ffffff 0%,#f8fbfe 100%);
}
@media (max-width:1100px){
 body.light-theme .header{
  position:relative;
  top:auto;
  border-radius:28px;
 }
 body.light-theme .panel{padding:22px}
}
@media (max-width:760px){
 body.light-theme .container,
 body.light-theme .header{width:calc(100% - 20px)}
 body.light-theme .header{
  flex-direction:column;
  align-items:flex-start;
  gap:12px;
  padding:14px;
 }
 body.light-theme .header-right{width:100%;justify-content:flex-start}
 body.light-theme .tabs{padding:8px;border-radius:20px}
 body.light-theme .tab{padding:10px 14px}
 body.light-theme .panel{padding:16px;border-radius:24px}
 body.light-theme .toolbar{
  padding:14px;
  border-radius:20px;
 }
 body.light-theme .summary{grid-template-columns:repeat(auto-fit,minmax(120px,1fr))}
 body.light-theme .scroll-table{border-radius:18px}
 body.light-theme #teamResultContainer:not(:empty){padding:12px;border-radius:20px}
}
</style>
</head>
<body class="light-theme">
<div class="header">
 <div class="brand">
  <img class="brand-logo" src="/logo.png" alt="Delta Force Data Center logo">
  <div>
   <h1>Delta Force Data Center</h1>
   <div class="brand-subtitle">三角洲数据分析中心</div>
  </div>
 </div>
 <div class="header-right">
  <button class="btn btn-muted" id="btnSwitchAccount" onclick="openPlayerMenu()">玩家 · 未选择</button>
  <a class="btn btn-feedback" id="btnFeedback" href="https://docs.qq.com/form/page/DS3pHdEpVZ2Rqd0JC" target="_blank" rel="noopener noreferrer">反馈中心</a>
  <button class="btn theme-toggle" id="themeToggle" onclick="toggleTheme()">切换暗色</button>
 </div>
</div>
<div class="container">
 <div class="msg" id="msg"></div>
 <div class="tabs">
   <div class="tab active" data-tab="records">战绩列表</div>
   <div class="tab" data-tab="items">带出物品</div>
   <div class="tab" data-tab="assets">账号资产</div>
    <div class="tab" data-tab="analysis">数据分析</div>
    <div class="tab" data-tab="trends">数据趋势</div>
    <div class="tab" data-tab="team">组队分析</div>
    <div class="tab" data-tab="fetch">数据抓取</div>
 </div>

 <div class="panel active" id="panel-records">
   <div class="toolbar">
    <label>每页</label><select id="pageSize"><option>20</option><option>50</option><option>100</option></select>
    <label>地图</label><select id="filterMap"><option value="">全部</option></select>
    <label>结果</label><select id="filterResult"><option value="">全部</option><option value="0">撤离成功</option><option value="1">撤离失败</option><option value="2">行动超时</option><option value="3">中途退出</option></select>
   <span style="border-left:1px solid #2a3a4a;margin:0 4px"></span>
    <label style="color:#8899aa;font-size:12px">开始</label>
    <button class="btn" style="font-size:12px;padding:4px 10px" id="btnRecStart" onclick="pickDate('recStartDate')">选择日期</button>
   <input type="date" id="recStartDate" style="position:absolute;opacity:0;pointer-events:none">
   <input type="time" id="recStartTime" value="00:00" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
    <label style="color:#8899aa;font-size:12px">结束</label>
    <button class="btn" style="font-size:12px;padding:4px 10px" id="btnRecEnd" onclick="pickDate('recEndDate')">选择日期</button>
   <input type="date" id="recEndDate" style="position:absolute;opacity:0;pointer-events:none">
   <input type="time" id="recEndTime" value="23:59" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
    <button class="btn" style="font-size:12px;padding:4px 10px" onclick="clearDateRange('records')">清空</button>
    <button class="btn btn-go" onclick="loadRecords()">查询</button>
  </div>
  <div class="scroll-table"><table><thead id="theadRec"></thead><tbody id="tbodyRec"></tbody></table></div>
  <div class="pager">
   <button class="btn btn-go" onclick="prevPage()">&lt;</button>
   <span id="pageInfo">-</span>
   <button class="btn btn-go" onclick="nextPage()"> &gt;</button>
  </div>
 </div>

 <div class="panel" id="panel-items">
   <div class="toolbar">
    <label style="color:#8899aa;font-size:12px">开始</label>
    <button class="btn" style="font-size:12px;padding:4px 10px" id="btnItemStart" onclick="pickDate('itemStartDate')">选择日期</button>
   <input type="date" id="itemStartDate" style="position:absolute;opacity:0;pointer-events:none">
   <input type="time" id="itemStartTime" value="00:00" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
    <label style="color:#8899aa;font-size:12px">结束</label>
    <button class="btn" style="font-size:12px;padding:4px 10px" id="btnItemEnd" onclick="pickDate('itemEndDate')">选择日期</button>
   <input type="date" id="itemEndDate" style="position:absolute;opacity:0;pointer-events:none">
   <input type="time" id="itemEndTime" value="23:59" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
    <button class="btn" style="font-size:12px;padding:4px 10px" onclick="clearDateRange('items')">清空</button>
    <button class="btn" id="btnItemOwnerScope" onclick="toggleItemOwnerScope()">范围: 全部</button>
    <button class="btn btn-go" onclick="loadItems()">刷新</button>
  </div>
  <div class="scroll-table"><table><thead id="theadItem"></thead><tbody id="tbodyItem"></tbody></table></div>
 </div>

 <div class="panel" id="panel-assets">
  <div class="summary" id="assetSummaryCards"></div>
  <div class="asset-tabs-row" style="margin-top:16px;display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap">
  <div class="tabs asset-tabs" id="assetCategoryTabs" style="margin-top:0"></div>
  <div id="assetUpdatedAt" style="font-size:12px;color:#8fa0ad;line-height:1.4;white-space:nowrap;flex:0 0 auto"></div>
 </div>
   <div class="toolbar" style="margin-top:16px">
   <label>等级</label><select id="assetGrade" onchange="loadAssets()"><option value="">全部</option></select>
   <button class="btn" style="font-size:12px;padding:4px 10px" onclick="clearAssetFilters()">清空筛选</button>
   </div>
   <div class="fetch-hint">
   <p><strong>刷新入口</strong>：请前往“数据抓取”页点击“刷新资产”，刷新时会同步更新最新物品列表和账号资产数据。</p>
   <p><strong>本地存储</strong>：资产数据会写入本地数据库 <code>collection</code> 表，后续仅做本地筛选和展示。</p>
   </div>
  <div class="scroll-table"><table><thead id="theadAsset"></thead><tbody id="tbodyAsset"></tbody></table></div>
 </div>

  <div class="panel" id="panel-analysis">
  <div class="toolbar">
   <label style="color:#8899aa;font-size:12px">开始</label>
   <button class="btn" style="font-size:12px;padding:4px 10px" id="btnAnaStart" onclick="pickDate('anaStartDate')">选择日期</button>
   <input type="date" id="anaStartDate" style="position:absolute;opacity:0;pointer-events:none">
   <input type="time" id="anaStartTime" value="00:00" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
   <label style="color:#8899aa;font-size:12px">结束</label>
   <button class="btn" style="font-size:12px;padding:4px 10px" id="btnAnaEnd" onclick="pickDate('anaEndDate')">选择日期</button>
   <input type="date" id="anaEndDate" style="position:absolute;opacity:0;pointer-events:none">
   <input type="time" id="anaEndTime" value="23:59" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
   <button class="btn" style="font-size:12px;padding:4px 10px" onclick="clearDateRange('analysis')">清空</button>
   <span style="border-left:1px solid #2a3a4a;margin:0 4px"></span>
   <button class="btn analysis-preset-btn" style="font-size:12px;padding:4px 10px" id="btnAnaPresetToday" onclick="setAnalysisPreset('today')">今天</button>
   <button class="btn analysis-preset-btn" style="font-size:12px;padding:4px 10px" id="btnAnaPresetYesterday" onclick="setAnalysisPreset('yesterday')">昨天</button>
   <button class="btn analysis-preset-btn" style="font-size:12px;padding:4px 10px" id="btnAnaPreset3" onclick="setAnalysisPreset(3)">近三天</button>
   <button class="btn analysis-preset-btn" style="font-size:12px;padding:4px 10px" id="btnAnaPreset7" onclick="setAnalysisPreset(7)">近一周</button>
   <button class="btn analysis-preset-btn" style="font-size:12px;padding:4px 10px" id="btnAnaPreset30" onclick="setAnalysisPreset(30)">近一个月</button>
   <button class="btn btn-go" onclick="loadAnalysis()">刷新分析</button>
  </div>
  <div id="analysisContent"></div>
 </div>

  <div class="panel" id="panel-trends">
  <div class="trend-sticky">
   <div class="toolbar">
    <label>时间范围</label>
    <button class="btn trend-range-btn" style="font-size:12px;padding:4px 10px" id="btnTrendRange7" onclick="setTrendRange('7')">近七天</button>
    <button class="btn trend-range-btn" style="font-size:12px;padding:4px 10px" id="btnTrendRange30" onclick="setTrendRange('30')">近30天</button>
    <button class="btn trend-range-btn" style="font-size:12px;padding:4px 10px" id="btnTrendRange90" onclick="setTrendRange('90')">近90天</button>
    <button class="btn trend-range-btn" style="font-size:12px;padding:4px 10px" id="btnTrendRangeAll" onclick="setTrendRange('all')">全部</button>
    <span style="border-left:1px solid #2a3a4a;margin:0 4px"></span>
    <label>展示维度</label><select id="trendBucket" onchange="loadTrends()"><option value="day">按日</option><option value="week">按周</option><option value="month">按月</option></select>
    <label>地图</label><select id="trendMap" onchange="loadTrends()"><option value="">全部地图</option></select>
    <button class="btn btn-go" onclick="loadTrends()">刷新趋势</button>
   </div>
   <div class="trend-meta" id="trendResolvedRange">当前范围：--</div>
   <div class="summary analysis-summary trend-summary" id="trendSummaryCards"></div>
  </div>
  <div id="trendContent"></div>
 </div>

   <div class="panel" id="panel-team">
    <div class="toolbar">
     <button class="btn btn-go" onclick="loadPlayerList()">刷新</button>
    </div>
     <div class="team-hint">选择 1-3 名队友，查看共同参与的对局数据</div>
     <div class="team-filter" style="display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap">
      <label style="color:#8899aa;font-size:12px">开始</label>
      <button class="btn" style="font-size:12px;padding:4px 10px" id="btnStartDate" onclick="pickDate('teamStartDate')">选择日期</button>
      <input type="date" id="teamStartDate" style="position:absolute;opacity:0;pointer-events:none">
      <input type="time" id="teamStartTime" value="00:00" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
      <label style="color:#8899aa;font-size:12px">结束</label>
      <button class="btn" style="font-size:12px;padding:4px 10px" id="btnEndDate" onclick="pickDate('teamEndDate')">选择日期</button>
      <input type="date" id="teamEndDate" style="position:absolute;opacity:0;pointer-events:none">
      <input type="time" id="teamEndTime" value="23:59" style="background:#1a2634;color:#ddeeff;border:1px solid #2a3a4a;border-radius:4px;padding:4px 8px;font-size:12px">
      <button class="btn" style="font-size:12px;padding:4px 10px" onclick="clearDateRange('team')">清空</button>
     <span style="border-left:1px solid #2a3a4a;margin:0 4px"></span>
      <button class="btn team-preset-btn" style="font-size:12px;padding:4px 10px" id="btnTeamPresetToday" onclick="setTeamPreset('today')">今天</button>
      <button class="btn team-preset-btn" style="font-size:12px;padding:4px 10px" id="btnTeamPresetYesterday" onclick="setTeamPreset('yesterday')">昨天</button>
      <button class="btn team-preset-btn" style="font-size:12px;padding:4px 10px" id="btnTeamPresetDayBefore" onclick="setTeamPreset('daybefore')">前天</button>
     </div>
    <div id="playerListContainer"></div>
    <div class="team-btns" id="teamBtns" style="display:none">
     <button class="btn btn-go" onclick="loadTeamAnalysis()">分析组队数据</button>
     <button class="btn btn-fetch" onclick="clearPlayerSelection()">清除选择</button>
     <button class="btn btn-muted" id="btnTeamMapFilter" onclick="openTeamMapFilter()">地图选择</button>
     <span style="border-left:2px solid #2a3a4a;margin:0 8px"></span>
     <button class="btn btn-go" onclick="exportTeamExcel()">导出 Excel</button>
     <button class="btn btn-fetch" onclick="downloadTeamImage()">导出图片</button>
     <button class="btn btn-fetch" id="btnTeamReport" onclick="generateTeamReport()">生成 PDF 报告</button>
     <button class="btn btn-muted" id="btnTeamReportManager" onclick="openTeamReportManager()">报告管理</button>
    </div>
    <div id="teamResultContainer" style="margin-top:16px"></div>
  </div>

  <div class="panel" id="panel-fetch">
 <div class="summary" id="summaryCards"></div>
  <div class="toolbar" style="margin-top:16px">
   <button class="btn btn-fetch" id="btnFetchLogin" onclick="doLogin()">登录 WeGame</button>
   <label>模式</label><select id="fetchQueue"><option value="sol">烽火地带</option><option value="mp">全面战场</option></select>
   <label>数量</label><input type="number" id="fetchCount" value="100" min="1" max="9999" style="width:80px">
  <button class="btn btn-fetch" id="btnSmartFetch" onclick="doSmartFetch()">智能抓取</button>
  <button class="btn btn-fetch" id="btnFetch" onclick="doFetch()">开始抓取</button>
  <button class="btn btn-muted" id="btnFetchDetails" onclick="doFetchDetails()">补全对局详情</button>
  <button class="btn btn-muted" id="btnAssetsSync" onclick="refreshAssets()">刷新资产</button>
  <button class="btn btn-muted" id="btnExportBackup" onclick="exportBattleBackup()">导出战绩备份</button>
  <button class="btn btn-muted" id="btnImportBackup" onclick="selectBattleBackupFile()">导入战绩备份</button>
  <button class="btn btn-danger" id="btnClearData" onclick="handleClearData()">清空数据</button>
  <input type="file" id="battleBackupFile" accept=".zip,application/zip" style="display:none" onchange="handleBattleBackupSelected(event)">
  </div>
  <div class="fetch-hint">
   <p><strong>智能抓取按钮</strong>：根据数据库内已有信息智能抓取缺失数据，第一次抓取时间较长。</p>
   <p><strong>开始抓取按钮</strong>：按用户填写的数量抓取最新对局，适合定量更新。</p>
   <p><strong>补全对局详情按钮</strong>：如上方总对局、详情、房间详情数量不一致，可以使用此功能补充抓取。</p>
   <p><strong>刷新资产按钮</strong>：仅在点击时访问账号资产接口，并同步更新最新物品列表，不会自动刷新。</p>
   <p><strong>战绩备份</strong>：导出 / 导入仅包含本地战绩数据库和 PDF 报告，不包含 WeGame 登录态和调试样本。</p>
   <p>如果抓取反复失败，账号可能需要手动输入滑块验证，请访问 <a href="https://www.wegame.com.cn/helper/df/" target="_blank" rel="noopener">WeGame 网站</a> 点击“查看最近战局”后，查看是否有验证码需要输入。</p>
  </div>
  <div class="fetch-progress" id="fetchProgress">
   <div class="row">
    <div class="meta"><span id="fetchTotalLabel">总进度</span><span id="fetchTotalText">-</span></div>
    <div class="bar"><div class="fill" id="fetchTotalFill"></div></div>
   </div>
   <div class="row">
    <div class="meta"><span id="fetchSubLabel">子进度</span><span id="fetchSubText">-</span></div>
    <div class="bar"><div class="fill" id="fetchSubFill"></div></div>
   </div>
  </div>
  <div id="fetchLog" style="background:#0a1018;border-radius:6px;padding:12px;font-family:monospace;font-size:12px;max-height:400px;overflow-y:auto;color:#8899aa"></div>
 </div>
</div>

<div class="account-overlay" id="accountGate">
 <div class="account-modal">
  <div class="account-modal-head">
   <div>
    <h3>选择账号</h3>
    <p id="accountGateHint">选择数据库中的账号后进入面板。</p>
   </div>
   <button class="close-link" id="accountGateClose" onclick="closeAccountGate()">关闭</button>
  </div>
  <div class="account-modal-body">
   <div id="accountGateBody"></div>
  </div>
 </div>
</div>

<div class="action-overlay" id="actionOverlay">
 <div class="action-card">
  <div class="action-card-head">
   <h3 id="actionOverlayTitle">处理中</h3>
   <button class="close-link" id="actionOverlayClose" onclick="closeActionOverlay()">关闭</button>
  </div>
  <div class="action-card-body" id="actionOverlayBody"></div>
  <div class="action-card-actions" id="actionOverlayActions"></div>
 </div>
</div>

<div class="account-overlay" id="softwareGate">
 <div class="account-modal">
  <div class="account-modal-head">
   <div>
    <h3>开源版</h3>
    <p id="softwareGateHint">开源版无需额外登录。</p>
   </div>
   <button class="close-link" id="softwareGateClose" onclick="closeSoftwareGate()">关闭</button>
  </div>
  <div class="account-modal-body" id="softwareGateBody"></div>
 </div>
</div>

<div class="action-overlay" id="teamReportOverlay">
 <div class="action-card report-card">
  <div class="action-card-head">
   <h3>报告管理</h3>
   <button class="close-link" onclick="closeTeamReportManager()">关闭</button>
  </div>
  <div class="action-card-body">
   <p class="dialog-note">报告会按提交顺序排队生成，您可在这里下载或删除；也可以稍后返回本页面继续下载已生成的报告。</p>
   <div class="report-batch-bar" id="teamReportBatchBar">
    <button class="btn btn-muted" onclick="selectAllTeamReports()">全选</button>
    <button class="btn btn-danger" onclick="deleteSelectedTeamReports()">删除所选</button>
    <button class="btn btn-muted" onclick="setTeamReportBatchMode(false)">退出批量删除</button>
   </div>
   <div class="report-list" id="teamReportList"></div>
  </div>
  <div class="action-card-actions">
   <button class="btn btn-danger" id="btnTeamReportBatch" onclick="setTeamReportBatchMode(true)">批量删除</button>
   <button class="btn btn-muted" onclick="refreshTeamReportList()">刷新</button>
   <button class="btn btn-fetch" onclick="closeTeamReportManager()">关闭</button>
  </div>
 </div>
</div>

<!-- modal -->
<div class="modal-overlay" id="modalOverlay" onclick="if(event.target===this)closeModal()">
 <div class="modal">
  <div class="modal-head">
    <h3 id="modalTitle">对局详情</h3>
   <button class="modal-close" onclick="closeModal()">&times;</button>
  </div>
  <div class="modal-body" id="modalBody"></div>
 </div>
</div>
<div class="battle-tag-tooltip" id="battleTagTooltip"></div>

<script>
let page=1,total=0,psize=20;
let itemRows=[],itemSort={key:'event_time',order:'desc'},itemOwnerScope='all';
let assetRows=[],assetFilters={categories:[],grades:[]},assetCurrentCategory='',assetQuickCollectibleOnly=false;
let selectedPlayers=[];
let recordsDateInited=false,itemsDateInited=false,analysisDateInited=false,teamDateInited=false;
let accountsCache=[],headerStatsCache=null,mapsLoaded=false,trendMapsLoaded=false;
let hasEnteredAccount=false,accountGateLocked=true;
let accountManageMode=false,pendingDeleteAccount=null;
let currentActionJobId='',currentActionLogCount=0,currentActionPollTimer=null;
let teamReportPollTimer=null;
let teamReportManagerTimer=null,teamReportKnownDone=new Set(),teamReportKnownInited=false,pendingTeamReportPayload=null,pendingDeleteReportId='',teamReportBatchMode=false,teamReportSelectedIds=new Set(),teamReportLastList=[];
let fetchConsecutiveFailures=0;
let msgTimer=null;
let analysisPresetDays=0;
let trendRange='30';
let teamPresetMode='';
let teamActivePlayerName='';
let teamActivePlayerKey='';
let teamSelfDeselected=false;
let teamMapFilterSelected=new Set();
let softwareSessionCache={logged_in:false,backend_ok:true,user:null,entitlements:null,accounts:[]};
let softwareAuthMode='login';
const FETCH_ACTION_IDS=['btnFetch','btnSmartFetch','btnFetchDetails','btnAssetsSync','btnClearData'];
const TEAM_OTHER_TOKEN='__OTHER_PLAYER__';
const ROLE_MAP={'10007':'红狼','10010':'威龙','10011':'无名','10012':'疾风','20003':'蜂医','20004':'蛊','20005':'蝶','30008':'牧羊人','30009':'乌鲁鲁','30010':'深蓝','30011':'比特','40005':'露娜','40010':'骇爪','40011':'银翼','40012':'回响','50001':'赤枭','50002':'赤枭亲卫','50003':'赤枭亲卫'};
const RESULT_MAP={0:'撤离成功',1:'撤离失败',2:'行动超时',3:'中途退出'};
const RESULT_CLASS={0:'b0',1:'b1',2:'b2',3:'b3'};
const TREND_METRICS=[
 {key:'total',label:'总对局',format:'int'},
 {key:'escaped',label:'撤离成功',format:'int'},
 {key:'evac_rate_num',label:'撤离率',format:'percent'},
 {key:'kd',label:'KD',format:'decimal'},
 {key:'total_player_kills',label:'玩家击杀',format:'int'},
 {key:'total_profit',label:'总盈亏',format:'money'},
 {key:'avg_profit',label:'平均盈亏',format:'money'},
 {key:'avg_duration',label:'平均时长',format:'duration'},
 {key:'max_collection',label:'最高带出',format:'money'},
 {key:'max_profit',label:'最高盈利',format:'money'}
];
function postJSON(url,body){return fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body||{})}).then(r=>r.json())}
function esc(v){return String(v??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
function appendFetchLogs(lines){(lines||[]).forEach(line=>appendFetchLog(line))}
function appendFetchLog(msg){
 let box=$('fetchLog');
 if(!box)return;
 let line=document.createElement('div');
 line.textContent='['+new Date().toLocaleTimeString('zh-CN',{hour12:false})+'] '+msg;
 box.prepend(line);
}
function resetFetchFailureCounter(){fetchConsecutiveFailures=0}
function noteFetchFailure(message){
 if(isLoginExpiredError(message))return;
 let text=errorText(message);
 if(text.includes('当前已有任务进行中'))return;
 fetchConsecutiveFailures++;
 if(fetchConsecutiveFailures<2)return;
 showActionOverlay({
  title:'',
  html:'<p>如果抓取反复失败，账号可能需要手动输入滑块验证，请访问 <a href="https://www.wegame.com.cn/helper/df/" target="_blank" rel="noopener">WeGame 网站</a> 点击“查看最近战局”后，查看是否有验证码需要输入。</p><p class="dialog-note">最近一次错误：'+esc(translateErrorMessage(message||''))+'</p>',
  actions:[{label:'关闭',className:'btn btn-fetch',onClick:'closeActionOverlay()'}],
  closable:true
 });
}
function renderBarTrack(pct,color,text){
 let outside=(Number(pct)||0)<18;
 return '<div class="bar-track"><div class="bar-fill" style="width:'+pct+'%;background:'+color+'"></div><div class="bar-text'+(outside?' outside':'')+'">'+esc(text)+'</div></div>';
}
function resetFetchProgress(){
 let wrap=$('fetchProgress');
 if(wrap)wrap.classList.remove('show');
 let ids=['fetchTotalFill','fetchSubFill'];
 ids.forEach(id=>{let el=$(id);if(el)el.style.width='0%'});
 let totalLabel=$('fetchTotalLabel'),totalText=$('fetchTotalText'),subLabel=$('fetchSubLabel'),subText=$('fetchSubText');
 if(totalLabel)totalLabel.textContent='总进度';
 if(totalText)totalText.textContent='-';
 if(subLabel)subLabel.textContent='子进度';
 if(subText)subText.textContent='-';
}
function renderFetchProgress(progress,action){
 let wrap=$('fetchProgress');
 if(!wrap)return;
 if(action!=='fetch' || !progress){
  wrap.classList.remove('show');
  return;
 }
 wrap.classList.add('show');
 let total=progress.total||{},sub=progress.sub||{};
 let totalFill=$('fetchTotalFill'),subFill=$('fetchSubFill');
 let totalLabel=$('fetchTotalLabel'),totalText=$('fetchTotalText'),subLabel=$('fetchSubLabel'),subText=$('fetchSubText');
 if(totalLabel)totalLabel.textContent=total.label||'总进度';
 if(totalText)totalText.textContent=total.text||'-';
 if(subLabel)subLabel.textContent=sub.label||'子进度';
 if(subText)subText.textContent=sub.text||'-';
 if(totalFill)totalFill.style.width=(Number(total.percent)||0)+'%';
 if(subFill)subFill.style.width=(Number(sub.percent)||0)+'%';
}
function stopActionPolling(){
 if(currentActionPollTimer){
  clearTimeout(currentActionPollTimer);
  currentActionPollTimer=null;
 }
 currentActionJobId='';
 currentActionLogCount=0;
}
function errorText(err){
 if(err&&typeof err==='object'&&'message' in err)return String(err.message||'');
 return String(err||'');
}
function translateErrorMessage(message){
 let text=errorText(message).trim();
 if(!text)return '未知错误';
 text=text.replace(/^Error:\s*/,'');
 let lower=text.toLowerCase();
 if(lower.includes('target page, context or browser has been closed'))return '用户关闭了浏览器窗口';
 if(lower.includes('browser has been closed'))return '浏览器窗口已关闭';
 if(lower.includes('page has been closed'))return '页面窗口已关闭';
 if(lower.includes('context has been closed'))return '浏览器登录上下文已关闭';
 if(lower.includes('failed to fetch'))return '网络请求失败，请检查程序是否仍在运行';
 if(lower.includes('networkerror when attempting to fetch resource'))return '网络请求失败，请稍后重试';
 if(text.includes('Cookie 文件不存在'))return 'WeGame 登录凭证不存在，请重新登录';
 if(text.includes('无法获取 openid'))return '登录信息已失效，无法获取账号标识，请重新登录';
 if(text.includes('未能获取账号信息'))return '登录已完成，但未成功读取账号信息，请重试';
 return text;
}
function isLoginExpiredError(message){
 let text=errorText(message).toLowerCase();
 return text.includes('cookie 文件不存在')
  || text.includes('cookie')
  || text.includes('凭证')
  || text.includes('登录已过期')
  || text.includes('登录信息过期')
  || text.includes('8025004')
  || text.includes('无法获取 openid')
  || text.includes('未能获取账号信息')
  || text.includes('请先重新登录')
  || text.includes('请先运行 delta-force-data-center login 登录');
}
function notifyLoginExpired(message){
 if(!isLoginExpiredError(message))return false;
 let detail=translateErrorMessage(message);
 showActionOverlay({
  title:'WeGame 登录已过期',
  html:'<p>当前登录信息已失效，请重新登录后再继续抓取或补全详情。</p><p class="dialog-note">原因：'+esc(detail)+'</p>',
  actions:[
   {label:'登录 WeGame',className:'btn btn-fetch',onClick:'closeActionOverlay();doLogin()'},
   {label:'关闭',className:'btn btn-muted',onClick:'closeActionOverlay()'}
  ],
  closable:true
 });
 return true;
}
function beginActionPolling(jobId,onDone){
 stopActionPolling();
 currentActionJobId=jobId;
 currentActionLogCount=0;
 const tick=()=>{
  fetch('/api/action-status?job_id='+encodeURIComponent(jobId)).then(r=>r.json()).then(d=>{
   if(!d.ok)throw new Error(d.error||'状态查询失败');
  let logs=Array.isArray(d.logs)?d.logs:[];
  if(logs.length>currentActionLogCount){
    appendFetchLogs(logs.slice(currentActionLogCount));
    currentActionLogCount=logs.length;
   }
   renderFetchProgress(d.progress,d.action);
   if(d.status==='running'){
    currentActionPollTimer=setTimeout(tick,900);
    return;
   }
   stopActionPolling();
   setButtonsDisabled(FETCH_ACTION_IDS,false);
   if(d.status==='done'){
    resetFetchFailureCounter();
    onDone(d.result||{});
   return;
  }
  let err=d.error||'未知错误';
  notifyLoginExpired(err);
   let msg=translateErrorMessage(err);
   showMsg('err','任务失败：'+msg);
   appendFetchLog('任务失败：'+msg);
   noteFetchFailure(err);
  }).catch(e=>{
  if(currentActionJobId!==jobId)return;
   currentActionPollTimer=setTimeout(tick,1200);
  });
 };
 tick();
}
function setButtonsDisabled(ids,disabled){ids.forEach(id=>{let el=$(id);if(el)el.disabled=disabled})}
function showActionOverlay(options,text){
 let opts=typeof options==='object'&&options!==null?options:{title:options,text:text};
 let overlay=$('actionOverlay');
 let titleEl=$('actionOverlayTitle');
 let bodyEl=$('actionOverlayBody');
 let actionsEl=$('actionOverlayActions');
 let closeBtn=$('actionOverlayClose');
 if(titleEl)titleEl.textContent=Object.prototype.hasOwnProperty.call(opts,'title')?opts.title:'处理中';
 if(bodyEl){
  if(opts.html)bodyEl.innerHTML=opts.html;
  else bodyEl.innerHTML='<p>'+esc(opts.text||'')+'</p>';
 }
 if(actionsEl){
  let actions=Array.isArray(opts.actions)?opts.actions:[];
  actionsEl.innerHTML=actions.map(action=>'<button class="'+esc(action.className||'btn btn-fetch')+'" onclick="'+esc(action.onClick||'closeActionOverlay()')+'">'+esc(action.label||'确定')+'</button>').join('');
 }
 if(closeBtn)closeBtn.style.display=opts.closable?'inline-block':'none';
 if(overlay)overlay.classList.add('show');
}
function closeActionOverlay(){
 let overlay=$('actionOverlay');
 if(overlay)overlay.classList.remove('show');
}
function hideActionOverlay(){closeActionOverlay()}

function $(id){return document.getElementById(id)}
function closeMsg(){
 let e=$('msg');
 if(!e)return;
 if(msgTimer){clearTimeout(msgTimer);msgTimer=null}
 e.className='msg';
 e.innerHTML='';
}
function showMsg(t,s){
 let e=$('msg');
 if(!e)return;
 if(msgTimer){clearTimeout(msgTimer);msgTimer=null}
 e.className='msg '+t+' show';
 e.innerHTML='<div class="msg-text">'+esc(s)+'</div><button class="msg-close" type="button" onclick="closeMsg()" aria-label="关闭提示">&times;</button>';
 msgTimer=setTimeout(closeMsg,60000);
}
function fmtDur(s){if(!s)return'-';let m=Math.floor(s/60),sec=s%60;return m+'分'+sec+'秒'}
function fmtPrice(v){if(v==null||v===undefined)return'-';return Number(v).toLocaleString()}
function fmtCompactNumber(v){
 if(v==null||v===undefined)return'-';
 let n=Number(v)||0,abs=Math.abs(n);
 if(abs>=1e9)return (n/1e9).toFixed(2).replace(/\.00$/,'')+'B';
 if(abs>=1e6)return (n/1e6).toFixed(2).replace(/\.00$/,'')+'M';
 if(abs>=1e3)return (n/1e3).toFixed(2).replace(/\.00$/,'')+'K';
 return fmtPrice(n);
}
function moneyToneStyle(v){
 let n=Number(v)||0;
 if(n>0)return document.body.classList.contains('light-theme')?'color:#1d9960;font-weight:700':'color:#57e58d;font-weight:700';
 if(n<0)return document.body.classList.contains('light-theme')?'color:#c45145;font-weight:700':'color:#ff8c7a;font-weight:700';
 return '';
}
function fmtMoneyWithTone(v){
 return '<span style="'+moneyToneStyle(v)+'">'+fmtPrice(v)+'</span>';
}
function roleName(id){return ROLE_MAP[String(id)]||id||'-'}
function resultBadge(r){return '<span class="badge '+(RESULT_CLASS[r]||'b3')+'">'+(RESULT_MAP[r]||r)+'</span>'}
const GRADE_MAP={1:'白',2:'绿',3:'蓝',4:'紫',5:'金',6:'红'};
const GRADE_COLOR={1:'#aaa',2:'#44ff88',3:'#4a9eff',4:'#c084fc',5:'#ffcc00',6:'#ff4444'};
function gradeLabel(g){if(!g)return'-';return '<span style="color:'+(GRADE_COLOR[g]||'#aaa')+';font-weight:bold">'+GRADE_MAP[g]+'</span>'}
function tagClass(dim){return dim==='负面评价'?'neg':dim==='中性评价'?'mid':'pos'}
function renderTagList(tags,extraClass){
 let list=Array.isArray(tags)?tags:[];
 if(!list.length)return '';
 return '<div class="tag-list '+(extraClass||'')+'">'+list.map(t=>{
  let tip=esc(t.tip||t.rule_text||'').replace(/\n/g,'&#10;');
  return '<span class="battle-tag '+tagClass(t.dimension)+'" data-tip="'+tip+'">'+esc(t.name||'标签')+'</span>';
 }).join('')+'</div>';
}
function renderDetailHead(title,tags){
 let tagHtml=renderTagList(tags,'detail-head-tags');
 let meta=tagHtml
  ? '<div class="detail-player-meta"><span class="meta-label">对局评价</span>'+tagHtml+'</div>'
  : '<div class="detail-player-meta empty"><span class="meta-label">对局评价：无</span></div>';
 return '<div class="detail-player-head"><div class="section-title">'+title+'</div>'+meta+'</div>';
}
let activeBattleTagTooltip=null;
function hideBattleTagTooltip(){
 let tip=$('battleTagTooltip');
 activeBattleTagTooltip=null;
 if(tip)tip.classList.remove('show');
}
function positionBattleTagTooltip(tag){
 let tip=$('battleTagTooltip');
 if(!tip||!tag)return;
 let rect=tag.getBoundingClientRect();
 let gap=12,pad=12;
 tip.classList.add('show');
 tip.style.visibility='hidden';
 tip.style.left='0px';
 tip.style.top='0px';
 let width=tip.offsetWidth||220;
 let height=tip.offsetHeight||48;
 let left=Math.min(Math.max(pad,rect.left),Math.max(pad,window.innerWidth-width-pad));
 let top=rect.bottom+gap;
 if(top+height>window.innerHeight-pad)top=rect.top-height-gap;
 if(top<pad)top=Math.max(pad,window.innerHeight-height-pad);
 tip.style.left=Math.round(left)+'px';
 tip.style.top=Math.round(top)+'px';
 tip.style.visibility='visible';
}
function showBattleTagTooltip(tag){
 let tip=$('battleTagTooltip');
 if(!tip||!tag)return;
 let text=(tag.getAttribute('data-tip')||'').replace(/&#10;/g,'\n');
 if(!text){hideBattleTagTooltip();return}
 activeBattleTagTooltip=tag;
 tip.textContent=text;
 positionBattleTagTooltip(tag);
}
document.addEventListener('mouseover',e=>{
 let tag=e.target.closest('.battle-tag,[data-tip]');
 if(!tag)return;
 showBattleTagTooltip(tag);
});
document.addEventListener('mouseout',e=>{
 let tag=e.target.closest('.battle-tag,[data-tip]');
 if(!tag)return;
 let next=e.relatedTarget&&e.relatedTarget.closest?e.relatedTarget.closest('.battle-tag,[data-tip]'):null;
 if(next===tag)return;
 hideBattleTagTooltip();
});
document.addEventListener('scroll',()=>{if(activeBattleTagTooltip)positionBattleTagTooltip(activeBattleTagTooltip)},true);
window.addEventListener('resize',()=>{if(activeBattleTagTooltip)positionBattleTagTooltip(activeBattleTagTooltip)});

function applyTheme(theme){
 let light=theme==='light';
 document.body.classList.toggle('light-theme',light);
 let btn=$('themeToggle');
 if(btn)btn.textContent=light?'切换暗色':'切换亮色';
}

function toggleTheme(){
 let next=document.body.classList.contains('light-theme')?'dark':'light';
 try{localStorage.setItem('delta-force-data-center-theme',next)}catch(e){}
 applyTheme(next);
}

function initTheme(){
 let theme='light';
 try{theme=localStorage.getItem('delta-force-data-center-theme')||'light'}catch(e){}
 applyTheme(theme);
}

function currentTab(){
 let tab=document.querySelector('.tab.active');
 return tab?tab.dataset.tab:'records';
}

function applyHeaderState(stats){
 headerStatsCache=stats||{};
 if(stats&&stats.software_session)applySoftwareState(stats.software_session);
 let switchBtn=$('btnSwitchAccount');
 if(switchBtn){
  let label=stats.active_account||'未选择';
  switchBtn.textContent='玩家 · '+label;
 }
}

function applySoftwareState(snapshot){
 softwareSessionCache=snapshot||{logged_in:false,backend_ok:true,user:null,entitlements:null,accounts:[]};
 let user=softwareSessionCache.user||null;
 let btn=$('btnSoftwareAccount');
 let badge=user&&softwareSessionCache.logged_in?('用户 · '+(user.username||'-')):'用户 · 未登录';
 if(!softwareSessionCache.backend_ok&&softwareSessionCache.error)badge='用户 · 后端异常';
 if(btn)btn.textContent=badge;
}

function openFeatureManagement(){
 let html='<p>开源版无需额外登录，所有本地功能均可直接使用。</p>';
 showActionOverlay({
  title:'开源版',
  html:html,
  actions:[{label:'关闭',className:'btn btn-muted',onClick:'closeActionOverlay()'}],
  closable:true
 });
}

function openSoftwareMenu(){
 let html='<div class="account-empty" style="text-align:left">开源版无需额外登录，也不限制可使用的游戏账号数量。</div>';
 showActionOverlay({
  title:'开源版',
  html:html,
  actions:[
   {label:'关闭',className:'btn btn-muted',onClick:'closeActionOverlay()'}
  ],
  closable:true
 });
}

function openPlayerMenu(){
 let stats=headerStatsCache||{};
 let activeName=stats.active_account||'未选择';
 let loginText=stats.logged_in?'当前 WeGame 已登录，可切换账号或退出登录。':'当前 WeGame 未登录，可先切换账号或重新登录。';
 showActionOverlay({
  title:'玩家 · '+esc(activeName),
  html:'<p>'+loginText+'</p><p class="dialog-note">玩家账号相关操作已移动到这里。</p>',
  actions:[
   {label:'切换账号',className:'btn btn-fetch',onClick:'closeActionOverlay();syncHeaderState({forceGate:true,allowClose:true})'},
   {label:'退出 WeGame',className:'btn btn-danger',onClick:'closeActionOverlay();doLogout()'},
   {label:'关闭',className:'btn btn-muted',onClick:'closeActionOverlay()'}
  ],
  closable:true
 });
}

function renderSoftwareGate(){
 let body=$('softwareGateBody');
 let hint=$('softwareGateHint');
 let closeBtn=$('softwareGateClose');
 if(!body||!hint||!closeBtn)return;
 closeBtn.style.display='inline-block';
 hint.textContent='开源版无需额外登录。';
 body.innerHTML='<div class="account-empty">所有本地功能均可直接使用，游戏账号数量不受限制。</div>';
}

function openSoftwareGate(){
 renderSoftwareGate();
 let gate=$('softwareGate');
 if(gate)gate.classList.add('show');
}

function closeSoftwareGate(){
 let gate=$('softwareGate');
 if(gate)gate.classList.remove('show');
}

function switchSoftwareAuthMode(mode){
 softwareAuthMode=mode==='register'?'register':'login';
 renderSoftwareGate();
}

function syncOpenSourceState(options){
 return fetch('/api/software-session').then(r=>r.json()).then(d=>{
  if(d.ok===false)throw new Error(d.error||'开源版状态读取失败');
  applySoftwareState(d);
  if(options&&options.forceOpen&&!softwareSessionCache.logged_in)openSoftwareGate();
  if(softwareSessionCache.logged_in&&options&&options.closeGate)closeSoftwareGate();
  return softwareSessionCache;
 }).catch(e=>{
  applySoftwareState({logged_in:false,backend_ok:false,error:String(e),backend_url:''});
  if(options&&options.forceOpen)openSoftwareGate();
  return softwareSessionCache;
 });
}

function submitSoftwareAuth(){
 let username=$('softwareUsername')?.value.trim()||'';
 let password=$('softwarePassword')?.value||'';
 if(!username||!password){
  showMsg('err','请输入用户名和密码');
  return;
 }
 let action=softwareAuthMode==='register'?'/api/software/register':'/api/software/login';
 postJSON(action,{username,password}).then(d=>{
  if(!d.ok)throw new Error(d.error||'操作失败');
  let next=softwareAuthMode==='register'?postJSON('/api/software/login',{username,password}):Promise.resolve(d);
  return next;
 }).then(d=>{
  if(!d.ok)throw new Error(d.error||'登录失败');
  return syncOpenSourceState({closeGate:true});
 }).then(()=>{
  showMsg('ok','开源版功能已启用');
  syncHeaderState({forceGate:true,allowClose:true});
 }).catch(e=>showMsg('err','开源版状态操作失败：'+translateErrorMessage(e)));
}

function logoutSoftwareAccount(){
 if(!confirm('开源版无需退出。是否关闭此提示？'))return;
 postJSON('/api/software/logout',{}).then(d=>{
  if(!d.ok)throw new Error(d.error||'退出失败');
  return syncOpenSourceState({forceOpen:true});
 }).then(()=>{
  showMsg('ok','开源版功能保持可用');
 }).catch(e=>showMsg('err','开源版状态操作失败：'+translateErrorMessage(e)));
}

function softwareFeatureEnabled(key){
 return true;
}

function showMemberFeaturePrompt(featureLabel){
 showActionOverlay({
  title:'功能可用',
  html:'<p>'+esc(featureLabel||'该功能')+' 在开源版中可直接使用。</p>',
  actions:[{label:'关闭',className:'btn btn-muted',onClick:'closeActionOverlay()'}],
  closable:true
 });
}

function ensureSoftwareAccess(featureKey,featureLabel){
 return true;
}

function accountManageHint(){
 return accountManageMode?'管理模式下可删除账号，删除会同时清空该账号数据。':'点击账号卡片可切换当前查看账号。';
}

function renderAccountGate(stats){
 let hint=$('accountGateHint');
 let body=$('accountGateBody');
 let closeBtn=$('accountGateClose');
 if(closeBtn)closeBtn.style.display=accountGateLocked?'none':'inline-block';
 if(!hint||!body)return;
 if(!accountsCache.length){
  hint.textContent='当前数据库还没有可用账号';
  body.innerHTML='<div class="account-empty">当前数据库还没有可用账号。请关闭此窗口后前往“数据抓取”页，点击“登录 WeGame”完成扫码登录。</div>';
  return;
 }
 hint.textContent=accountManageMode?'管理数据库中的账号。':'请选择数据库中的账号后进入面板。';
 let cards=accountsCache.map(acc=>{
  let active=acc.is_active?' active':'';
  let current=acc.is_active?'<span style="margin-left:8px;font-size:12px;color:#00d4aa">当前</span>':'';
  let manage=accountManageMode?' manage-mode':'';
 let pid=esc(acc.player_id||'');
  let name=esc(acc.player_name||acc.player_id||'未知账号');
  let level=acc.level?('等级 '+acc.level):'未记录等级';
  let last=esc(String(acc.last_login||'').replace('T',' '));
  let deleteBtn=accountManageMode?'<button class="btn btn-danger account-delete-btn" data-player-id="'+pid+'" data-player-name="'+name+'">删除账号</button>':''; 
  return '<div class="account-card'+active+manage+'" data-player-id="'+pid+'"><div class="account-card-head"><div class="name">'+name+current+'</div><div class="account-card-tools">'+deleteBtn+'</div></div><div class="meta">PlayerID: '+pid+'<br>'+level+'<br>最近登录: '+(last||'-')+'</div></div>';
 }).join('');
 let manageBtnText=accountManageMode?'完成管理':'管理账号';
 body.innerHTML='<div class="account-actions"><div class="left"><button class="btn btn-muted" id="btnToggleAccountManage" onclick="toggleAccountManageMode()">'+manageBtnText+'</button></div><div class="right">'+accountManageHint()+'</div></div><div class="account-list">'+cards+'</div>';
 body.querySelectorAll('.account-card').forEach(card=>{
  card.onclick=()=>{
   if(accountManageMode)return;
   selectAccount(card.dataset.playerId);
  };
 });
 body.querySelectorAll('.account-delete-btn').forEach(btn=>{
  btn.onclick=(e)=>{
   e.stopPropagation();
   handleDeleteAccount(btn.dataset.playerId,btn.dataset.playerName);
  };
 });
}

function openAccountGate(allowClose){
 accountGateLocked=!allowClose;
 accountManageMode=false;
 renderAccountGate(headerStatsCache||{});
 let gate=$('accountGate');
 if(gate)gate.classList.add('show');
}

function closeAccountGate(){
 if(accountGateLocked)return;
 let gate=$('accountGate');
 if(gate)gate.classList.remove('show');
}

function syncHeaderState(options){
 let opts=options||{};
 return Promise.all([
  fetch('/api/stats').then(r=>r.json()),
  fetch('/api/accounts').then(r=>r.json())
 ]).then(([stats,accountData])=>{
  accountsCache=Array.isArray(accountData.accounts)?accountData.accounts:(Array.isArray(accountData)?accountData:[]);
  applyHeaderState(stats);
  renderAccountGate(stats);
  if(stats.active_player_id)hasEnteredAccount=true;
  let mustChoose=opts.forceGate||!accountsCache.length||!stats.active_player_id;
  if(mustChoose){
   let canClose=opts.allowClose===true||!accountsCache.length;
   openAccountGate(canClose);
  }else if(opts.closeGate){
   let gate=$('accountGate');
   if(gate)gate.classList.remove('show');
  }
  return stats;
 })
}

function selectAccount(playerId){
 if(!playerId)return;
 postJSON('/api/account/select',{player_id:playerId}).then(d=>{
   if(!d.ok){
    showMsg('err','切换账号失败: '+(d.error||'未知错误'));
    return
   }
  hasEnteredAccount=true;
  accountManageMode=false;
  let gate=$('accountGate');
  if(gate)gate.classList.remove('show');
  resetScopedState();
  showMsg('ok','已切换到 '+(d.active_account||playerId));
  syncHeaderState({closeGate:true}).then(()=>{loadSummary();refreshVisiblePanel()});
 }).catch(e=>showMsg('err','切换账号失败: '+e))
}

function toggleAccountManageMode(){
 accountManageMode=!accountManageMode;
 renderAccountGate(headerStatsCache||{});
}

function handleDeleteAccount(playerId,playerName){
 pendingDeleteAccount={player_id:playerId,name:playerName||playerId};
 showActionOverlay({
  title:'确认删除账号',
  html:'<p>删除账号会同时删除该账号关联的所有战绩、物品和详情数据。</p><p>请输入 <span class=\"danger-keyword\">删除账号</span> 后再执行操作。</p><div style=\"margin-top:14px\"><input id=\"deleteAccountConfirmInput\" class=\"dialog-input\" type=\"text\" placeholder=\"请输入：删除账号\"></div><p class=\"dialog-note\">目标账号：'+esc(playerName||playerId)+'</p>',
  actions:[
   {label:'取消',className:'btn btn-muted',onClick:'closeActionOverlay()'},
   {label:'确认删除',className:'btn btn-danger',onClick:'confirmDeleteAccount()'}
  ],
  closable:true
 });
 setTimeout(()=>{let input=$('deleteAccountConfirmInput');if(input)input.focus()},30);
}

function confirmDeleteAccount(){
 let input=$('deleteAccountConfirmInput');
 let value=input?input.value.trim():'';
 if(value!=='删除账号'){
  showMsg('err','请输入“删除账号”后再执行删除');
  if(input)input.focus();
  return;
 }
 let target=pendingDeleteAccount;
 if(!target)return;
 closeActionOverlay();
 postJSON('/api/account/delete',{player_id:target.player_id}).then(d=>{
  pendingDeleteAccount=null;
  appendFetchLogs(d.logs);
  if(!d.ok){
   showMsg('err','删除账号失败：'+translateErrorMessage(d.error||'未知错误'));
   return;
  }
  hasEnteredAccount=!!d.active_player_id;
  accountManageMode=false;
  resetScopedState();
  showMsg('ok','已删除账号 '+(d.deleted_account||target.name));
  syncHeaderState({forceGate:!d.active_player_id,allowClose:true,closeGate:!!d.active_player_id}).then(()=>{
   loadSummary();
   refreshVisiblePanel();
  });
 }).catch(e=>{
  pendingDeleteAccount=null;
  showMsg('err','删除账号失败：'+translateErrorMessage(e));
 })
}

function doLogin(){
 let btn=$('btnFetchLogin');
 if(btn)btn.disabled=true;
 showActionOverlay({title:'请完成 WeGame 登录',text:'系统即将弹出 WeGame 窗口，请在弹出的窗口中扫码登录。若已进入 WeGame 三角洲助手页面，建议点击“查看最近战局”主动进入对局记录页，以便系统更快完成登录态校验与接口获取。登录完成后页面会自动切换到该账号。',closable:false});
 showMsg('info','正在打开 WeGame 登录窗口。完成扫码后，如已进入助手页面，建议点击“查看最近战局”以加快接口获取');
 appendFetchLog('开始 WeGame 登录流程');
 setTimeout(()=>postJSON('/api/login',{}).then(d=>{
   if(btn)btn.disabled=false;
   hideActionOverlay();
   if(!d.ok){
   let msg=translateErrorMessage(d.error||'未知错误');
   notifyLoginExpired(d.error);
   showMsg('err','登录失败：'+msg);
   appendFetchLogs(d.logs);
   appendFetchLog('登录失败：'+msg);
    return;
   }
   hasEnteredAccount=true;
   appendFetchLogs(d.logs);
   appendFetchLog('登录成功: '+(d.active_account||'已写入账号信息'));
   resetFetchFailureCounter();
   showMsg('ok','登录成功，已进入当前登录账号');
   resetScopedState();
   syncHeaderState({closeGate:true}).then(()=>{loadSummary();refreshVisiblePanel()});
 }).catch(e=>{
   if(btn)btn.disabled=false;
   hideActionOverlay();
   let msg=translateErrorMessage(e);
   notifyLoginExpired(e);
   showMsg('err','登录失败：'+msg);
   appendFetchLog('登录请求失败：'+msg);
 }),80)
}

function doLogout(){
 if(!confirm('确认清除当前 WeGame 登录凭证吗？'))return;
 let btn=$('btnSwitchAccount');
 if(btn)btn.disabled=true;
 postJSON('/api/logout',{}).then(d=>{
  if(btn)btn.disabled=false;
  if(!d.ok){showMsg('err','退出登录失败：'+translateErrorMessage(d.error||'未知错误'));return}
  showMsg('ok','已清除登录凭证');
  appendFetchLog('已清除 WeGame 登录凭证');
  syncHeaderState({closeGate:false});
 }).catch(e=>{
  if(btn)btn.disabled=false;
  showMsg('err','退出登录失败：'+translateErrorMessage(e));
 })
}

function bindCurrentGameAccount(){
 postJSON('/api/software/bind-current',{}).then(d=>{
  if(!d.ok)throw new Error(d.error||'操作失败');
  if(d.software_session)applySoftwareState(d.software_session);
  showMsg('ok','当前游戏账号可直接使用');
  syncHeaderState({forceGate:true,allowClose:true});
 }).catch(e=>showMsg('err','账号状态更新失败：'+translateErrorMessage(e)));
}

function resetScopedState(){
 page=1;
 total=0;
 selectedPlayers=[];
 analysisPresetDays=0;
 trendRange='30';
 teamPresetMode='';
 teamActivePlayerName='';
 teamActivePlayerKey='';
 teamSelfDeselected=false;
 teamMapFilterSelected.clear();
 lastTeamData=null;
 recordsDateInited=false;
 itemsDateInited=false;
 analysisDateInited=false;
 teamDateInited=false;
 updateAnalysisPresetButtons();
 if(typeof updateTrendRangeButtons==='function')updateTrendRangeButtons();
 updateTeamPresetButtons();
 if(typeof setDateRange==='function'){
  setDateRange('recStartDate','btnRecStart','recEndDate','btnRecEnd','','');
  setDateRange('itemStartDate','btnItemStart','itemEndDate','btnItemEnd','','');
  setDateRange('anaStartDate','btnAnaStart','anaEndDate','btnAnaEnd','','');
  setDateRange('teamStartDate','btnStartDate','teamEndDate','btnEndDate','','');
 }
 let pageInfo=$('pageInfo');
 if(pageInfo)pageInfo.textContent='-';
 let playerList=$('playerListContainer');
 if(playerList)playerList.innerHTML='';
 let teamResult=$('teamResultContainer');
 if(teamResult)teamResult.innerHTML='';
 let trendContent=$('trendContent');
 if(trendContent)trendContent.innerHTML='';
 let trendSummary=$('trendSummaryCards');
 if(trendSummary)trendSummary.innerHTML='';
 let teamBtns=$('teamBtns');
 if(teamBtns)teamBtns.style.display='none';
 updateTeamMapFilterButton();
}

function refreshVisiblePanel(){
 if(!ensureSoftwareAccess('', '本地功能'))return;
 let tab=currentTab();
 if(tab==='assets'&&!ensureSoftwareAccess('assets','账号资产'))return;
 if(tab==='assets'){loadAssets();return}
 if(tab==='analysis'&&!ensureSoftwareAccess('analysis','数据分析'))return;
 if(tab==='trends'&&!ensureSoftwareAccess('trend_analysis','数据趋势'))return;
 if(tab==='team'&&!ensureSoftwareAccess('team_analysis','组队分析'))return;
 if(tab==='records'){loadMapFilter();initRecordsDateFilter();loadRecords();return}
 if(tab==='items'){initItemsDateFilter();loadItems();return}
 if(tab==='analysis'){initAnalysisDateFilter();loadAnalysis();return}
 if(tab==='trends'){loadTrendMapFilter();loadTrends();return}
 if(tab==='team'){initTeamDateFilter();loadPlayerList();return}
 if(tab==='fetch'){loadSummary()}
}

function refreshAfterFetchDataChanged(){
 loadSummary();
 page=1;
 loadMapFilter();
 initRecordsDateFilter();
 loadRecords();
 let tab=currentTab();
 if(tab!=='records'&&tab!=='fetch')refreshVisiblePanel();
}

document.querySelectorAll('.tab').forEach(t=>{
 t.onclick=()=>{
  if(!ensureSoftwareAccess('', '本地功能'))return;
  if(t.dataset.tab==='assets'&&!ensureSoftwareAccess('assets','账号资产'))return;
  if(t.dataset.tab==='assets'){document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));document.querySelectorAll('.panel').forEach(x=>x.classList.remove('active'));t.classList.add('active');$('panel-assets').classList.add('active');loadAssets();return}
  if(t.dataset.tab==='analysis'&&!ensureSoftwareAccess('analysis','数据分析'))return;
  if(t.dataset.tab==='trends'&&!ensureSoftwareAccess('trend_analysis','数据趋势'))return;
  if(t.dataset.tab==='team'&&!ensureSoftwareAccess('team_analysis','组队分析'))return;
  document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
  document.querySelectorAll('.panel').forEach(x=>x.classList.remove('active'));
  t.classList.add('active');
  $('panel-'+t.dataset.tab).classList.add('active');
  if(t.dataset.tab==='records'){loadMapFilter();initRecordsDateFilter();loadRecords()}
  if(t.dataset.tab==='items'){initItemsDateFilter();loadItems()}
  if(t.dataset.tab==='assets'){loadAssets()}
  if(t.dataset.tab==='analysis'){initAnalysisDateFilter();loadAnalysis()}
  if(t.dataset.tab==='trends'){loadTrendMapFilter();loadTrends()}
  if(t.dataset.tab==='team'){initTeamDateFilter();loadPlayerList()}
  if(t.dataset.tab==='fetch')loadSummary()
 }
});

function loadStatsLegacy(){
 fetch('/api/stats').then(r=>r.json()).then(d=>{
  let acc=d.active_account?' | 账号: '+d.active_account:'';
  $('headerStats').textContent='战绩 '+d.total_records+' 条 | 物品 '+d.total_items+' 件 | 撤离率 '+(d.total_records?((d.escaped/d.total_records*100).toFixed(1)):'0')+'%'+acc
 })
}

function loadMapFilterLegacy(){
 fetch('/api/maps').then(r=>r.json()).then(d=>{
  let sel=$('filterMap');
  if(!sel)return;
  d.forEach(m=>{let o=document.createElement('option');o.value=m.map_id;o.textContent=m.map_name;sel.appendChild(o)})
 })
}

function loadRecords(){
 psize=parseInt($('pageSize')?.value||20);
 let map=$('filterMap')?.value||'', result=$('filterResult')?.value??'';
 let sd=$('recStartDate')?.value||'',st=$('recStartTime')?.value||'00:00',ed=$('recEndDate')?.value||'',et=$('recEndTime')?.value||'23:59';
 let url='/api/records?page='+page+'&size='+psize;
 if(map)url+='&map_id='+map;if(result!=='')url+='&game_result='+result;
 if(sd)url+='&start='+encodeURIComponent(sd+' '+st);if(ed)url+='&end='+encodeURIComponent(ed+' '+et);
 fetch(url).then(r=>r.json()).then(d=>{
  total=d.total;page=d.page;
  $('pageInfo').textContent=page+' / '+Math.ceil(total/psize)+' ('+total+'条)';
  $('theadRec').innerHTML='<tr><th style="width:170px">对局编号</th><th style="width:132px">时间</th><th style="width:118px">地图</th><th style="width:96px">干员</th><th style="width:90px">结果</th><th style="width:86px">时长</th><th style="width:72px">击杀</th><th style="width:90px">玩家击杀</th><th class="money" style="width:118px">带入装备价值</th><th class="money" style="width:118px">带出价值</th><th class="money" style="width:118px">盈亏</th><th style="width:72px">物品数</th><th class="tags-cell" style="width:240px">对局评价</th></tr>';
  let rows=d.data.map(r=>'<tr onclick="showDetail(\''+r.room_id+'\')" title="点击查看对局详情"><td>'+r.room_id+'</td><td>'+r.event_time+'</td><td>'+(r.map_name||'-')+'</td><td>'+(r.role_name||'-')+'</td><td>'+resultBadge(r.game_result)+'</td><td>'+fmtDur(r.duration_s)+'</td><td>'+r.kill_cnt+'</td><td>'+r.kill_player+'</td><td class="money">'+fmtPrice(r.original_equipment_price)+'</td><td class="money">'+fmtPrice(r.gained_price)+'</td><td class="money">'+fmtMoneyWithTone(r.profit_loss)+'</td><td>'+r.item_count+'</td><td class="tags-cell">'+(renderTagList(r.tags)||'-')+'</td></tr>').join('');
  $('tbodyRec').innerHTML=rows||'<tr><td colspan="13" style="text-align:center;color:#555">暂无数据</td></tr>'
 })
}
function nextPage(){if(page<Math.ceil(total/psize)){page++;loadRecords()}}
function prevPage(){if(page>1){page--;loadRecords()}}

function loadItems(){
 let sd=$('itemStartDate')?.value||'',st=$('itemStartTime')?.value||'00:00',ed=$('itemEndDate')?.value||'',et=$('itemEndTime')?.value||'23:59';
 let url='/api/items';
 let qs=[];
 if(sd)qs.push('start='+encodeURIComponent(sd+' '+st));
 if(ed)qs.push('end='+encodeURIComponent(ed+' '+et));
 if(itemOwnerScope!=='all')qs.push('owner_scope='+encodeURIComponent(itemOwnerScope));
 if(qs.length){
  url+='?';
  url+=qs.join('&');
 }
 fetch(url).then(r=>r.json()).then(d=>{
  itemRows=Array.isArray(d)?d:[];
  renderItems();
 })
}

function itemOwnerScopeLabel(){
 if(itemOwnerScope==='self')return '自己';
 if(itemOwnerScope==='teammate')return '队友';
 return '全部';
}

function updateItemOwnerScopeButton(){
 let btn=$('btnItemOwnerScope');
 if(btn)btn.textContent='范围: '+itemOwnerScopeLabel();
}

function toggleItemOwnerScope(){
 itemOwnerScope=itemOwnerScope==='all'?'self':itemOwnerScope==='self'?'teammate':'all';
 updateItemOwnerScopeButton();
 loadItems();
}

function itemSortArrow(key){
 if(itemSort.key!==key)return '↕';
 return itemSort.order==='asc'?'↑':'↓';
}

function sortItems(key){
 if(itemSort.key===key)itemSort.order=itemSort.order==='asc'?'desc':'asc';
 else{itemSort.key=key;itemSort.order='desc'}
 renderItems();
}

function renderItems(){
  $('theadItem').innerHTML='<tr><th style="width:170px">对局编号</th><th style="width:132px;cursor:pointer;user-select:none" onclick="sortItems(\'event_time\')">时间 '+itemSortArrow('event_time')+'</th><th style="width:118px">地图</th><th style="width:120px">获取者</th><th style="width:120px">干员</th><th style="width:320px">物品名</th><th style="width:80px">品质</th><th style="width:70px">数量</th><th class="money" style="width:110px;cursor:pointer;user-select:none" onclick="sortItems(\'price\')">单价 '+itemSortArrow('price')+'</th></tr>';
 let rows=[...itemRows];
 rows.sort((a,b)=>{
  if(itemSort.key==='price'){
   let av=Number(a.price)||0,bv=Number(b.price)||0;
   return av-bv;
  }
  let at=String(a.event_time||''),bt=String(b.event_time||'');
  return at.localeCompare(bt);
 });
 if(itemSort.order==='desc')rows.reverse();
 let html=rows.map(r=>'<tr onclick="showDetail(\''+r.room_id+'\')" title="点击查看对局详情"><td style="width:170px">'+r.room_id+'</td><td style="width:132px">'+r.event_time+'</td><td style="width:118px">'+(r.map_name||'-')+'</td><td style="width:120px">'+(r.owner_name||'-')+'</td><td style="width:120px">'+(r.role_name||'-')+'</td><td style="width:320px;max-width:320px;overflow:hidden;text-overflow:ellipsis">'+(r.item_name||'-')+'</td><td style="width:80px">'+gradeLabel(r.grade)+'</td><td style="width:70px">'+r.num+'</td><td class="money" style="width:110px">'+fmtPrice(r.price)+'</td></tr>').join('');
 $('tbodyItem').innerHTML=html||'<tr><td colspan="9" style="text-align:center;color:#555">暂无数据</td></tr>';
}

function renderAssetCategoryTabs(categories){
 let wrap=$('assetCategoryTabs');
 if(!wrap)return;
 let list=Array.isArray(categories)&&categories.length?categories:[
  {code:'operator',name:'干员',count:0},
  {code:'gun',name:'枪械',count:0},
  {code:'dagger',name:'近战',count:0},
  {code:'vehicle',name:'载具',count:0},
  {code:'pendant',name:'挂饰',count:0}
 ];
 list=list.filter(opt=>String(opt.code||'').trim()!=='');
 if(!assetCurrentCategory&&list.length)assetCurrentCategory=String(list[0].code||'operator');
 wrap.innerHTML=list.map(opt=>'<button class="tab'+((opt.code||'')===assetCurrentCategory?' active':'')+'" type="button" onclick="setAssetCategory(\''+esc(opt.code||'')+'\')"><span class="asset-tab-name">'+esc(opt.name||'')+'</span><span class="asset-tab-count">('+fmtPrice(opt.count||0)+')</span></button>').join('');
}

function setAssetCategory(code){
 assetCurrentCategory=String(code||'operator');
 assetQuickCollectibleOnly=false;
 renderAssetCategoryTabs(assetFilters.categories||[]);
 loadAssets();
}

function renderAssetFilterOptions(filters){
 assetFilters=filters||{categories:[],grades:[]};
 renderAssetCategoryTabs(assetFilters.categories||[]);
 let grade=$('assetGrade');
 if(grade){
  let current=grade.value||'';
  grade.innerHTML=(assetFilters.grades||[]).map(opt=>'<option value="'+esc(opt.value||'')+'">'+esc(opt.label||'全部')+'</option>').join('')||'<option value=\"\">全部</option>';
  grade.value=current;
 }
}

function fmtAssetUpdatedAt(value){
 if(!value)return '数据更新时间：--';
 try{
  let normalized=String(value).trim().replace(' ','T');
  let d=new Date(normalized);
  if(!isNaN(d.getTime()))return '数据更新时间：'+d.toLocaleString();
 }catch(e){}
 return '数据更新时间：'+esc(String(value));
}

function renderAssetUpdatedAt(value){
 let el=$('assetUpdatedAt');
 if(!el)return;
 el.textContent=fmtAssetUpdatedAt(value||'');
}

function renderAssetSummary(summary){
 let leftCards=[
  {lbl:'典藏传说枪械',val:summary.collectible_guns||0,category:'gun',grade:'5',collectibleOnly:true},
  {lbl:'干员研究外观',val:summary.operator_count||0,category:'operator',grade:'6'},
  {lbl:'传说近战武器',val:summary.dagger_count||0,category:'dagger',grade:'5'}
 ];
 let rightCards=[
  {lbl:'资产总额',val:summary.total_price||0,compact:true},
  {lbl:'固定资产总额',val:summary.noncurrent_asset||0,compact:true},
  {lbl:'流动资产总额',val:summary.current_asset||0,compact:true},
  {lbl:'哈夫币余额',val:summary.hafcoinnum||0,compact:true}
 ];
 let renderCard=c=>{
  let raw=Number(c.val||0);
  let shown=c.compact?fmtCompactNumber(raw):fmtPrice(raw);
  let tip=c.compact?esc(fmtPrice(raw)+' 哈夫币'):'';
  let action=c.category?' class="card asset-summary-action" role="button" tabindex="0" onclick="applyAssetQuickFilter(\''+c.category+'\',\''+c.grade+'\')" onkeydown="if(event.key===\'Enter\'||event.key===\' \'){event.preventDefault();applyAssetQuickFilter(\''+c.category+'\',\''+c.grade+'\')}"':' class="card"';
  return '<div'+action+(tip?' data-tip="'+tip+'"':'')+' title="'+fmtPrice(raw)+'"><div class="val">'+shown+'</div><div class="lbl">'+c.lbl+'</div></div>';
 };
$('assetSummaryCards').innerHTML=
  '<div class="asset-summary-group left">'+leftCards.map(renderCard).join('')+'</div>'
  +'<div class="summary-separator" aria-hidden="true"><div class="summary-separator-line"></div></div>'
  +'<div class="asset-summary-group right">'+rightCards.map(renderCard).join('')+'</div>';
}

function applyAssetQuickFilter(category, grade){
 assetCurrentCategory=String(category||assetCurrentCategory||'operator');
 assetQuickCollectibleOnly=!!(category==='gun' && String(grade||'')==='5');
 renderAssetCategoryTabs(assetFilters.categories||[]);
 let gradeSel=$('assetGrade');
 if(gradeSel)gradeSel.value=String(grade||'');
 loadAssets();
}

function clearAssetFilters(){
 let gradeSel=$('assetGrade');
 if(gradeSel)gradeSel.value='';
 assetQuickCollectibleOnly=false;
 loadAssets();
}

function assetImageCell(row){
 let src=row.pre_pic||row.pic||'';
 if(!src)return '-';
 return '<img src="'+esc(src)+'" alt="'+esc(row.item_name||row.item_id||'asset')+'" style="width:250px;height:auto;display:block;object-fit:contain;border-radius:12px;background:rgba(255,255,255,.7);padding:6px;box-sizing:border-box">';
}

function assetTagList(row){
 let tags=[];
 let raw=String(row.tag_text||'').trim();
 if(raw){
  raw.split(/[\s,，、|/]+/).map(v=>String(v||'').trim()).filter(Boolean).forEach(v=>{
   if(!tags.includes(v))tags.push(v);
  });
 }
 if(Number(row.is_archive||0) && !tags.includes('隐藏款'))tags.push('隐藏款');
 if(!tags.length)return '-';
 return '<div class="asset-tag-list">'+tags.map(tag=>{
  let cls=tag==='联动'?'asset-tag-collab':(tag==='隐藏款'?'asset-tag-hidden':'asset-tag-hidden');
  return '<span class="asset-tag '+cls+'">'+esc(tag)+'</span>';
 }).join('')+'</div>';
}

const ASSET_GRADE_ORDER={'6':6,'5':5,'4':4,'3':3,'2':2,'1':1};
const ASSET_RARITY_ORDER={'极品':2,'优品':1,'--':0,'':0};
const ASSET_FINISH_ORDER={'极品':6,'优品':5,'S':4,'A':3,'B':2,'C':1,'':0,'--':0};

function assetCategoryTitle(code){
 if(code==='operator')return '干员皮肤';
 if(code==='gun')return '枪械外观';
 if(code==='dagger')return '近战外观';
 if(code==='vehicle')return '载具外观';
 if(code==='pendant')return '挂饰';
 return '名称';
}

function sortAssetRows(rows){
 return [...rows].sort((a,b)=>{
  let gd=(ASSET_GRADE_ORDER[String(b.grade||'')]||0)-(ASSET_GRADE_ORDER[String(a.grade||'')]||0);
  if(gd!==0)return gd;
  if(assetCurrentCategory==='gun'){
   let rd=(ASSET_RARITY_ORDER[String(b.collectible_rarity_text||'')]||0)-(ASSET_RARITY_ORDER[String(a.collectible_rarity_text||'')]||0);
   if(rd!==0)return rd;
   let fd=(ASSET_FINISH_ORDER[String(b.finish_grade||'')]||0)-(ASSET_FINISH_ORDER[String(a.finish_grade||'')]||0);
   if(fd!==0)return fd;
  }
  let cn=String(a.category_name||a.category_code||'').localeCompare(String(b.category_name||b.category_code||''),'zh-CN');
  if(cn!==0)return cn;
  let nm=String(a.item_name||'').localeCompare(String(b.item_name||''),'zh-CN');
  if(nm!==0)return nm;
  return String(a.item_id||'').localeCompare(String(b.item_id||''),'zh-CN');
 });
}

function renderAssets(){
 let headers=['<th style="width:72px">预览</th>','<th style="width:320px">'+assetCategoryTitle(assetCurrentCategory)+'</th>'];
 headers.push('<th style="width:80px">等级</th>');
 if(assetCurrentCategory==='gun'){
  headers.push('<th style="width:90px">典藏</th>');
  headers.push('<th style="width:100px">稀有度</th>');
  headers.push('<th style="width:100px">成色</th>');
 }
 $('theadAsset').innerHTML='<tr>'+headers.join('')+'</tr>';
 let rows=sortAssetRows(assetRows);
 let html=rows.map(r=>{
  let cols=['<td style="width:270px;min-width:270px">'+assetImageCell(r)+'</td>','<td style="min-width:460px">'+esc(r.item_name||'-')+'</td>'];
  cols.push('<td>'+gradeLabel(r.grade)+'</td>');
  if(assetCurrentCategory==='gun'){
   cols.push('<td>'+(Number(r.is_collectible_gun||0)?'是':'否')+'</td>');
   cols.push('<td>'+esc(r.collectible_rarity_text||'--')+'</td>');
   cols.push('<td>'+esc(r.finish_grade||'--')+'</td>');
  }
  return '<tr>'+cols.join('')+'</tr>';
 }).join('');
 $('tbodyAsset').innerHTML=html||'<tr><td colspan="'+headers.length+'" style="text-align:center;color:#555">暂无资产数据，请前往“数据抓取”页点击“刷新资产”获取。</td></tr>';
}

function loadAssets(){
 if(!ensureSoftwareAccess('assets','账号资产'))return;
 let grade=$('assetGrade')?.value||'';
 let url='/api/assets?category='+encodeURIComponent(assetCurrentCategory||'operator')+'&grade='+encodeURIComponent(grade)+'&collectible_only='+(assetQuickCollectibleOnly?'1':'0');
 fetch(url).then(r=>r.json()).then(d=>{
  assetRows=Array.isArray(d.rows)?d.rows:[];
  renderAssetFilterOptions(d.filters||{});
  renderAssetSummary(d.summary||{});
  renderAssetUpdatedAt(d.summary?.last_fetched_at||'');
  renderAssets();
 }).catch(e=>{
  showMsg('err','加载资产失败：'+translateErrorMessage(e));
 });
}

function refreshAssets(){
 if(!ensureSoftwareAccess('assets','账号资产'))return;
 let btn=$('btnAssetsSync');
 if(btn){btn.disabled=true;btn.textContent='刷新中...';}
 $('fetchLog').innerHTML='';
 appendFetchLog('开始刷新账号资产并同步物品列表');
  postJSON('/api/assets-refresh',{}).then(d=>{
   if(!d.ok)throw new Error(d.error||'刷新失败');
   let total=(d.total_entries??d.summary?.total_entries??0);
   let itemCount=fmtPrice(d.item_count||0);
   let catalogMsg=d.catalog_updated?'已同步最新物品列表':'物品列表已是最新';
   appendFetchLogs(d.logs);
   appendFetchLog('资产刷新完成：共 '+fmtPrice(total)+' 条，'+catalogMsg+'（'+itemCount+' 条）');
   showMsg('ok','资产刷新完成：共 '+fmtPrice(total)+' 条，'+catalogMsg+'（'+itemCount+' 条）');
   loadAssets();
  }).catch(e=>{
   notifyLoginExpired(e);
   appendFetchLog('刷新资产失败：'+translateErrorMessage(e));
   showMsg('err','刷新资产失败：'+translateErrorMessage(e));
  }).finally(()=>{
   if(btn){btn.disabled=false;btn.textContent='刷新资产';}
  });
}

function refreshAssetCatalog(){
 if(!ensureSoftwareAccess('assets','账号资产'))return;
 let btn=$('btnAssetsCatalog');
 if(btn){btn.disabled=true;btn.textContent='更新中...';}
 postJSON('/api/assets-catalog-refresh',{}).then(d=>{
  if(!d.ok)throw new Error(d.error||'更新失败');
  let msg=d.updated
   ? '物品列表已更新，共 '+fmtPrice(d.item_count||0)+' 条'
   : '物品列表已是最新，共 '+fmtPrice(d.item_count||0)+' 条';
  showMsg('ok',msg);
  loadAssets();
 }).catch(e=>{
  notifyLoginExpired(e);
  showMsg('err','更新物品列表失败：'+translateErrorMessage(e));
 }).finally(()=>{
  if(btn){btn.disabled=false;btn.textContent='更新物品列表';}
 });
}

function loadSummary(){
 fetch('/api/stats').then(r=>r.json()).then(d=>{
  let cards=[
   {lbl:'总对局',val:d.total_records||0},
   {lbl:'详情',val:d.detail_records||0},
   {lbl:'房间详情',val:d.room_detail_records||0}
   ];
  $('summaryCards').innerHTML=cards.map(c=>'<div class="card"><div class="val">'+c.val+'</div><div class="lbl">'+c.lbl+'</div></div>').join('')
 })
}

function loadMapFilter(){
 if(mapsLoaded)return;
  fetch('/api/maps').then(r=>r.json()).then(d=>{
  let sel=$('filterMap');
  if(!sel)return;
  sel.innerHTML='<option value="">全部</option>';
  d.forEach(m=>{let o=document.createElement('option');o.value=m.map_id;o.textContent=m.map_name;sel.appendChild(o)});
  mapsLoaded=true;
 })
}

function doFetch(){
 if(!ensureSoftwareAccess('', '本地功能'))return;
 let q=$('fetchQueue').value,n=parseInt($('fetchCount').value||'100',10);
 setButtonsDisabled(FETCH_ACTION_IDS,true);
 $('fetchLog').innerHTML='';
 resetFetchProgress();
 showMsg('info','抓取进行中...');
 appendFetchLog('开始抓取 '+(q==='sol'?'烽火地带':'全面战场')+'，数量 '+n);
 postJSON('/api/fetch',{queue:q,count:n})
 .then(d=>{
   if(!d.ok){
    setButtonsDisabled(FETCH_ACTION_IDS,false);
   notifyLoginExpired(d.error);
   let msg=translateErrorMessage(d.error||'未知错误');
   showMsg('err','抓取失败：'+msg);
   appendFetchLog('抓取失败：'+msg);
   noteFetchFailure(d.error||msg);
   return;
  }
   beginActionPolling(d.job_id,result=>{
    showMsg('ok','抓取完成: 新增 '+result.new_records+' 条，共 '+result.total_records+' 条');
    appendFetchLog('抓取完成: 新增 '+result.new_records+' 条，共 '+result.total_records+' 条');
    resetScopedState();
   syncHeaderState({closeGate:true}).then(()=>refreshAfterFetchDataChanged());
   });
 }).catch(e=>{let msg=translateErrorMessage(e);setButtonsDisabled(FETCH_ACTION_IDS,false);notifyLoginExpired(e);showMsg('err','请求失败：'+msg);appendFetchLog('抓取请求失败：'+msg);noteFetchFailure(e)})
}

function doSmartFetch(){
 if(!ensureSoftwareAccess('', '本地功能'))return;
 let q=$('fetchQueue').value,n=parseInt($('fetchCount').value||'100',10);
 setButtonsDisabled(FETCH_ACTION_IDS,true);
 $('fetchLog').innerHTML='';
 resetFetchProgress();
 showMsg('info','智能抓取进行中...');
 appendFetchLog('开始智能抓取 '+(q==='sol'?'烽火地带':'全面战场')+'，上限 '+n);
 postJSON('/api/fetch-smart',{queue:q,count:n})
 .then(d=>{
   if(!d.ok){
    setButtonsDisabled(FETCH_ACTION_IDS,false);
    notifyLoginExpired(d.error);
    let msg=translateErrorMessage(d.error||'未知错误');
    showMsg('err','智能抓取失败：'+msg);
    appendFetchLog('智能抓取失败：'+msg);
    noteFetchFailure(d.error||msg);
    return;
   }
   beginActionPolling(d.job_id,result=>{
    showMsg('ok','智能抓取完成: 新增 '+result.new_records+' 条，共 '+result.total_records+' 条');
    appendFetchLog('智能抓取完成: 新增 '+result.new_records+' 条，共 '+result.total_records+' 条');
    resetScopedState();
   syncHeaderState({closeGate:true}).then(()=>refreshAfterFetchDataChanged());
   });
 }).catch(e=>{let msg=translateErrorMessage(e);setButtonsDisabled(FETCH_ACTION_IDS,false);notifyLoginExpired(e);showMsg('err','智能抓取请求失败：'+msg);appendFetchLog('智能抓取请求失败：'+msg);noteFetchFailure(e)})
}

function doFetchDetails(){
 if(!ensureSoftwareAccess('', '本地功能'))return;
 setButtonsDisabled(FETCH_ACTION_IDS,true);
 resetFetchProgress();
 showMsg('info','正在补全对局详情...');
 appendFetchLog('开始补全缺失的对局详情');
 postJSON('/api/fetch-details',{})
 .then(d=>{
   if(!d.ok){
    setButtonsDisabled(FETCH_ACTION_IDS,false);
    notifyLoginExpired(d.error);
    let msg=translateErrorMessage(d.error||'未知错误');
    showMsg('err','补全失败：'+msg);
    appendFetchLog('补全失败：'+msg);
   noteFetchFailure(d.error||msg);
   return;
  }
 beginActionPolling(d.job_id,result=>{
  showMsg('ok','补全完成: 缺失 '+result.missing+' 条，获取 '+result.fetched_report+' 条详情 / '+result.fetched_room+' 条房间详情');
  appendFetchLog('补全完成: 缺失 '+result.missing+' 条，获取 '+result.fetched_report+' 条详情 / '+result.fetched_room+' 条房间详情');
  syncHeaderState({closeGate:true}).then(()=>refreshAfterFetchDataChanged());
  });
 }).catch(e=>{let msg=translateErrorMessage(e);setButtonsDisabled(FETCH_ACTION_IDS,false);notifyLoginExpired(e);showMsg('err','补全请求失败：'+msg);appendFetchLog('补全请求失败：'+msg);noteFetchFailure(e)})
}

function base64ToUint8Array(raw){
 let text=String(raw||'');
 let binary=atob(text);
 let out=new Uint8Array(binary.length);
 for(let i=0;i<binary.length;i++)out[i]=binary.charCodeAt(i);
 return out;
}

function exportBattleBackup(){
 let btn=$('btnExportBackup');
 if(btn){btn.disabled=true;btn.textContent='导出中...';}
 postJSON('/api/backup/export',{}).then(d=>{
  if(!d.ok)throw new Error(d.error||'导出失败');
  let bytes=base64ToUint8Array(d.content_base64||'');
  let blob=new Blob([bytes],{type:'application/zip'});
  let a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download=d.filename||'dfdc_backup.zip';
  document.body.appendChild(a);
  a.click();
  setTimeout(()=>{URL.revokeObjectURL(a.href);a.remove();},300);
  let summary=d.manifest&&d.manifest.database&&d.manifest.database.counts?d.manifest.database.counts:{};
  showMsg('ok','战绩备份已导出：对局 '+(summary.records||0)+' 条，报告 '+(((d.manifest||{}).reports||{}).count||0)+' 份');
 }).catch(e=>{
  alert('导出失败: '+normalizeErrorMessage(e));
 }).finally(()=>{
  if(btn){btn.disabled=false;btn.textContent='导出战绩备份';}
 });
}

function selectBattleBackupFile(){
 let input=$('battleBackupFile');
 if(!input)return;
 input.value='';
 input.click();
}

function handleBattleBackupSelected(event){
 let file=event&&event.target&&event.target.files&&event.target.files[0];
 if(!file)return;
 if(!confirm('导入只会恢复本地战绩数据库和 PDF 报告，不会恢复 WeGame 登录态。是否继续导入？')){
  event.target.value='';
  return;
 }
 let btn=$('btnImportBackup');
 if(btn){btn.disabled=true;btn.textContent='导入中...';}
 let reader=new FileReader();
 reader.onload=()=>{
  try{
   let raw=String(reader.result||'');
   let contentBase64=raw.includes(',')?raw.split(',').pop():'';
   postJSON('/api/backup/import',{filename:file.name,content_base64:contentBase64}).then(d=>{
    if(!d.ok)throw new Error(d.error||'导入失败');
    let msg='导入完成：新增对局 '+(d.records_added||0)+' 条，跳过重复 '+(d.records_skipped||0)+' 条，恢复报告 '+(d.reports_added||0)+' 份';
    if((d.reports_skipped||0)>0)msg+='，已跳过报告 '+d.reports_skipped+' 份';
    showMsg('ok',msg);
    syncHeaderState({closeGate:true}).then(()=>refreshAfterFetchDataChanged());
   }).catch(e=>{
    alert('导入失败: '+normalizeErrorMessage(e));
   }).finally(()=>{
    if(btn){btn.disabled=false;btn.textContent='导入战绩备份';}
    event.target.value='';
   });
  }catch(e){
   if(btn){btn.disabled=false;btn.textContent='导入战绩备份';}
   event.target.value='';
   alert('导入失败: '+normalizeErrorMessage(e));
  }
 };
 reader.onerror=()=>{
  if(btn){btn.disabled=false;btn.textContent='导入战绩备份';}
  event.target.value='';
  alert('导入失败: 无法读取备份文件');
 };
 reader.readAsDataURL(file);
}

function handleClearData(){
 if(!ensureSoftwareAccess('', '本地功能'))return;
 showActionOverlay({
  title:'确认清空当前账号数据',
  html:'<p>该操作只会删除当前选中账号的战绩、物品和详情数据，不会删除账号记录。</p><p>请输入 <span class=\"danger-keyword\">清空</span> 后再执行操作。</p><div style=\"margin-top:14px\"><input id=\"clearDataConfirmInput\" class=\"dialog-input\" type=\"text\" placeholder=\"请输入：清空\"></div>',
  actions:[
   {label:'取消',className:'btn btn-muted',onClick:'closeActionOverlay()'},
   {label:'确认清空',className:'btn btn-danger',onClick:'confirmClearData()'}
  ],
  closable:true
 });
 setTimeout(()=>{let input=$('clearDataConfirmInput');if(input)input.focus()},30);
}

function confirmClearData(){
 let input=$('clearDataConfirmInput');
 let value=input?input.value.trim():'';
 if(value!=='清空'){
  showMsg('err','请输入“清空”后再执行删除');
  if(input)input.focus();
  return;
 }
 closeActionOverlay();
 setButtonsDisabled(FETCH_ACTION_IDS,true);
 postJSON('/api/clear-data',{})
 .then(d=>{
  setButtonsDisabled(FETCH_ACTION_IDS,false);
   appendFetchLogs(d.logs);
   if(!d.ok){
    let msg=translateErrorMessage(d.error||'未知错误');
    showMsg('err','清空失败：'+msg);
    appendFetchLog('清空失败：'+msg);
   return;
  }
  showMsg('ok','已清空当前账号 '+d.cleared_records+' 条战绩和 '+d.cleared_items+' 件物品');
  appendFetchLog('已清空当前账号 '+d.cleared_records+' 条战绩和 '+d.cleared_items+' 件物品');
  resetScopedState();
  syncHeaderState({closeGate:true}).then(()=>{loadSummary();refreshVisiblePanel()});
 }).catch(e=>{let msg=translateErrorMessage(e);setButtonsDisabled(FETCH_ACTION_IDS,false);showMsg('err','清空请求失败：'+msg);appendFetchLog('清空请求失败：'+msg)})
}

function showDetail(roomId){
 $('modalTitle').textContent='对局详情 '+roomId;
 $('modalBody').innerHTML='<div style="text-align:center;padding:40px;color:#8899aa">加载中...</div>';
 $('modalOverlay').classList.add('show');
 fetch('/api/detail?room_id='+roomId).then(r=>r.json()).then(d=>{
  let html='';
  let self=d.self||{};
  let tm=d.teammates||[];
  let items=(d.items||[]).slice().sort((a,b)=>{
   let diff=(Number(b.price)||0)-(Number(a.price)||0);
   if(diff!==0)return diff;
   return (Number(b.num)||0)-(Number(a.num)||0);
  });
  let byPlayer={};
  let usedItemKeys={};

  function playerKey(p){
   if(p&&p.player_id!=null&&p.player_id!=='')return 'id:'+p.player_id;
   return 'name:'+((p&&p.player_name)||'未知');
  }
  items.forEach(it=>{
   let key=playerKey(it);
   if(!byPlayer[key])byPlayer[key]=[];
   byPlayer[key].push(it);
  });
  function renderItemsForPlayer(p,title){
   let key=playerKey(p);
   let list=byPlayer[key]||[];
   if(!list.length)return '';
   usedItemKeys[key]=true;
   let part='<div class="detail-player-items">';
   part+='<div class="detail-player-items-head"><div class="detail-player-items-title">'+title+'</div><div class="detail-player-items-count">'+list.length+' 件</div></div>';
   part+='<div class="item-grid">';
   list.forEach(it=>{
    let gradeColor=GRADE_COLOR[it.grade]||'#aaa';let gradeText=GRADE_MAP[it.grade]||'';
    part+='<div class="item-card">';
    if(it.pic)part+='<img src="'+it.pic+'" alt="" onerror="this.style.display=\'none\'">';
    part+='<div class="item-info"><div class="item-name" style="color:'+gradeColor+'">'+it.item_name+'</div><div class="item-meta">'+(gradeText?'<span style="color:'+gradeColor+'">'+gradeText+'</span> · ':'')+'x'+it.num+' · '+fmtPrice(it.price)+'</div></div></div>';
   });
   part+='</div>';
   part+='</div>';
   return part;
  }

  html+='<div class="detail-grid">';
  html+=di('时间',self.event_time||'-');
  html+=di('地图',self.map_name||'-');
  html+=di('模式',self.game_rule!=null?(self.game_rule===4?'烽火地带':'规则'+self.game_rule):'-');
  html+=di('对局类型',self.is_rank_match?'排位':'匹配');
  html+=di('游戏时长',fmtDur(self.duration_s));
  html+=di('队伍',self.team_id!=null?'#'+self.team_id:'-');
  html+='</div>';

  if(self.player_name!==undefined){
   html+='<div class="detail-player-block">';
   html+=renderDetailHead('我的数据',self.tags);
   html+='<div class="player-row is-self">';
   html+=di('昵称',self.player_name||'-');
   html+=di('干员',self.role_name||roleName(self.armed_force_id));
   html+=di('结果',resultBadge(self.game_result));
   html+=di('击杀玩家',self.kill_player!=null?self.kill_player:'-');
   html+=di('救援',self.rescue!=null?self.rescue:'-');
   html+=di('带出价值',fmtPrice(self.gained_price));
   html+=di('盈亏',fmtPrice(self.profit_loss),moneyToneStyle(self.profit_loss));
   html+=di('带入装备价值',fmtPrice(self.original_equipment_price));
   html+=di('曼德尔砖',self.has_blue_box?'有':'无');
   html+=di('中途退出',self.is_leave?'是':'否');
   html+='</div>';
   html+=renderItemsForPlayer(self,'带出高价值物品');
   html+='</div>';
  }


  if(tm.length>0){
   tm.forEach(t=>{
    html+='<div class="detail-player-block">';
    html+=renderDetailHead('队友 · '+(t.player_name||'-'),t.tags);
   html+='<div class="player-row">';
   html+=di('昵称',t.player_name||'-');
   html+=di('干员',t.role_name||roleName(t.armed_force_id));
   html+=di('结果',resultBadge(t.game_result));
   html+=di('击杀玩家',t.kill_player!=null?t.kill_player:'-');
   html+=di('救援',t.rescue!=null?t.rescue:'-');
   html+=di('带出价值',fmtPrice(t.gained_price));
   html+=di('盈亏',fmtPrice(t.profit_loss),moneyToneStyle(t.profit_loss));
   html+=di('带入装备价值',fmtPrice(t.original_equipment_price));
   html+=di('曼德尔砖',t.has_blue_box?'有':'无');
   html+=di('中途退出',t.is_leave?'是':'否');
   html+='</div>';
    html+=renderItemsForPlayer(t,'带出高价值物品');
    html+='</div>';
   });
  }

  let unmatchedKeys=Object.keys(byPlayer).filter(key=>!usedItemKeys[key]&&byPlayer[key]&&byPlayer[key].length);
  if(unmatchedKeys.length){
   html+='<div class="detail-player-block">';
   html+='<div class="section-title">未匹配玩家物品</div>';
   unmatchedKeys.forEach(key=>{
    let first=byPlayer[key][0]||{};
    let title=first.player_name||'未知玩家';
    html+=renderItemsForPlayer(first,title+' 的高价值物品');
   });
   html+='</div>';
  }

  if(!self.player_name && tm.length===0 && items.length===0){
   html+='<div style="text-align:center;padding:20px;color:#8899aa">暂无详情数据</div>';
  }

  $('modalBody').innerHTML=html;
 }).catch(e=>{
  $('modalBody').innerHTML='<div style="text-align:center;padding:20px;color:#ff6b4a">加载失败: '+e+'</div>';
 });
}

function closeModal(){$('modalOverlay').classList.remove('show');hideBattleTagTooltip()}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeModal()});

function di(label,val,extraStyle){return '<div class="detail-item"><div class="d-label">'+label+'</div><div class="d-val"'+(extraStyle?' style="'+extraStyle+'"':'')+'>'+val+'</div></div>';}

function loadAnalysis(){
 if(!ensureSoftwareAccess('analysis','数据分析'))return;
 let sd=$('anaStartDate')?.value||'',st=$('anaStartTime')?.value||'00:00',ed=$('anaEndDate')?.value||'',et=$('anaEndTime')?.value||'23:59';
 let url='/api/analysis';
 if(sd||ed){
  url+='?';
  if(sd)url+='start='+encodeURIComponent(sd+' '+st);
  if(sd&&ed)url+='&';
  if(ed)url+='end='+encodeURIComponent(ed+' '+et);
 }
 fetch(url).then(r=>r.json()).then(d=>{
  if(d&&d.ok===false){
   if(d.requires_member){showMemberFeaturePrompt('数据分析');return}
   if(d.requires_software_login){openSoftwareGate();return}
   $('analysisContent').innerHTML='<div style="color:#ff6b4a">'+esc(d.error||'加载失败')+'</div>';return
  }
  if(d.error){$('analysisContent').innerHTML='<div style="color:#ff6b4a">'+d.error+'</div>';return}
  let html='';

  let mapStats=[...(d.map_stats||[])];
  let mapStatsByRate=[...mapStats].sort((a,b)=>(Number(b.evac_rate_num)||0)-(Number(a.evac_rate_num)||0)||(Number(b.count)||0)-(Number(a.count)||0));
  let mapStatsByCount=[...mapStats].sort((a,b)=>(Number(b.count)||0)-(Number(a.count)||0)||(Number(b.evac_rate_num)||0)-(Number(a.evac_rate_num)||0));
  let roleStats=[...(d.role_stats||[])];
  let roleStatsByRate=[...roleStats].sort((a,b)=>(Number(b.evac_rate_num)||0)-(Number(a.evac_rate_num)||0)||(Number(b.count)||0)-(Number(a.count)||0));
  let roleStatsByCount=[...roleStats].sort((a,b)=>(Number(b.count)||0)-(Number(a.count)||0)||(Number(b.evac_rate_num)||0)-(Number(a.evac_rate_num)||0));

  html+='<div class="summary analysis-summary">';
   let cards=[
    {lbl:'总对局',val:d.total},{lbl:'撤离成功',val:d.escaped},
    {lbl:'撤离率',val:d.evac_rate},{lbl:'KD',val:d.kd},
    {lbl:'玩家击杀',val:d.total_player_kills},{lbl:'总盈亏',val:fmtPrice(d.total_profit),key:'total_profit',num:Number(d.total_profit)||0},
    {lbl:'平均盈亏',val:fmtPrice(Math.round(d.avg_profit))},{lbl:'平均时长',val:fmtDur(Math.round(d.avg_duration))},
    {lbl:'最高带出',val:fmtPrice(d.max_collection)},{lbl:'最高盈利',val:fmtPrice(d.max_profit)}
   ];
  html+=cards.map(c=>{
   let cls='val';
   if(c.key==='total_profit'){
    if(c.num>0)cls+=' profit-pos';
    else if(c.num<0)cls+=' profit-neg';
   }
   return '<div class="card"><div class="'+cls+'">'+c.val+'</div><div class="lbl">'+c.lbl+'</div></div>';
  }).join('');
  html+='</div>';

   if(mapStats.length){
    html+='<div class="analysis-section"><div class="analysis-section-head"><h3>地图分布</h3><div class="analysis-note">按当前筛选范围统计</div></div>';
    html+='<div class="chart-row"><div class="chart-box"><h4>对局次数</h4><div class="bar-chart">';
   let maxCnt=Math.max(...mapStatsByCount.map(m=>m.count));
   mapStatsByCount.forEach(m=>{
    let pct=maxCnt?Math.round(m.count/maxCnt*100):0;
    html+='<div class="bar-row"><div class="bar-label">'+(m.map_name||m.map_id)+'</div>'+renderBarTrack(pct,'#00d4aa',String(m.count))+'</div>';
   });
    html+='</div></div>';
    html+='<div class="chart-box"><h4>撤离率</h4><div class="bar-chart">';
   mapStatsByRate.forEach(m=>{
     let rate=m.evac_rate_num;let pct=Math.round(rate);
     let color=rate<40?'#ff6b4a':rate<60?'#ffcc4a':'#00ff88';
    html+='<div class="bar-row"><div class="bar-label">'+(m.map_name||m.map_id)+'</div>'+renderBarTrack(pct,color,m.evac_rate)+'</div>';
   });
    html+='</div></div></div>';
    html+='<table class="heat-table"><thead><tr><th>地图</th><th class="num">对局</th><th class="num">撤离</th><th class="num">撤离率</th><th class="num">平均玩家击杀</th><th class="num">平均盈亏</th><th class="num">平均时长</th></tr></thead><tbody>';
   mapStatsByRate.forEach(m=>{
    html+='<tr><td>'+(m.map_name||m.map_id)+'</td><td class="num">'+m.count+'</td><td class="num">'+m.escaped+'</td><td class="num">'+m.evac_rate+'</td><td class="num">'+m.avg_player_kills+'</td><td class="num">'+fmtPrice(m.avg_profit)+'</td><td class="num">'+fmtDur(m.avg_duration)+'</td></tr>';
   });
   html+='</tbody></table></div>';
  }

   if(roleStats.length){
    html+='<div class="analysis-section"><div class="analysis-section-head"><h3>干员使用</h3><div class="analysis-note">按当前筛选范围统计</div></div>';
    html+='<div class="chart-row"><div class="chart-box"><h4>使用次数</h4><div class="bar-chart">';
   let maxR=Math.max(...roleStatsByCount.map(r=>r.count));
   roleStatsByCount.forEach(r=>{
    let pct=maxR?Math.round(r.count/maxR*100):0;
    html+='<div class="bar-row"><div class="bar-label">'+(r.role_name||r.role_id)+'</div>'+renderBarTrack(pct,'#4a9eff',String(r.count))+'</div>';
   });
    html+='</div></div>';
    html+='<div class="chart-box"><h4>撤离率</h4><div class="bar-chart">';
   roleStatsByRate.forEach(r=>{
     let rate=r.evac_rate_num;let pct=Math.round(rate);
     let color=rate<40?'#ff6b4a':rate<60?'#ffcc4a':'#00ff88';
    html+='<div class="bar-row"><div class="bar-label">'+(r.role_name||r.role_id)+'</div>'+renderBarTrack(pct,color,r.evac_rate)+'</div>';
   });
    html+='</div></div></div>';
    html+='<table class="heat-table"><thead><tr><th>干员</th><th class="num">场次</th><th class="num">撤离率</th><th class="num">平均玩家击杀</th><th class="num">平均盈亏</th></tr></thead><tbody>';
   roleStatsByRate.forEach(r=>{
    html+='<tr><td>'+(r.role_name||r.role_id)+'</td><td class="num">'+r.count+'</td><td class="num">'+r.evac_rate+'</td><td class="num">'+r.avg_player_kills+'</td><td class="num">'+fmtPrice(r.avg_profit)+'</td></tr>';
   });
   html+='</tbody></table></div>';
  }

   if(d.result_dist){
    html+='<div class="analysis-section"><div class="analysis-section-head"><h3>结果分布</h3><div class="analysis-note">按当前筛选范围统计</div></div><div class="chart-row">';
    html+='<div class="chart-box"><h4>结果占比</h4><div class="bar-chart">';
    let rd=d.result_dist;let maxRd=Math.max(rd.success,rd.fail,rd.timeout,rd.leave,1);
    [['success','撤离成功','#00ff88'],['fail','撤离失败','#ff6b4a'],['timeout','行动超时','#ffcc4a'],['leave','中途退出','#888']].forEach(([k,l,c])=>{
    let pct=maxRd?Math.round(rd[k]/maxRd*100):0;
    html+='<div class="bar-row"><div class="bar-label">'+l+'</div>'+renderBarTrack(pct,c,String(rd[k]))+'</div>';
   });
   html+='</div></div></div></div>';
  }

   if(d.top_games&&d.top_games.length){
    html+='<div class="analysis-section"><div class="analysis-section-head"><h3>高光时刻</h3><div class="analysis-note">按盈亏排序的前 10 局</div></div>';
    html+='<table class="heat-table"><thead><tr><th>时间</th><th>地图</th><th>干员</th><th>结果</th><th class="num">玩家击杀</th><th class="num">带出价值</th><th class="num">盈亏</th></tr></thead><tbody>';
   d.top_games.forEach(g=>{
    html+='<tr onclick="showDetail(\''+g.room_id+'\')" title="点击查看对局详情" style="cursor:pointer"><td>'+g.event_time+'</td><td>'+(g.map_name||'-')+'</td><td>'+(g.role_name||'-')+'</td><td>'+resultBadge(g.game_result)+'</td><td class="num">'+g.kill_player+'</td><td class="num">'+fmtPrice(g.gained_price)+'</td><td class="num">'+fmtPrice(g.profit_loss)+'</td></tr>';
   });
   html+='</tbody></table></div>';
  }

  $('analysisContent').innerHTML=html;
 })
}

function trendRangeLabel(value){
 if(value==='7')return '近七天';
 if(value==='30')return '近30天';
 if(value==='90')return '近90天';
 return '全部';
}

function updateTrendRangeButtons(){
 let ids={'7':'btnTrendRange7','30':'btnTrendRange30','90':'btnTrendRange90','all':'btnTrendRangeAll'};
 Object.entries(ids).forEach(([value,id])=>{
  let btn=$(id);
  if(btn)btn.classList.toggle('is-active',trendRange===value);
 });
}

function setTrendRange(value){
 trendRange=String(value||'30');
 updateTrendRangeButtons();
 loadTrends();
}

function trendMapDifficulty(name){
 let parts=String(name||'').split('-').map(v=>v.trim()).filter(Boolean);
 return parts.length>1?parts.slice(1).join('-'):'未标注';
}

function loadTrendMapFilter(){
 if(trendMapsLoaded)return;
 fetch('/api/maps').then(r=>r.json()).then(d=>{
  let sel=$('trendMap');
  if(!sel)return;
  let rows=Array.isArray(d)?d:[];
  let counts={};
  rows.forEach(m=>{let name=String(m.map_name||m.map_id||'');counts[name]=(counts[name]||0)+1});
  let groups={};
  rows.forEach(m=>{
   let name=String(m.map_name||m.map_id||'未知地图');
   let diff=trendMapDifficulty(name);
   if(!groups[diff])groups[diff]=[];
   groups[diff].push(m);
  });
  let order=['常规','普通','机密','绝密','永夜','水淹','未标注'];
  let keys=Object.keys(groups).sort((a,b)=>{
   let ai=order.indexOf(a),bi=order.indexOf(b);
   if(ai<0)ai=999;if(bi<0)bi=999;
   return ai-bi||a.localeCompare(b,'zh-CN');
  });
  let html='<option value="">全部地图</option>';
  keys.forEach(key=>{
   html+='<optgroup label="'+esc(key)+'">';
   groups[key].sort((a,b)=>String(a.map_name||'').localeCompare(String(b.map_name||''),'zh-CN')).forEach(m=>{
    let name=String(m.map_name||m.map_id||'');
    let label=counts[name]>1?name+' #'+m.map_id:name;
    html+='<option value="'+esc(m.map_id)+'">'+esc(label)+'</option>';
   });
   html+='</optgroup>';
  });
  sel.innerHTML=html;
  trendMapsLoaded=true;
 });
}

function trendMetricFormat(metric,value){
 let n=Number(value)||0;
 if(metric.format==='percent')return n.toFixed(1).replace(/\.0$/,'')+'%';
 if(metric.format==='decimal')return n.toFixed(2).replace(/\.00$/,'');
 if(metric.format==='money')return fmtPrice(Math.round(n));
 if(metric.format==='duration')return fmtDur(Math.round(n));
 return fmtPrice(Math.round(n));
}

function trendAxisFormat(metric,value){
 let n=Number(value)||0;
 if(metric.format==='percent')return n.toFixed(0)+'%';
 if(metric.format==='decimal')return n.toFixed(1);
 if(metric.format==='duration')return Math.round(n/60)+'分';
 if(metric.format==='money')return fmtCompactNumber(Math.round(n));
 return fmtCompactNumber(Math.round(n));
}

function renderTrendSummary(summary){
 let cards=TREND_METRICS.map(metric=>{
  let val=summary&&summary[metric.key]!=null?summary[metric.key]:0;
  let cls='val';
  if(metric.format==='money'){
   let n=Number(val)||0;
   if(n>0)cls+=' profit-pos';
   else if(n<0)cls+=' profit-neg';
  }
  let key=esc(metric.key);
  return '<a class="card" href="#trend-'+key+'" data-trend-key="'+key+'" onclick="scrollTrendMetric(\''+key+'\')" onkeydown="handleTrendCardKey(event,\''+key+'\')"><div class="'+cls+'">'+trendMetricFormat(metric,val)+'</div><div class="lbl">'+metric.label+'</div></a>';
 }).join('');
 let wrap=$('trendSummaryCards');
 if(wrap){
  wrap.innerHTML=cards;
  bindTrendSummaryCards();
 }
}

function scrollTrendMetric(key){
 let el=$('trend-'+key);
 if(!el)return;
 let sticky=document.querySelector('.trend-sticky');
 let offset=96;
 if(sticky)offset=Math.min(Math.max(sticky.getBoundingClientRect().height+36,offset),560);
 let top=el.getBoundingClientRect().top+window.pageYOffset-offset;
 window.scrollTo({top:Math.max(0,top),behavior:'smooth'});
}
function handleTrendCardKey(event,key){
 if(event.key!=='Enter'&&event.key!==' ')return;
 event.preventDefault();
 event.stopPropagation();
 scrollTrendMetric(key);
}
function bindTrendSummaryCards(){
 let wrap=$('trendSummaryCards');
 if(!wrap)return;
 wrap.querySelectorAll('.card[data-trend-key]').forEach(card=>{
  if(card.dataset.trendJumpBound==='1')return;
  card.dataset.trendJumpBound='1';
  card.addEventListener('click',event=>{
   scrollTrendMetric(card.getAttribute('data-trend-key')||'');
  });
  card.addEventListener('keydown',event=>handleTrendCardKey(event,card.getAttribute('data-trend-key')||''));
 });
}

document.addEventListener('click',e=>{
 let card=e.target.closest('#trendSummaryCards .card[data-trend-key]');
 if(!card)return;
 scrollTrendMetric(card.getAttribute('data-trend-key')||'');
});
document.addEventListener('keydown',e=>{
 if(e.key!=='Enter'&&e.key!==' ')return;
 let card=e.target.closest('#trendSummaryCards .card[data-trend-key]');
 if(!card)return;
 e.preventDefault();
 scrollTrendMetric(card.getAttribute('data-trend-key')||'');
});

function renderTrendChartStats(metric,rows,values){
 if(!rows.length)return '';
 let latest=values[values.length-1];
 let rawMin=Math.min(...values);
 let rawMax=Math.max(...values);
 let avg=values.reduce((sum,v)=>sum+v,0)/values.length;
 let maxIndex=values.indexOf(rawMax);
 let minIndex=values.indexOf(rawMin);
 let items=[
  {k:'最新',v:trendMetricFormat(metric,latest),sub:rows[rows.length-1].label},
  {k:'峰值',v:trendMetricFormat(metric,rawMax),sub:rows[maxIndex]?.label||''},
  {k:'低值',v:trendMetricFormat(metric,rawMin),sub:rows[minIndex]?.label||''},
  {k:'平均',v:trendMetricFormat(metric,avg),sub:rows.length+' 个点'}
 ];
 return '<div class="trend-chart-stats">'+items.map(item=>'<div class="trend-chart-stat"><div class="k">'+esc(item.k)+' · '+esc(item.sub)+'</div><div class="v">'+esc(item.v)+'</div></div>').join('')+'</div>';
}

function renderTrendDataStrip(metric,rows,values){
 if(!rows.length)return '';
 let start=Math.max(0,rows.length-16);
 let chips=rows.slice(start).map((row,idx)=>{
  let actual=start+idx;
  return '<div class="trend-data-chip"><div class="date">'+esc(row.label||row.bucket||'')+'</div><div class="value">'+esc(trendMetricFormat(metric,values[actual]))+'</div></div>';
 }).join('');
 return '<div class="trend-data-strip">'+chips+'</div>';
}

function renderTrendLineChart(metric,rows){
 rows=Array.isArray(rows)?rows:[];
 if(!rows.length)return '<div class="trend-empty">当前筛选范围暂无趋势数据</div>';
 let values=rows.map(r=>Number(r[metric.key])||0);
 let rawMin=Math.min(...values),rawMax=Math.max(...values);
 let min=rawMin,max=rawMax;
 if(min===max){
  let delta=Math.max(Math.abs(max)*0.12,1);
  min-=delta;max+=delta;
 }
 let width=Math.max(760,rows.length*58+100),height=260,padL=62,padR=24,padT=24,padB=44;
 let innerW=width-padL-padR,innerH=height-padT-padB;
 let xAt=i=>padL+(rows.length===1?innerW/2:(i/(rows.length-1))*innerW);
 let yAt=v=>padT+((max-v)/(max-min))*innerH;
 let points=values.map((v,i)=>xAt(i).toFixed(1)+','+yAt(v).toFixed(1)).join(' ');
 let bottom=padT+innerH;
 let area=rows.length?xAt(0).toFixed(1)+','+bottom.toFixed(1)+' '+points+' '+xAt(rows.length-1).toFixed(1)+','+bottom.toFixed(1):'';
 let svg='<svg class="trend-svg" style="width:'+width+'px" viewBox="0 0 '+width+' '+height+'" preserveAspectRatio="none">';
 for(let i=0;i<=4;i++){
  let v=min+(max-min)*(i/4);
  let y=yAt(v);
  svg+='<line class="trend-grid" x1="'+padL+'" y1="'+y.toFixed(1)+'" x2="'+(width-padR)+'" y2="'+y.toFixed(1)+'"></line>';
  svg+='<text class="trend-value-label" x="8" y="'+(y+4).toFixed(1)+'">'+esc(trendAxisFormat(metric,v))+'</text>';
 }
 if(min<0&&max>0){
  let zy=yAt(0);
  svg+='<line class="trend-axis" x1="'+padL+'" y1="'+zy.toFixed(1)+'" x2="'+(width-padR)+'" y2="'+zy.toFixed(1)+'"></line>';
 }
 svg+='<line class="trend-axis" x1="'+padL+'" y1="'+bottom+'" x2="'+(width-padR)+'" y2="'+bottom+'"></line>';
 svg+='<polygon class="trend-area" points="'+area+'"></polygon>';
 svg+='<polyline class="trend-line" points="'+points+'"></polyline>';
 let labelIndexes=new Set([0,rows.length-1,values.indexOf(rawMin),values.indexOf(rawMax)]);
 if(rows.length<=14)values.forEach((_,i)=>labelIndexes.add(i));
 else{
  let labelStep=Math.max(1,Math.ceil(rows.length/5));
  values.forEach((_,i)=>{if(i%labelStep===0)labelIndexes.add(i)});
 }
 values.forEach((v,i)=>{
  let cx=xAt(i),cy=yAt(v);
  svg+='<circle class="trend-dot" cx="'+cx.toFixed(1)+'" cy="'+cy.toFixed(1)+'" r="4"><title>'+esc(rows[i].label+' · '+trendMetricFormat(metric,v))+'</title></circle>';
  if(labelIndexes.has(i)){
   let labelY=cy<padT+16?cy+18:cy-8;
   svg+='<text class="trend-dot-label" x="'+cx.toFixed(1)+'" y="'+labelY.toFixed(1)+'" text-anchor="middle">'+esc(trendMetricFormat(metric,v))+'</text>';
  }
 });
 let step=Math.max(1,Math.ceil(rows.length/8));
 rows.forEach((r,i)=>{
  if(i%step!==0&&i!==rows.length-1)return;
  svg+='<text class="trend-axis-label" x="'+xAt(i).toFixed(1)+'" y="'+(height-12)+'" text-anchor="middle">'+esc(r.label)+'</text>';
 });
 svg+='</svg>';
 return renderTrendChartStats(metric,rows,values)+'<div class="trend-chart-body">'+svg+'</div>'+renderTrendDataStrip(metric,rows,values);
}

function renderTrendCharts(rows){
 let html=TREND_METRICS.map(metric=>{
  return '<div class="analysis-section trend-chart" id="trend-'+metric.key+'"><div class="analysis-section-head"><h3>'+metric.label+'趋势</h3><div class="analysis-note">点击上方指标卡可定位到这里</div></div>'+renderTrendLineChart(metric,rows)+'</div>';
 }).join('');
 let wrap=$('trendContent');
 if(wrap)wrap.innerHTML=html;
}

function loadTrends(){
 if(!ensureSoftwareAccess('trend_analysis','数据趋势'))return;
 updateTrendRangeButtons();
 let bucket=$('trendBucket')?.value||'day';
 let map=$('trendMap')?.value||'';
 let url='/api/trends?range='+encodeURIComponent(trendRange)+'&bucket='+encodeURIComponent(bucket);
 if(map)url+='&map_id='+encodeURIComponent(map);
 fetch(url).then(r=>r.json()).then(d=>{
  if(d&&d.ok===false){
   if(d.requires_member){showMemberFeaturePrompt('数据趋势');return}
   if(d.requires_software_login){openSoftwareGate();return}
   $('trendContent').innerHTML='<div style="color:#ff6b4a">'+esc(d.error||'加载失败')+'</div>';return
  }
  if(d.error){$('trendContent').innerHTML='<div style="color:#ff6b4a">'+esc(d.error)+'</div>';return}
  let range=d.range||{};
  let bucketLabel=bucket==='week'?'按周':bucket==='month'?'按月':'按日';
  let meta=$('trendResolvedRange');
  if(meta)meta.textContent='当前范围：'+trendRangeLabel(trendRange)+' · '+bucketLabel+' · '+(range.start||'--')+' 至 '+(range.end||'--');
  renderTrendSummary(d.summary||{});
  renderTrendCharts(d.rows||[]);
 });
}

const GC_LIGHT=['#ffe0e0','#fff0d0','#ffffd0','#d0f0d0','#b0e8b0'];
function gc(val,min,max,rev){
 if(min===max)return 'background:'+GC_LIGHT[2]+';color:#1a1a2a';
 let r=(val-min)/(max-min);if(rev)r=1-r;
 let i=r<=0.2?0:r<=0.4?1:r<=0.6?2:r<=0.8?3:4;
 return 'background:'+GC_LIGHT[i]+';color:#1a1a2a';
}
function gcSigned(val,min,max){
 val=Number(val)||0; min=Number(min)||0; max=Number(max)||0;
 if(val===0 || (min===0 && max===0))return 'background:'+GC_LIGHT[2]+';color:#1a1a2a';
 if(val>0){
  let posMax=Math.max(0,max);
  if(posMax<=0)return 'background:'+GC_LIGHT[2]+';color:#1a1a2a';
  let r=Math.max(0,Math.min(1,val/posMax));
  let i=r<=0.33?2:r<=0.66?3:4;
  return 'background:'+GC_LIGHT[i]+';color:#1a1a2a';
 }
 let negMin=Math.min(0,min);
 if(negMin>=0)return 'background:'+GC_LIGHT[2]+';color:#1a1a2a';
 let r=Math.max(0,Math.min(1,Math.abs(val/negMin)));
 let i=r<=0.33?2:r<=0.66?1:0;
 return 'background:'+GC_LIGHT[i]+';color:#1a1a2a';
}

function fmtDateInput(dt){
 let y=dt.getFullYear(),m=String(dt.getMonth()+1).padStart(2,'0'),d=String(dt.getDate()).padStart(2,'0');
 return y+'-'+m+'-'+d;
}

function setDateField(inputId,btnId,val){
 let inp=$(inputId),btn=$(btnId);
 if(inp)inp.value=val;
  if(btn)btn.textContent=val||'选择日期';
}

function setDateRange(startInputId,startBtnId,endInputId,endBtnId,startVal,endVal){
 setDateField(startInputId,startBtnId,startVal);
 setDateField(endInputId,endBtnId,endVal);
}
function updateAnalysisPresetButtons(){
 let ids={today:'btnAnaPresetToday',yesterday:'btnAnaPresetYesterday',3:'btnAnaPreset3',7:'btnAnaPreset7',30:'btnAnaPreset30'};
 Object.entries(ids).forEach(([mode,id])=>{
  let btn=$(id);
  if(btn)btn.classList.toggle('is-active',String(analysisPresetDays)===String(mode));
 });
}
function updateTeamPresetButtons(){
 let ids={today:'btnTeamPresetToday',yesterday:'btnTeamPresetYesterday',daybefore:'btnTeamPresetDayBefore'};
 Object.entries(ids).forEach(([mode,id])=>{
  let btn=$(id);
  if(btn)btn.classList.toggle('is-active',teamPresetMode===mode);
 });
}
function teamPlayerKey(p){
 if(!p)return'';
 if(p.key)return String(p.key);
 if(p.player_key)return String(p.player_key);
 if(p.player_id!=null&&p.player_id!=='')return 'id:'+String(p.player_id);
 return 'name:'+String(p.name||p.player_name||'');
}
function shortTeamPlayerId(id){
 let s=String(id||'');
 return s.length>6?s.slice(-6):s;
}
function teamPlayerDisplay(p){
 let name=String((p&& (p.name||p.player_name))||'');
 if(p&&p.duplicate_name&&p.player_id)return name+' · ID '+shortTeamPlayerId(p.player_id);
 return name;
}
function escTeamArg(v){
 return String(v||'').replace(/\\/g,'\\\\').replace(/'/g,"\\'");
}
function getTeamSelectedNames(){
 return selectedPlayers.filter(key=>key!==TEAM_OTHER_TOKEN);
}
function hasTeamOtherSelection(){
 return selectedPlayers.includes(TEAM_OTHER_TOKEN);
}
function getTeamFilterQueryParts(){
 let parts=[];
 let sd=$('teamStartDate')?.value||'',st=$('teamStartTime')?.value||'00:00',ed=$('teamEndDate')?.value||'',et=$('teamEndTime')?.value||'23:59';
 if(sd)parts.push('start='+encodeURIComponent(sd+' '+st));
 if(ed)parts.push('end='+encodeURIComponent(ed+' '+et));
 return parts;
}
function normalizeTeamSelections(availablePlayers,activeKey,activeName){
 let keys=new Set((availablePlayers||[]).map(p=>teamPlayerKey(p)).filter(Boolean));
 let real=getTeamSelectedNames().filter(key=>keys.has(key));
 if(activeKey){
  teamActivePlayerKey=activeKey;
  teamActivePlayerName=activeName;
  if(keys.has(activeKey)&&!real.includes(activeKey)&&!teamSelfDeselected)real.unshift(activeKey);
 }
 real=[...new Set(real)];
 if(teamActivePlayerKey&&real.includes(teamActivePlayerKey)){
  real=[teamActivePlayerKey].concat(real.filter(key=>key!==teamActivePlayerKey));
 }
 if(hasTeamOtherSelection()&&real.length===2)selectedPlayers=real.concat(TEAM_OTHER_TOKEN);
 else selectedPlayers=real.slice(0,3);
}
function handleTeamFilterChanged(){
 teamPresetMode='';
 updateTeamPresetButtons();
 loadPlayerList();
}

function clearDateRange(scope){
 if(scope==='records'){
  setDateRange('recStartDate','btnRecStart','recEndDate','btnRecEnd','','');
  if($('recStartTime'))$('recStartTime').value='00:00';
  if($('recEndTime'))$('recEndTime').value='23:59';
  page=1;
  loadRecords();
  return;
 }
 if(scope==='items'){
  setDateRange('itemStartDate','btnItemStart','itemEndDate','btnItemEnd','','');
  if($('itemStartTime'))$('itemStartTime').value='00:00';
  if($('itemEndTime'))$('itemEndTime').value='23:59';
  loadItems();
  return;
 }
 if(scope==='analysis'){
  analysisPresetDays=0;
  updateAnalysisPresetButtons();
  setDateRange('anaStartDate','btnAnaStart','anaEndDate','btnAnaEnd','','');
  if($('anaStartTime'))$('anaStartTime').value='00:00';
  if($('anaEndTime'))$('anaEndTime').value='23:59';
  loadAnalysis();
  return;
 }
 if(scope==='team'){
  teamPresetMode='';
  updateTeamPresetButtons();
  setDateRange('teamStartDate','btnStartDate','teamEndDate','btnEndDate','','');
  if($('teamStartTime'))$('teamStartTime').value='00:00';
  if($('teamEndTime'))$('teamEndTime').value='23:59';
  loadPlayerList();
 }
}

teamDateInited=teamDateInited||false;
const DATE_BTN_MAP={
 'recStartDate':'btnRecStart','recEndDate':'btnRecEnd',
 'itemStartDate':'btnItemStart','itemEndDate':'btnItemEnd',
 'anaStartDate':'btnAnaStart','anaEndDate':'btnAnaEnd',
 'teamStartDate':'btnStartDate','teamEndDate':'btnEndDate'
};
function restoreDateInputAnchor(inp){
 if(!inp)return;
 inp.style.position='absolute';
 inp.style.left='';
 inp.style.top='';
 inp.style.width='';
 inp.style.height='';
 inp.style.margin='';
 inp.style.opacity='0';
 inp.style.pointerEvents='none';
 inp.style.zIndex='';
 inp.style.border='';
 inp.style.background='';
}
function pickDate(inputId){
 let inp=$(inputId);
 if(!inp)return;
 let btn=$(DATE_BTN_MAP[inputId]||('btn'+inputId));
 if(btn){
  let rect=btn.getBoundingClientRect();
  inp.style.position='fixed';
  inp.style.left=Math.max(8,Math.floor(rect.left))+'px';
  inp.style.top=Math.max(8,Math.floor(rect.top))+'px';
  inp.style.width=Math.max(44,Math.ceil(rect.width))+'px';
  inp.style.height=Math.max(36,Math.ceil(rect.height))+'px';
  inp.style.margin='0';
  inp.style.border='0';
  inp.style.background='transparent';
 }
 inp.style.opacity='0.01';
 inp.style.pointerEvents='auto';
 inp.style.zIndex='9999';
 inp.showPicker?inp.showPicker():inp.click();
 inp.addEventListener('change',function h(){
  if(btn)btn.textContent=inp.value||'选择日期';
  if(inputId==='anaStartDate'||inputId==='anaEndDate'){
   analysisPresetDays=0;
   updateAnalysisPresetButtons();
  }
  inp.removeEventListener('change',h);
  restoreDateInputAnchor(inp);
 },{once:true});
 inp.addEventListener('blur',function b(){
  restoreDateInputAnchor(inp);
  inp.removeEventListener('blur',b);
 },{once:true});
}

function setAnalysisPreset(days){
 analysisPresetDays=days;
 updateAnalysisPresetButtons();
 let end=new Date();
 let start=new Date(end);
 if(days==='yesterday'){
  start.setDate(start.getDate()-1);
  end.setDate(end.getDate()-1);
 }else if(days!=='today'){
  start.setDate(end.getDate()-(Number(days)-1));
 }
 let s=fmtDateInput(start),e=fmtDateInput(end);
 setDateRange('anaStartDate','btnAnaStart','anaEndDate','btnAnaEnd',s,e);
 loadAnalysis();
}

function setTeamPreset(mode){
 let d=new Date();
 if(mode==='yesterday')d.setDate(d.getDate()-1);
 if(mode==='daybefore')d.setDate(d.getDate()-2);
 let v=fmtDateInput(d);
 teamPresetMode=mode;
 updateTeamPresetButtons();
 setDateRange('teamStartDate','btnStartDate','teamEndDate','btnEndDate',v,v);
 loadPlayerList();
}

function initTeamDateFilter(){
 if(teamDateInited)return;
 teamDateInited=true;
 teamPresetMode='';
 updateTeamPresetButtons();
 setDateRange('teamStartDate','btnStartDate','teamEndDate','btnEndDate','','');
 ['teamStartDate','teamEndDate','teamStartTime','teamEndTime'].forEach(id=>{
  let el=$(id);
  if(el)el.addEventListener('change',handleTeamFilterChanged);
 });
}

recordsDateInited=recordsDateInited||false;
function initRecordsDateFilter(){
 if(recordsDateInited)return;
 recordsDateInited=true;
 setDateRange('recStartDate','btnRecStart','recEndDate','btnRecEnd','','');
}

itemsDateInited=itemsDateInited||false;
function initItemsDateFilter(){
 if(itemsDateInited)return;
 fetch('/api/items').then(r=>r.json()).then(d=>{
  if(d&&d.length){
   let first=d[d.length-1].event_time.split(' ');
   let last=d[0].event_time.split(' ');
   $('itemStartDate').value=first[0];
   $('itemEndDate').value=last[0];
   $('btnItemStart').textContent=first[0];
   $('btnItemEnd').textContent=last[0];
   itemsDateInited=true;
  }
 });
}

analysisDateInited=analysisDateInited||false;
function initAnalysisDateFilter(){
 if(analysisDateInited)return;
 fetch('/api/records?size=1').then(r=>r.json()).then(d=>{
  if(d.data&&d.data.length){
   let first=d.data[d.data.length-1].event_time.split(' ');
   let last=d.data[0].event_time.split(' ');
   $('anaStartDate').value=first[0];
   $('anaEndDate').value=last[0];
   $('btnAnaStart').textContent=first[0];
   $('btnAnaEnd').textContent=last[0];
   analysisDateInited=true;
  }
 });
}

function loadPlayerList(){
 if(!ensureSoftwareAccess('team_analysis','组队分析'))return;
 let parts=getTeamFilterQueryParts();
 let realSelected=getTeamSelectedNames();
 if(realSelected.length>0)parts.push('with='+encodeURIComponent(realSelected.join(',')));
 let url='/api/players'+(parts.length?('?'+parts.join('&')):'');
 fetch(url).then(r=>r.json()).then(d=>{
  if(d&&d.ok===false){
   if(d.requires_member){showMemberFeaturePrompt('组队分析');return}
   if(d.requires_software_login){openSoftwareGate();return}
   $('playerListContainer').innerHTML='<div style="color:#ff6b4a;padding:20px;text-align:center">'+esc(d.error||'加载失败')+'</div>';
   return;
  }
  let players=d.players||[];
  let selfPlayer=players.find(p=>p.is_self)||{};
  let activeName=d.active_player_name||selfPlayer.name||'';
  let activeKey=d.active_player_key||teamPlayerKey(selfPlayer)||'';
  normalizeTeamSelections(players,activeKey,activeName);
  if(d.error||!players.length){
   let emptyText=realSelected.length>0?'当前时间范围内没有匹配到共同队友':'暂无玩家数据（需要先补全对局详情）';
   $('playerListContainer').innerHTML='<div style="color:#8899aa;padding:20px;text-align:center">'+emptyText+'</div>';
   $('teamBtns').style.display='none';
   $('teamResultContainer').innerHTML='';
   return;
  }
  let html='<div class="player-grid">';
  let currentReal=getTeamSelectedNames();
  let canSelectOther=currentReal.length===2&&players.some(p=>!currentReal.includes(teamPlayerKey(p)));
  if(!canSelectOther&&hasTeamOtherSelection())selectedPlayers=currentReal.slice();
  players.forEach(p=>{
   let key=teamPlayerKey(p);
   let sel=selectedPlayers.includes(key)?'selected':'';
   let idLine=(p.duplicate_name&&p.player_id)?'<div class="p-id">ID '+esc(shortTeamPlayerId(p.player_id))+'</div>':'';
   html+='<div class="player-item '+sel+'" onclick="togglePlayer(\''+esc(escTeamArg(key))+'\')" title="'+esc(teamPlayerDisplay(p))+'">';
   html+='<div class="p-check">'+(sel?'✓':'')+'</div>';
   html+='<div class="p-main"><div class="p-name">'+esc(p.name||p.player_name||'')+'</div>'+idLine+'</div>';
   html+='<div class="p-count">'+p.count+'</div>';
   html+='</div>';
  });
  if(canSelectOther){
   let otherSel=hasTeamOtherSelection()?'selected':'';
   html+='<div class="player-item special '+otherSel+'" onclick="togglePlayer(\''+TEAM_OTHER_TOKEN+'\')">';
   html+='<div class="p-check">'+(otherSel?'✓':'')+'</div>';
   html+='<div class="p-main"><div class="p-name">其他玩家</div></div>';
   html+='<div class="p-count">聚合</div>';
   html+='</div>';
  }
  html+='</div>';
  $('playerListContainer').innerHTML=html;
  let hasSelection=getTeamSelectedNames().length>0;
  $('teamBtns').style.display=hasSelection?'flex':'none';
  if(hasSelection)loadTeamAnalysis();
  else $('teamResultContainer').innerHTML='';
 })
}

function togglePlayer(key){
 if(key===TEAM_OTHER_TOKEN){
  let real=getTeamSelectedNames();
  if(real.length!==2)return;
  if(hasTeamOtherSelection())selectedPlayers=real.slice();
  else selectedPlayers=real.concat(TEAM_OTHER_TOKEN);
  loadPlayerList();
  return;
 }
 let idx=selectedPlayers.indexOf(key);
 if(idx>=0){
  selectedPlayers=selectedPlayers.filter(v=>v!==key);
  if(key===teamActivePlayerKey)teamSelfDeselected=true;
 }else{
  let real=getTeamSelectedNames();
  let maxReal=hasTeamOtherSelection()?2:3;
  if(real.length>=maxReal)return;
  selectedPlayers.push(key);
  if(key===teamActivePlayerKey)teamSelfDeselected=false;
 }
 let nextReal=getTeamSelectedNames();
 if(teamActivePlayerKey&&nextReal.includes(teamActivePlayerKey)){
  nextReal=[teamActivePlayerKey].concat(nextReal.filter(v=>v!==teamActivePlayerKey));
 }
 selectedPlayers=hasTeamOtherSelection()&&nextReal.length===2?nextReal.concat(TEAM_OTHER_TOKEN):nextReal;
 loadPlayerList();
}

function clearPlayerSelection(){
 selectedPlayers=[];
 teamSelfDeselected=true;
 teamMapFilterSelected.clear();
 loadPlayerList();
 $('teamResultContainer').innerHTML='';
}

function openTeamMapFilter(){
 let maps=getTeamAvailableMaps(lastTeamData);
 if(!maps.length){
  showMsg('err','请先生成组队分析结果');
  return;
 }
 if(!teamMapFilterSelected.size)teamMapFilterSelected=new Set(maps);
 let rows=maps.map((name,idx)=>{
  let checked=teamMapFilterSelected.has(name)?' checked':'';
  return '<label style="display:flex;align-items:center;gap:8px;padding:8px 0"><input type="checkbox" class="team-map-option" value="'+esc(name)+'"'+checked+'> '+esc(name)+'</label>';
 }).join('');
 showActionOverlay({
  title:'选择地图',
  html:'<p class="dialog-note">取消不需要分析的地图后，组队分析表格会隐藏这些地图的对局并重新计算汇总。</p><div style="max-height:320px;overflow:auto;margin-top:10px">'+rows+'</div>',
  actions:[
   {label:'全选',className:'btn btn-muted',onClick:'selectTeamMapOptions(true)'},
   {label:'清空',className:'btn btn-muted',onClick:'selectTeamMapOptions(false)'},
   {label:'应用筛选',className:'btn btn-fetch',onClick:'applyTeamMapFilter()'}
  ],
  closable:true
 });
}

function selectTeamMapOptions(checked){
 document.querySelectorAll('.team-map-option').forEach(input=>{input.checked=checked});
}

function applyTeamMapFilter(){
 let selected=[...document.querySelectorAll('.team-map-option:checked')].map(input=>input.value);
 if(!selected.length){
  showMsg('err','至少选择一个地图');
  return;
 }
 teamMapFilterSelected=new Set(selected);
 closeActionOverlay();
 renderTeamAnalysis();
}

function loadTeamAnalysis(){
 if(!ensureSoftwareAccess('team_analysis','组队分析'))return;
 if(!getTeamSelectedNames().length){$('teamResultContainer').innerHTML='';return}
 let url='/api/team-analysis?players='+encodeURIComponent(selectedPlayers.join(','));
 let parts=getTeamFilterQueryParts();
 if(parts.length)url+='&'+parts.join('&');
 fetch(url).then(r=>r.json()).then(d=>{
  if(d&&d.ok===false){
   if(d.requires_member){showMemberFeaturePrompt('组队分析');return}
   if(d.requires_software_login){openSoftwareGate();return}
   $('teamResultContainer').innerHTML='<div style="color:#ff6b4a">'+esc(d.error||'加载失败')+'</div>';return
  }
  if(d.error){$('teamResultContainer').innerHTML='<div style="color:#ff6b4a">'+d.error+'</div>';return}
  lastTeamData=d;
  teamMapFilterSelected=new Set(getTeamAvailableMaps(d));
  renderTeamAnalysis();
 })
}

function getTeamAvailableMaps(data){
 let games=(data&&data.games)||[];
 return [...new Set(games.map(g=>g.map_name||'未知地图'))].sort((a,b)=>a.localeCompare(b,'zh-CN'));
}

function updateTeamMapFilterButton(){
 let btn=$('btnTeamMapFilter');
 if(!btn)return;
 let maps=getTeamAvailableMaps(lastTeamData);
 if(!maps.length){
  btn.textContent='地图选择';
  btn.disabled=true;
  return;
 }
 btn.disabled=false;
 let selected=[...teamMapFilterSelected].filter(m=>maps.includes(m));
 btn.textContent=selected.length===maps.length?'地图选择':'地图选择('+selected.length+'/'+maps.length+')';
}

function filteredTeamData(){
 if(!lastTeamData)return null;
 let maps=getTeamAvailableMaps(lastTeamData);
 let selected=teamMapFilterSelected.size?teamMapFilterSelected:new Set(maps);
 let games=(lastTeamData.games||[]).filter(g=>selected.has(g.map_name||'未知地图'));
 let players=(lastTeamData.players||[]).map((p,pi)=>{
  let stat={...p,games:0,escaped:0,kills:0,eq:0,out:0,profit:0,roles:[]};
  games.forEach(g=>{
   let pr=(g.player_records||[])[pi];
   if(!pr)return;
   stat.games++;
   if(pr.game_result===0)stat.escaped++;
   stat.kills+=Number(pr.kill_player)||0;
   stat.eq+=Number(pr.original_equipment_price)||0;
   stat.out+=Number(pr.gained_price)||0;
   stat.profit+=Number(pr.profit_loss)||0;
   if(pr.armed_force_id)stat.roles.push(String(pr.armed_force_id));
  });
  let fail=stat.games-stat.escaped;
  let kd=fail>0?stat.kills/fail:stat.kills;
  let avgEq=stat.games?stat.eq/stat.games:0;
  let rate=stat.eq>0?stat.profit/stat.eq:0;
  let roleCounts={};
  stat.roles.forEach(r=>{roleCounts[r]=(roleCounts[r]||0)+1});
  let commonRoleId=Object.keys(roleCounts).sort((a,b)=>roleCounts[b]-roleCounts[a])[0]||'';
  return {
   ...p,
   games:stat.games,
   evac_rate_num:stat.games?stat.escaped/stat.games:0,
   evac_rate:stat.games?(stat.escaped/stat.games*100).toFixed(1)+'%':'0.0%',
   kd:kd,
   avg_eq:avgEq,
   total_out:stat.out,
   total_profit:stat.profit,
   avg_rate_num:rate,
   avg_rate:(rate*100).toFixed(1)+'%',
   common_role:commonRoleId?roleName(commonRoleId):(stat.games?p.common_role:'-')
  };
 });
 let totalEscaped=games.filter(g=>g.game_result===0).length;
 let totalProfit=games.reduce((sum,g)=>{
  return sum+(g.player_records||[]).reduce((s,pr)=>s+(Number(pr?.profit_loss)||0),0);
 },0);
 let totalDuration=games.reduce((sum,g)=>sum+(Number(g.duration_s)||0),0);
 let mapCounts={};
 games.forEach(g=>{let name=g.map_name||'未知地图';mapCounts[name]=(mapCounts[name]||0)+1});
 let commonMap=Object.keys(mapCounts).sort((a,b)=>mapCounts[b]-mapCounts[a])[0]||'';
 let activePlayers=players.filter(p=>p.games>0);
 let killKing=activePlayers.length?activePlayers.reduce((best,p)=>p.kd>best.kd?p:best,activePlayers[0]):{};
 let profitKing=activePlayers.length?activePlayers.reduce((best,p)=>p.total_profit>best.total_profit?p:best,activePlayers[0]):{};
 let summary={
  ...(lastTeamData.summary||{}),
  total_games:games.length,
  total_escaped:totalEscaped,
  evac_rate:games.length?(totalEscaped/games.length*100).toFixed(1)+'%':'0.0%',
  team_kd:activePlayers.length?(activePlayers.reduce((s,p)=>s+(Number(p.kd)||0),0)/activePlayers.length).toFixed(2):'0.00',
  total_profit:totalProfit,
  avg_duration:games.length?totalDuration/games.length:0,
  common_map:commonMap||'-',
  kill_king:killKing.player_name||'',
  kill_king_kd:Number(killKing.kd)||0,
  profit_king:profitKing.player_name||'',
  profit_king_val:Number(profitKing.total_profit)||0
 };
 return {...lastTeamData,games,players,summary};
}

function teamAnalysisPlayerLabel(p,players){
 let name=esc((p&&p.player_name)||'-');
 let duplicate=(players||[]).filter(x=>String(x.player_name||'')===String((p&&p.player_name)||'')).length>1;
 if(duplicate&&p&&p.player_id)return name+' <span style="font-size:11px;color:#7f95a6;font-weight:400">ID '+esc(shortTeamPlayerId(p.player_id))+'</span>';
 return name;
}

function teamAnalysisFileLabel(){
 let players=(lastTeamData&&lastTeamData.players)||[];
 let counts={};
 players.forEach(p=>{let n=String(p.player_name||'');counts[n]=(counts[n]||0)+1});
 let names=players.map(p=>{
  let n=p.player_name||shortTeamPlayerId(p.player_id)||'玩家';
  if(counts[String(p.player_name||'')]>1&&p.player_id)n+='_'+shortTeamPlayerId(p.player_id);
  return n;
 }).filter(Boolean).join('_');
 return names||getTeamSelectedNames().join('_');
}

function teamAnalysisReportLabel(){
 let players=(lastTeamData&&lastTeamData.players)||[];
 let counts={};
 players.forEach(p=>{let n=String(p.player_name||'');counts[n]=(counts[n]||0)+1});
 let names=players.map(p=>{
  let n=p.player_name||shortTeamPlayerId(p.player_id)||'玩家';
  if(counts[String(p.player_name||'')]>1&&p.player_id)n+='(ID '+shortTeamPlayerId(p.player_id)+')';
  return n;
 }).filter(Boolean).join('、');
 return names;
}

function renderTeamAnalysis(){
 updateTeamMapFilterButton();
 let d=filteredTeamData();
 if(!d){$('teamResultContainer').innerHTML='';return}
 let games=d.games,players=d.players,summary=d.summary;
 if(!games||!games.length){
  $('teamResultContainer').innerHTML='<div style="color:#8899aa">当前地图筛选无数据，请重新选择地图。</div>';
  return;
 }

  let html='';
  html+='<div class="summary team-summary">';
  html+='<div class="card"><div class="val">'+summary.total_games+'</div><div class="lbl">总场次</div></div>';
  html+='<div class="card"><div class="val">'+summary.total_escaped+'</div><div class="lbl">撤离成功</div></div>';
  html+='<div class="card"><div class="val">'+summary.evac_rate+'</div><div class="lbl">撤离率</div></div>';
  html+='<div class="card"><div class="val">'+summary.team_kd+'</div><div class="lbl">队伍 KD</div></div>';
  html+='<div class="card"><div class="sub">'+(summary.kill_king||'-')+'</div><div class="val">'+summary.kill_king_kd.toFixed(2)+'</div><div class="lbl">击杀王</div></div>';
  html+='<div class="card"><div class="sub">'+(summary.profit_king||'-')+'</div><div class="val">'+fmtPrice(summary.profit_king_val)+'</div><div class="lbl">收益王</div></div>';
  html+='<div class="card"><div class="val">'+fmtDur(Math.round(summary.avg_duration))+'</div><div class="lbl">平均时长</div></div>';
  html+='</div>';
  html+='<div class="matrix-note"><span class="note-icon">i</span><div><strong>查看对局详情</strong><br>点击表头中的具体时间，即可打开该局的对局详情弹窗。</div></div>';

   let ncols=games.length+3;
   html+='<div class="matrix-wrap" id="teamMatrixWrap" style="max-height:calc(100vh - 400px)"><table class="matrix-table"><colgroup>';
   html+='<col style="width:90px"><col style="width:100px"><col style="width:100px">';
   games.forEach(()=>{html+='<col style="width:110px">'});
   html+='</colgroup><thead>';

    let dateGroups=[];
    let curDate='',curSpan=0;
    let dateSepCols=new Set();
    games.forEach((g,i)=>{
     let d=g.event_time.split(' ')[0]||'';
     if(d!==curDate){
      if(curSpan>0){
       dateGroups.push({date:curDate,span:curSpan});
       dateSepCols.add(i-1);
      }
      curDate=d;curSpan=1;
     }else curSpan++;
    });
    if(curSpan>0)dateGroups.push({date:curDate,span:curSpan});

    html+='<tr><th class="col-tag" rowspan="2">汇总</th><th class="col-sum" rowspan="2">汇总数据</th><th class="col-detail" rowspan="2">明细</th>';
    let dgIdx=0;
    dateGroups.forEach((dg,dgi)=>{
     let sep=dgi<dateGroups.length-1?'border-right:3px solid #4a7aaa':'';
     html+='<th class="col-time" colspan="'+dg.span+'" style="'+sep+'">'+dg.date+'</th>';
    });
    html+='</tr><tr>';
    games.forEach((g,i)=>{
     let t=(g.event_time.split(' ')[1]||g.event_time).substring(0,5);
     let sep=dateSepCols.has(i)?'border-right:3px solid #4a7aaa':'';
     let timeTitle='点击查看 '+(g.event_time||'')+' 的对局详情';
     html+='<th class="col-time time-link" style="'+sep+'" onclick="showDetail(\''+g.room_id+'\')" title="'+esc(timeTitle)+'">'+t+'</th>';
    });
    html+='</tr></thead><tbody>';

   let rcls={0:'cell-ok',1:'cell-fail',2:'cell-timeout',3:'cell-leave'};
   function teamMetricRange(getter){
    let vals=[];
    games.forEach(g=>{
     (g.player_records||[]).filter(Boolean).forEach(pr=>{
      let v=Number(getter(pr,g));
      if(Number.isFinite(v))vals.push(v);
     });
    });
    if(!vals.length)return {min:0,max:0};
    return {min:Math.min(...vals),max:Math.max(...vals)};
   }
   function teamMetricStyle(getter,value){
    let range=teamMetricRange(getter);
    return gc(Number(value)||0,range.min,range.max);
   }
   function teamMetricStyleSigned(getter,value){
    let range=teamMetricRange(getter);
    return gcSigned(Number(value)||0,range.min,range.max);
   }
   function gameProfitRate(pr){
    if(!pr)return 0;
    let eq=parseInt(pr.original_equipment_price)||0;
    let pl=parseInt(pr.profit_loss)||0;
    return eq>0?pl/eq:0;
   }

   html+='<tr class="global-row"><td class="col-tag">常用地图</td><td class="col-sum">'+summary.common_map+'</td><td class="col-detail">地图</td>';
   games.forEach(g=>{html+='<td>'+(g.map_name||'-')+'</td>'});
   html+='</tr>';

   html+='<tr class="global-row"><td class="col-tag">平均时长</td><td class="col-sum">'+fmtDur(Math.round(summary.avg_duration))+'</td><td class="col-detail">时长</td>';
   let durs=games.map(g=>g.duration_s||0),minD=Math.min(...durs),maxD=Math.max(...durs);
   games.forEach(g=>{let v=g.duration_s||0;html+='<td class="heat-cell" style="'+gc(v,minD,maxD)+'">'+fmtDur(v)+'</td>'});
   html+='</tr>';

   html+='<tr class="global-row"><td class="col-tag">撤离率</td><td class="col-sum">'+summary.evac_rate+'</td><td class="col-detail">结果</td>';
   games.forEach(g=>{html+='<td class="'+(rcls[g.game_result]||'')+'">'+(RESULT_MAP[g.game_result]||g.game_result)+'</td>'});
   html+='</tr>';

players.forEach((p,pi)=>{
     let restCols=ncols-2;
     html+='<tr class="player-head"><td colspan="2">'+teamAnalysisPlayerLabel(p,players)+'</td>'+(restCols>0?'<td colspan="'+restCols+'"></td>':'')+'</tr>';

    html+='<tr class="role-row"><td class="col-tag">常用角色</td><td class="col-sum">'+(p.common_role||'-')+'</td><td class="col-detail">角色</td>';
    games.forEach(g=>{let pr=g.player_records[pi];html+=pr?'<td>'+roleName(pr.armed_force_id)+'</td>':'<td>-</td>'});
    html+='</tr>';

    html+='<tr><td class="col-tag">撤离率</td><td class="col-sum">'+p.evac_rate+'</td><td class="col-detail">结果</td>';
    games.forEach(g=>{
     let pr=g.player_records[pi];
     if(pr)html+='<td class="'+(rcls[pr.game_result]||'')+'">'+(RESULT_MAP[pr.game_result]||pr.game_result)+'</td>';
     else html+='<td>-</td>';
    });
    html+='</tr>';

    let maxKD=Math.max(...players.map(x=>x.kd),1);

    html+='<tr><td class="col-tag">KD</td><td class="col-sum" style="'+gc(p.kd,0,maxKD)+'">'+p.kd.toFixed(2)+'</td><td class="col-detail">击杀玩家</td>';
    games.forEach(g=>{let pr=g.player_records[pi];let kp=pr?(pr.kill_player||0):0;html+=pr?'<td class="heat-cell" style="'+teamMetricStyle(pr=>pr.kill_player||0,kp)+'">'+kp+'</td>':'<td>-</td>'});
    html+='</tr>';

     let minAvgEq=Math.min(...players.map(x=>Number(x.avg_eq)||0)),maxAvgEq=Math.max(...players.map(x=>Number(x.avg_eq)||0));
     html+='<tr><td class="col-tag">带入装备价值</td><td class="col-sum" style="'+gc(p.avg_eq,minAvgEq,maxAvgEq)+'">'+fmtPrice(Math.round(p.avg_eq))+'</td><td class="col-detail">带入装备价值</td>';
    games.forEach(g=>{let pr=g.player_records[pi];let v=parseInt(pr?.original_equipment_price)||0;html+=pr?'<td class="heat-cell" style="'+teamMetricStyle(pr=>parseInt(pr.original_equipment_price)||0,v)+'">'+fmtPrice(v)+'</td>':'<td>-</td>'});
    html+='</tr>';

     html+='<tr><td class="col-tag">带出价值</td><td class="col-sum" style="'+gc(p.total_out,Math.min(...players.map(x=>Number(x.total_out)||0)),Math.max(...players.map(x=>Number(x.total_out)||0),1))+'">'+fmtPrice(p.total_out)+'</td><td class="col-detail">带出价值</td>';
    games.forEach(g=>{let pr=g.player_records[pi];let v=parseInt(pr?.gained_price)||0;html+=pr?'<td class="heat-cell" style="'+teamMetricStyle(pr=>parseInt(pr.gained_price)||0,v)+'">'+fmtPrice(v)+'</td>':'<td>-</td>'});
    html+='</tr>';

     html+='<tr><td class="col-tag">盈亏</td><td class="col-sum" style="'+gcSigned(p.total_profit,Math.min(...players.map(x=>Number(x.total_profit)||0)),Math.max(...players.map(x=>Number(x.total_profit)||0)))+'">'+fmtPrice(p.total_profit)+'</td><td class="col-detail">盈亏</td>';
    games.forEach(g=>{let pr=g.player_records[pi];let v=parseInt(pr?.profit_loss)||0;html+=pr?'<td class="heat-cell" style="'+teamMetricStyleSigned(pr=>parseInt(pr.profit_loss)||0,v)+'">'+fmtPrice(v)+'</td>':'<td>-</td>'});
    html+='</tr>';

     html+='<tr><td class="col-tag">收益率</td><td class="col-sum" style="'+gcSigned(p.avg_rate_num,-0.5,2)+'">'+p.avg_rate+'</td><td class="col-detail">收益率</td>';
    games.forEach(g=>{let pr=g.player_records[pi];
     if(pr){
      let rate=gameProfitRate(pr),pct=(rate*100).toFixed(1)+'%';
      html+='<td class="heat-cell" style="'+teamMetricStyleSigned(pr=>gameProfitRate(pr),rate)+'">'+pct+'</td>';
     }else html+='<td>-</td>';
    });
    html+='</tr>';
   });

  html+='</tbody></table></div>';

  $('teamResultContainer').innerHTML=html;
  bindTeamMatrixScroll();
}

let lastTeamData=null;

function bindTeamMatrixScroll(){
 let wrap=$('teamMatrixWrap');
 if(!wrap||wrap.dataset.wheelBound==='1')return;
 wrap.dataset.wheelBound='1';
 wrap.addEventListener('wheel',function(e){
  if(wrap.scrollWidth<=wrap.clientWidth+1)return;
  let dx=e.deltaX;
  if(!dx&&e.shiftKey)dx=e.deltaY;
  if(!dx)return;
  wrap.scrollLeft+=dx;
  e.preventDefault();
 },{passive:false});
}

function exportTeamExcel(){
 if(!ensureSoftwareAccess('team_export','组队分析导出'))return;
 if(!getTeamSelectedNames().length)return;
 let payload={players:selectedPlayers.join(',')};
 let sd=$('teamStartDate')?.value||'',st=$('teamStartTime')?.value||'00:00',ed=$('teamEndDate')?.value||'',et=$('teamEndTime')?.value||'23:59';
 if(sd)payload.start=sd+' '+st;
 if(ed)payload.end=ed+' '+et;
 fetch('/api/export-team',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)}).then(r=>{
  if(!r.ok)return r.json().then(e=>{throw new Error(e.error||'导出失败')});
  return r.blob();
 }).then(blob=>{
 let url=URL.createObjectURL(blob);
 let a=document.createElement('a');a.href=url;
  let names=teamAnalysisFileLabel();
  a.download='组队分析_'+names+'.xlsx';
  document.body.appendChild(a);a.click();URL.revokeObjectURL(url);a.remove();
 }).catch(e=>alert('导出失败: '+e.message));
}

function buildTeamActionPayload(){
 let payload={players:selectedPlayers.join(',')};
 let label=teamAnalysisReportLabel();
 if(label)payload.players_label=label;
 let sd=$('teamStartDate')?.value||'',st=$('teamStartTime')?.value||'00:00',ed=$('teamEndDate')?.value||'',et=$('teamEndTime')?.value||'23:59';
 if(sd)payload.start=sd+' '+st;
 if(ed)payload.end=ed+' '+et;
 return payload;
}

function setTeamReportSubmitting(submitting){
 let btn=$('btnTeamReport');
 if(!btn)return;
 btn.disabled=submitting;
 btn.textContent=submitting?'提交中...':'生成 PDF 报告';
}

function showTeamReportError(message){
 let msg=translateErrorMessage(message);
 showActionOverlay({
  title:'报告操作失败',
  html:'<p>'+esc(msg)+'</p><p class="dialog-note">请检查 AI 配置、模型服务、当前筛选范围或报告任务状态后重试。</p>',
  actions:[{label:'关闭',className:'btn btn-muted',onClick:'closeActionOverlay()'}],
  closable:true
 });
}

function downloadTeamReport(url,filename){
 if(!url)return;
 let a=document.createElement('a');
 a.href=url;
 if(filename)a.download=filename;
 document.body.appendChild(a);
 a.click();
 a.remove();
}

function teamReportStatusText(status){
 return {pending:'排队中',running:'生成中',canceling:'取消中',done:'已完成',failed:'失败',canceled:'已取消'}[status]||status||'-';
}

function teamReportTimeText(ts){
 if(!ts)return'-';
 let d=new Date(Number(ts)*1000);
 if(Number.isNaN(d.getTime()))return'-';
 return d.toLocaleString();
}

function teamReportCountdownText(r){
 if(!r||r.status!=='running')return'';
 let total=(r.progress&&r.progress.total)||{};
 if(Number(total.current)!==3)return'';
 let timeout=Number(r.ai_timeout_seconds)||0;
 let started=Number(r.ai_started_at)||0;
 if(!timeout||!started)return'';
 let elapsed=Math.max(0,Math.floor(Date.now()/1000-started));
 let left=Math.max(0,timeout-elapsed);
 return '数据计算处理中，最长还需等待'+left+'秒';
}

function teamReportProgressPercent(r){
 if(!r)return 0;
 if(r.status==='done')return 100;
 let total=(r.progress&&r.progress.total)||{};
 if(r.status==='running'&&Number(total.current)===3){
  let timeout=Number(r.ai_timeout_seconds)||0;
  let started=Number(r.ai_started_at)||0;
  if(timeout&&started){
   let elapsed=Math.max(0,Date.now()/1000-started);
   return Math.max(0,Math.min(99,elapsed/timeout*100));
  }
 }
 return Math.max(0,Math.min(Number(r.percent||0),100));
}

function isTeamReportDeletable(r){
 return !!r&&['done','failed','canceled'].includes(r.status);
}

function setTeamReportBatchMode(enabled){
 teamReportBatchMode=!!enabled;
 if(!teamReportBatchMode)teamReportSelectedIds.clear();
 let bar=$('teamReportBatchBar'),btn=$('btnTeamReportBatch');
 if(bar)bar.classList.toggle('show',teamReportBatchMode);
 if(btn)btn.style.display=teamReportBatchMode?'none':'';
 renderTeamReportList(teamReportLastList);
}

function toggleTeamReportSelection(id,checked){
 if(checked)teamReportSelectedIds.add(id);
 else teamReportSelectedIds.delete(id);
}

function selectAllTeamReports(){
 (teamReportLastList||[]).forEach(r=>{if(isTeamReportDeletable(r))teamReportSelectedIds.add(r.id)});
 renderTeamReportList(teamReportLastList);
}

function deleteSelectedTeamReports(){
 let ids=[...teamReportSelectedIds];
 if(!ids.length){
  showMsg('err','请先选择需要删除的报告');
  return;
 }
 showActionOverlay({
  title:'确认批量删除 PDF 报告',
  html:'<p>将删除已选择的 '+ids.length+' 个 PDF 报告文件，该操作不可撤销。</p>',
  actions:[
   {label:'取消',className:'btn btn-muted',onClick:'closeActionOverlay()'},
   {label:'确认删除',className:'btn btn-danger',onClick:'confirmDeleteSelectedTeamReports()'}
  ],
  closable:true
 });
}

function confirmDeleteSelectedTeamReports(){
 let ids=[...teamReportSelectedIds];
 if(!ids.length)return;
 closeActionOverlay();
 Promise.all(ids.map(id=>postJSON('/api/team-report-delete',{job_id:id}).then(d=>({id,d})).catch(e=>({id,error:e}))))
 .then(results=>{
  let failed=results.filter(item=>item.error||!item.d?.ok);
  teamReportSelectedIds.clear();
  if(failed.length){
   showTeamReportError('部分报告删除失败：'+failed.length+' 个');
  }else{
   showMsg('ok','已删除 '+results.length+' 个报告');
  }
  refreshTeamReportList();
 });
}

function openTeamReportManager(){
 let overlay=$('teamReportOverlay');
 if(overlay)overlay.classList.add('show');
 refreshTeamReportList();
}

function closeTeamReportManager(){
 let overlay=$('teamReportOverlay');
 if(overlay)overlay.classList.remove('show');
 setTeamReportBatchMode(false);
}

function renderTeamReportList(reports){
 let box=$('teamReportList');
 if(!box)return;
 let list=Array.isArray(reports)?reports:[];
 teamReportLastList=list;
 let bar=$('teamReportBatchBar'),btn=$('btnTeamReportBatch');
 if(bar)bar.classList.toggle('show',teamReportBatchMode);
 if(btn)btn.style.display=teamReportBatchMode?'none':'';
 if(!list.length){
  box.innerHTML='<div class="report-empty">暂无报告任务。点击“生成 PDF 报告”后会在这里显示排队和生成进度。</div>';
  return;
 }
 box.innerHTML=list.map(r=>{
  let status=esc(r.status||'');
  let progress=teamReportProgressPercent(r);
  let total=(r.progress&&r.progress.total)||{};
  let sub=(r.progress&&r.progress.sub)||{};
  let countdown=teamReportCountdownText(r);
  let progressText=esc(r.status==='done'?'已完成':(countdown||total.text||sub.text||teamReportStatusText(r.status)));
  let time='提交：'+teamReportTimeText(r.created_at);
  if(r.completed_at)time+=' · 完成：'+teamReportTimeText(r.completed_at);
  let range=(r.start||r.end)?('范围：'+(r.start||'不限')+' 至 '+(r.end||'不限')):'范围：不限';
  let note=r.error?'<div class="report-item-meta" style="color:#ff9a8f">原因：'+esc(translateErrorMessage(r.error))+'</div>':'';
  let actions='';
  if(r.status==='pending'||r.status==='running')actions+='<button class="btn btn-danger" onclick="cancelTeamReport(\''+esc(r.id)+'\')">取消</button>';
  if(r.status==='done')actions+='<button class="btn btn-go" onclick="downloadTeamReport(\''+esc(r.download_url)+'\',\''+esc(r.filename||'')+'\')">下载</button><button class="btn btn-danger" onclick="deleteTeamReport(\''+esc(r.id)+'\')">删除</button>';
  if(r.status==='failed'||r.status==='canceled')actions+='<button class="btn btn-danger" onclick="deleteTeamReport(\''+esc(r.id)+'\')">删除</button>';
  let checkbox='';
  if(teamReportBatchMode){
   let disabled=isTeamReportDeletable(r)?'':' disabled';
   let checked=teamReportSelectedIds.has(r.id)?' checked':'';
   checkbox='<input type="checkbox"'+checked+disabled+' onchange="toggleTeamReportSelection(\''+esc(r.id)+'\',this.checked)">';
   actions='';
  }
  return '<div class="report-item"><div class="'+(teamReportBatchMode?'report-select':'')+'">'+checkbox+'<div style="flex:1"><div class="report-item-head"><div><div class="report-item-title">'+esc(r.players||'组队分析报告')+'</div><div class="report-item-meta">'+esc(range)+'<br>'+esc(time)+'<br>'+progressText+'</div>'+note+'</div><div class="report-item-actions"><span class="report-status '+status+'">'+esc(teamReportStatusText(r.status))+'</span>'+actions+'</div></div><div class="report-progress"><div class="fill" style="width:'+Math.max(0,Math.min(progress,100))+'%"></div></div></div></div></div>';
 }).join('');
}

function refreshTeamReportList(){
 fetch('/api/team-report-list').then(async r=>{
  let d=await r.json().catch(()=>({error:'报告列表加载失败'}));
  if(!r.ok||!d.ok)throw new Error(d.error||'报告列表加载失败');
  return d.reports||[];
 }).then(reports=>{
  renderTeamReportList(reports);
  let hasActive=reports.some(r=>r.status==='pending'||r.status==='running'||r.status==='canceling');
  reports.forEach(r=>{
   if(r.status==='done'){
    if(teamReportKnownInited&&!teamReportKnownDone.has(r.id))showMsg('ok','PDF 报告已生成，可在“报告管理”中下载');
    teamReportKnownDone.add(r.id);
   }
  });
  teamReportKnownInited=true;
  if(teamReportManagerTimer){clearTimeout(teamReportManagerTimer);teamReportManagerTimer=null}
  if(hasActive)teamReportManagerTimer=setTimeout(refreshTeamReportList,1000);
 }).catch(e=>{
  let box=$('teamReportList');
  if(box)box.innerHTML='<div class="report-empty">报告列表加载失败：'+esc(translateErrorMessage(e))+'</div>';
 });
}

function cancelTeamReport(jobId){
 postJSON('/api/team-report-cancel',{job_id:jobId}).then(d=>{
  if(!d.ok)throw new Error(d.error||'取消失败');
  showMsg('info',d.message||'已取消报告任务');
  refreshTeamReportList();
 }).catch(e=>showTeamReportError(e));
}

function deleteTeamReport(jobId){
 pendingDeleteReportId=jobId;
 showActionOverlay({
  title:'确认删除 PDF 报告',
  html:'<p>删除后会同时移除保存数据库中的 PDF 文件，无法在报告管理中继续下载，该操作不可撤销。</p>',
  actions:[
   {label:'取消',className:'btn btn-muted',onClick:'pendingDeleteReportId=\'\';closeActionOverlay()'},
   {label:'确认删除',className:'btn btn-danger',onClick:'confirmDeleteTeamReport()'}
  ],
  closable:true
 });
}

function confirmDeleteTeamReport(){
 let jobId=pendingDeleteReportId;
 if(!jobId)return;
 closeActionOverlay();
 postJSON('/api/team-report-delete',{job_id:jobId}).then(d=>{
  pendingDeleteReportId='';
  if(!d.ok)throw new Error(d.error||'删除失败');
  showMsg('ok','报告已删除');
  refreshTeamReportList();
 }).catch(e=>{pendingDeleteReportId='';showTeamReportError(e)});
}

function confirmGenerateTeamReport(){
 if(!pendingTeamReportPayload)return;
 closeActionOverlay();
 setTeamReportSubmitting(true);
 fetch('/api/team-report-pdf',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(pendingTeamReportPayload)}).then(async r=>{
  let d=await r.json().catch(()=>({error:'报告任务启动失败'}));
  if(!r.ok||!d.ok)throw new Error(d.error||'报告任务启动失败');
  return d;
 }).then(d=>{
  pendingTeamReportPayload=null;
  setTeamReportSubmitting(false);
  showMsg('info','PDF 报告已加入生成队列，可在“报告管理”中查看进度');
  openTeamReportManager();
 }).catch(e=>{
  pendingTeamReportPayload=null;
  setTeamReportSubmitting(false);
  showTeamReportError(e);
 });
}

function showGenerateTeamReportConfirm(){
 if(!pendingTeamReportPayload)return;
 showActionOverlay({
  title:'确认生成 PDF 报告',
  html:'<p>AI PDF 报告生成可能需要较长时间，任务会进入后台队列并按提交顺序处理。</p><p class="dialog-note">AI API 端点、模型和密钥保存在 config.ai.json，可在此处编辑后立即生效。</p>',
  actions:[
   {label:'取消',className:'btn btn-muted',onClick:'pendingTeamReportPayload=null;closeActionOverlay()'},
   {label:'编辑 AI 设置',className:'btn btn-muted',onClick:'openAiConfigEditor()'},
   {label:'开始生成',className:'btn btn-fetch',onClick:'confirmGenerateTeamReport()'}
  ],
  closable:true
 });
}

function openAiConfigEditor(){
 fetch('/api/ai-config').then(r=>r.json()).then(d=>{
  if(!d.ok)throw new Error(d.error||'AI 配置读取失败');
  let cfg=d.config||{};
  showActionOverlay({
   title:'AI 报告配置',
   html:
    '<div style="display:grid;gap:10px">'+
    '<label class="dialog-note">API 端点</label><input id="aiCfgBaseUrl" class="dialog-input" value="'+esc(cfg.base_url||'')+'" placeholder="https://api.openai.com/v1">'+
    '<label class="dialog-note">模型</label><input id="aiCfgModel" class="dialog-input" value="'+esc(cfg.model||'')+'" placeholder="gpt-4.1-mini">'+
    '<label class="dialog-note">API Key</label><input id="aiCfgApiKey" class="dialog-input" type="password" value="'+esc(cfg.api_key||'')+'" placeholder="sk-...">'+
    '<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">'+
    '<div><label class="dialog-note">超时秒数</label><input id="aiCfgTimeout" class="dialog-input" type="number" min="1" value="'+esc(cfg.timeout_seconds||120)+'"></div>'+
    '<div><label class="dialog-note">Temperature</label><input id="aiCfgTemperature" class="dialog-input" type="number" min="0" max="2" step="0.1" value="'+esc(cfg.temperature??0.3)+'"></div>'+
    '</div>'+
    '<p class="dialog-note">保存后会写入项目根目录的 config.ai.json，下一次生成 PDF 会立即使用新配置。</p>'+
    '</div>',
   actions:[
    {label:'返回',className:'btn btn-muted',onClick:'showGenerateTeamReportConfirm()'},
    {label:'保存配置',className:'btn btn-fetch',onClick:'saveAiConfigFromEditor()'}
   ],
   closable:true
  });
 }).catch(e=>showTeamReportError(e));
}

function saveAiConfigFromEditor(){
 let payload={
  base_url:$('aiCfgBaseUrl')?.value.trim()||'',
  model:$('aiCfgModel')?.value.trim()||'',
  api_key:$('aiCfgApiKey')?.value.trim()||'',
  timeout_seconds:parseInt($('aiCfgTimeout')?.value||'120',10)||120,
  temperature:parseFloat($('aiCfgTemperature')?.value||'0.3')
 };
 postJSON('/api/ai-config',payload).then(d=>{
  if(!d.ok)throw new Error(d.error||'AI 配置保存失败');
  showMsg('ok','AI 配置已保存');
  showGenerateTeamReportConfirm();
 }).catch(e=>showTeamReportError(e));
}

function generateTeamReport(){
 if(!ensureSoftwareAccess('ai_report','AI PDF 报告'))return;
 if(!getTeamSelectedNames().length){
  showMsg('err','请先选择玩家并生成组队分析');
  return;
 }
 if(!lastTeamData||!lastTeamData.games||!lastTeamData.games.length){
  showMsg('err','请先点击“分析组队数据”生成结果');
  return;
 }
 pendingTeamReportPayload=buildTeamActionPayload();
 showGenerateTeamReportConfirm();
}

function downloadTeamImage(){
 if(!ensureSoftwareAccess('team_export','组队分析导出'))return;
 let wrap=$('teamMatrixWrap');if(!wrap)return;
 let table=wrap.querySelector('.matrix-table');if(!table)return;
 let btn=event.target;btn.textContent='生成中...';btn.disabled=true;
 let orig=wrap.style.cssText;
 let panel=wrap.closest('.panel');
 let origPanel=panel?panel.style.cssText:'';
 let origTable=table.style.cssText;
 wrap.style.maxHeight='none';wrap.style.overflow='visible';wrap.style.width=table.scrollWidth+'px';wrap.style.display='inline-block';
 table.style.width=table.scrollWidth+'px';
 if(panel){panel.style.overflow='visible';panel.style.maxHeight='none'}
 setTimeout(()=>{
  html2canvas(wrap,{backgroundColor:'#162029',scale:2,logging:false,useCORS:true,scrollX:0,scrollY:0,width:table.scrollWidth,height:wrap.scrollHeight}).then(canvas=>{
   wrap.style.cssText=orig;
   table.style.cssText=origTable;
   if(panel)panel.style.cssText=origPanel;
   canvas.toBlob(function(blob){
    if(!blob){alert('生成失败');btn.textContent='导出图片';btn.disabled=false;return}
    let url=URL.createObjectURL(blob);
    let a=document.createElement('a');a.href=url;
    let names=teamAnalysisFileLabel();
    a.download='组队分析_'+names+'.png';
    document.body.appendChild(a);a.click();
    URL.revokeObjectURL(url);a.remove();
    btn.textContent='导出图片';btn.disabled=false;
   },'image/png');
  }).catch(e=>{
   wrap.style.cssText=orig;
   table.style.cssText=origTable;
   if(panel)panel.style.cssText=origPanel;
   alert('生成失败');btn.textContent='导出图片';btn.disabled=false;
  });
 },200);
}

initTheme();
updateItemOwnerScopeButton();
syncOpenSourceState({}).then(()=>syncHeaderState({closeGate:true})).then(()=>{loadSummary();refreshVisiblePanel()});
</script>
</body>
</html>"""


class APIHandler(BaseHTTPRequestHandler):
    db: Database = None
    fetch_callback = None
    action_lock = threading.Lock()

    def log_message(self, format, *args):
        pass

    def _json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _html(self, html):
        body = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _svg(self, svg):
        body = svg.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "image/svg+xml; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _png_file(self, path):
        if not path.exists():
            self._svg(FAVICON_SVG)
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "image/png")
        self.send_header("Cache-Control", "public, max-age=3600")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length", 0))
        if not length:
            return {}
        return json.loads(self.rfile.read(length))

    def _try_acquire_action(self):
        if self.action_lock.acquire(blocking=False):
            return True
        self._json({"ok": False, "error": "当前已有任务进行中，请等待完成后再试"}, 409)
        return False

    def _is_client_disconnect(self, err: Exception) -> bool:
        return isinstance(err, (BrokenPipeError, ConnectionAbortedError, ConnectionResetError))

    def _write_server_error(self, err: Exception):
        import traceback

        traceback.print_exc()
        try:
            self._json({"error": f"服务端异常: {err}"}, 500)
        except Exception as write_err:
            if self._is_client_disconnect(write_err):
                return
            raise

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            params = parse_qs(parsed.query)

            if path == "/" or path == "/index.html":
                self._html(HTML_PAGE)
            elif path == "/logo.png":
                self._png_file(LOGO_PATH)
            elif path == "/favicon.svg":
                self._svg(FAVICON_SVG)
            elif path == "/api/stats":
                self._api_stats()
            elif path == "/api/records":
                self._api_records(params)
            elif path == "/api/items":
                self._api_items(params)
            elif path == "/api/assets":
                self._api_assets(params)
            elif path == "/api/maps":
                self._api_maps()
            elif path == "/api/accounts":
                self._api_accounts()
            elif path == "/api/software-session":
                self._api_software_session()
            elif path == "/api/ai-config":
                self._api_ai_config()
            elif path == "/api/action-status":
                self._api_action_status(params)
            elif path == "/api/detail":
                self._api_detail(params)
            elif path == "/api/analysis":
                self._api_analysis(params)
            elif path == "/api/trends":
                self._api_trends(params)
            elif path == "/api/players":
                self._api_players(params)
            elif path == "/api/team-analysis":
                self._api_team_analysis(params)
            elif path == "/api/team-report-status":
                self._api_team_report_status(params)
            elif path == "/api/team-report-download":
                self._api_team_report_download(params)
            elif path == "/api/team-report-list":
                self._api_team_report_list()
            else:
                self.send_error(404)
        except Exception as e:
            if self._is_client_disconnect(e):
                return
            self._write_server_error(e)

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/api/login":
                self._api_login()
            elif parsed.path == "/api/logout":
                self._api_logout()
            elif parsed.path == "/api/software/register":
                self._api_software_register()
            elif parsed.path == "/api/software/login":
                self._api_software_login()
            elif parsed.path == "/api/software/logout":
                self._api_software_logout()
            elif parsed.path == "/api/software/bind-current":
                self._api_software_bind_current()
            elif parsed.path == "/api/ai-config":
                self._api_ai_config_save()
            elif parsed.path == "/api/account/select":
                self._api_account_select()
            elif parsed.path == "/api/account/delete":
                self._api_account_delete()
            elif parsed.path == "/api/fetch":
                self._api_fetch()
            elif parsed.path == "/api/fetch-smart":
                self._api_fetch_smart()
            elif parsed.path == "/api/fetch-details":
                self._api_fetch_details()
            elif parsed.path == "/api/backup/export":
                self._api_backup_export()
            elif parsed.path == "/api/backup/import":
                self._api_backup_import()
            elif parsed.path == "/api/assets-refresh":
                self._api_assets_refresh()
            elif parsed.path == "/api/assets-catalog-refresh":
                self._api_assets_catalog_refresh()
            elif parsed.path == "/api/clear-data":
                self._api_clear_data()
            elif parsed.path == "/api/export-team":
                self._api_export_team()
            elif parsed.path == "/api/team-report-pdf":
                self._api_team_report_pdf()
            elif parsed.path == "/api/team-report-cancel":
                self._api_team_report_cancel()
            elif parsed.path == "/api/team-report-delete":
                self._api_team_report_delete()
            else:
                self.send_error(404)
        except Exception as e:
            if self._is_client_disconnect(e):
                return
            self._write_server_error(e)

    def _get_battle_tag_rules(self):
        rows = self.db.get_battle_tag_rules()
        return rows, {row["tag_name"]: row for row in rows}

    def _make_tag_payload(self, tag_name, rules_by_name):
        rule = rules_by_name.get(tag_name)
        if not rule:
            return {
                "name": tag_name,
                "dimension": "中性评价",
                "rule_text": "",
                "note": "",
                "tip": "",
            }
        return {
            "name": rule["tag_name"],
            "dimension": rule["dimension"],
            "rule_text": rule["rule_text"],
            "note": rule["note"],
            "tip": _rule_tip(rule),
        }

    def _evaluate_common_tags(self, rec, red_item_count=0):
        tags = []
        game_result = _safe_int(rec.get("game_result"))
        if _safe_int(rec.get("kill_player")) > 3 and game_result == 0:
            tags.append("大杀四方")
        gained_price = _safe_int(rec.get("gained_price"))
        if gained_price > 1_000_000 and game_result == 0:
            tags.append("百万撤离")
        if gained_price > 3_000_000 and game_result == 0:
            tags.append("盆满钵满")
        equip_price = _safe_int(rec.get("original_equipment_price"))
        if equip_price > 2_000_000:
            tags.append("猛攻哥")
        if _safe_int(rec.get("duration_s")) < 120 and game_result == 1:
            tags.append("落地成盒")
        if _safe_int(rec.get("has_blue_box")):
            tags.append("砖厂老板")
        if (
            _is_non_normal_map(rec.get("map_name"))
            and equip_price < 300_000
            and gained_price > 1_000_000
            and game_result == 0
        ):
            tags.append("以小博大")
        if red_item_count > 1:
            tags.append("再吃亿点")
        if equip_price > 2_000_000 and game_result == 0 and _safe_int(rec.get("profit_loss")) < 500_000:
            tags.append("打白工")
        return tags

    def _build_streak_tags(self, player_id):
        where = ""
        args = []
        if player_id:
            where = " WHERE player_id = ?"
            args.append(player_id)
        rows = self.db.conn.execute(
            f"""SELECT room_id, game_result
                FROM Record
                {where}
                ORDER BY event_time ASC, room_id ASC""",
            args,
        ).fetchall()
        result = defaultdict(list)
        start = 0
        while start < len(rows):
            current = _safe_int(rows[start]["game_result"], -1)
            end = start + 1
            while end < len(rows) and _safe_int(rows[end]["game_result"], -2) == current:
                end += 1
            length = end - start
            if current == 1 and length >= 5:
                for idx in range(start + 4, end):
                    result[rows[idx]["room_id"]].append("连跪")
            elif current == 0 and length >= 3:
                for idx in range(start + 2, end):
                    result[rows[idx]["room_id"]].append("手感火热")
            start = end
        return result

    def _api_stats(self):
        pid = self.db.get_active_player_id()
        stats = self.db.get_battle_stats(player_id=pid)
        stats["detail_records"] = self.db.get_detail_count(player_id=pid)
        stats["room_detail_records"] = self.db.get_battle_detail_count_distinct(
            player_id=pid
        )
        try:
            if pid:
                item_count = self.db.conn.execute(
                    "SELECT COUNT(*) FROM battles_items bi JOIN Record r ON bi.room_id = r.room_id WHERE r.player_id = ?",
                    (pid,),
                ).fetchone()[0]
            else:
                item_count = self.db.conn.execute(
                    "SELECT COUNT(*) FROM battles_items"
                ).fetchone()[0]
        except Exception:
            item_count = 0
        stats["total_items"] = item_count
        stats["total_records"] = stats.get("total", 0) or 0
        acc = self.db.get_active_account()
        if acc:
            stats["active_account"] = acc.get("player_name", "") or acc.get(
                "player_id", ""
            )
            stats["active_player_id"] = acc.get("player_id", "")
        else:
            stats["active_account"] = ""
            stats["active_player_id"] = ""
        stats["logged_in"] = vault_store.has_cookies()
        stats["software_session"] = _open_source_session_snapshot()
        self._json(stats)

    def _api_records(self, params):
        page = int(params.get("page", [1])[0])
        size = int(params.get("size", [20])[0])
        map_id = params.get("map_id", [""])[0]
        game_result = params.get("game_result", [""])[0]
        start_dt = params.get("start", [""])[0]
        end_dt = params.get("end", [""])[0]

        pid = self.db.get_active_player_id()

        where_parts = []
        where_args = []
        if pid:
            where_parts.append("r.player_id = ?")
            where_args.append(pid)
        if map_id:
            where_parts.append("r.map_id = ?")
            where_args.append(int(map_id))
        if game_result != "":
            where_parts.append("r.game_result = ?")
            where_args.append(int(game_result))
        if start_dt:
            where_parts.append("r.event_time >= ?")
            where_args.append(start_dt)
        if end_dt:
            where_parts.append("r.event_time <= ?")
            where_args.append(end_dt)

        where = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""

        total = self.db.conn.execute(
            f"SELECT COUNT(*) FROM Record r{where}", where_args
        ).fetchone()[0]

        offset = (page - 1) * size
        rows = self.db.conn.execute(
            f"""SELECT r.*, m.map_name, rl.role_name,
                (SELECT COUNT(*) FROM battles_items bi WHERE bi.room_id = r.room_id) as item_count
               FROM Record r
               LEFT JOIN MapID m ON r.map_id = m.map_id
               LEFT JOIN Role rl ON r.role_id = rl.role_id
               {where}
               ORDER BY r.event_time DESC
               LIMIT ? OFFSET ?""",
            where_args + [size, offset],
        ).fetchall()

        rule_rows, rules_by_name = self._get_battle_tag_rules()
        streak_tags = self._build_streak_tags(pid)
        row_dicts = [dict(r) for r in rows]
        room_ids = [row["room_id"] for row in row_dicts]
        red_counts = {}
        if room_ids:
            placeholders = ",".join("?" * len(room_ids))
            args = list(room_ids)
            red_where = f"room_id IN ({placeholders}) AND grade = 6"
            if pid:
                red_where += " AND player_id = ?"
                args.append(pid)
            red_rows = self.db.conn.execute(
                f"""SELECT room_id, SUM(COALESCE(num, 0)) as red_count
                    FROM battles_items
                    WHERE {red_where}
                    GROUP BY room_id""",
                args,
            ).fetchall()
            red_counts = {r["room_id"]: _safe_int(r["red_count"]) for r in red_rows}

        for row in row_dicts:
            tag_names = self._evaluate_common_tags(
                row,
                red_counts.get(row["room_id"], 0),
            )
            tag_names.extend(streak_tags.get(row["room_id"], []))
            row["tags"] = [
                self._make_tag_payload(name, rules_by_name) for name in tag_names
            ]

        self._json(
            {
                "page": page,
                "size": size,
                "total": total,
                "tag_rules": rule_rows,
                "data": row_dicts,
            }
        )

    def _api_items(self, params):
        pid = self.db.get_active_player_id()
        start_dt = params.get("start", [""])[0]
        end_dt = params.get("end", [""])[0]
        owner_scope = params.get("owner_scope", ["all"])[0]

        where_parts = []
        where_args = []
        if pid:
            where_parts.append("r.player_id = ?")
            where_args.append(pid)
            if owner_scope == "self":
                where_parts.append("bi.player_id = ?")
                where_args.append(pid)
            elif owner_scope == "teammate":
                where_parts.append("bi.player_id IS NOT NULL")
                where_parts.append("bi.player_id != ?")
                where_args.append(pid)
        if start_dt:
            where_parts.append("r.event_time >= ?")
            where_args.append(start_dt)
        if end_dt:
            where_parts.append("r.event_time <= ?")
            where_args.append(end_dt)

        where = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""

        rows = self.db.conn.execute(
            f"""SELECT bi.*, r.event_time, m.map_name,
                       COALESCE(
                           bd.player_name,
                           CASE WHEN bi.player_id = r.player_id THEN r.player_name END,
                           bi.player_id
                       ) as owner_name,
                       COALESCE(
                           rl_bd.role_name,
                           CASE WHEN bi.player_id = r.player_id THEN rl_self.role_name END
                       ) as role_name
               FROM battles_items bi
               JOIN Record r ON bi.room_id = r.room_id
               LEFT JOIN BattleDetail bd ON bd.room_id = bi.room_id AND bd.player_id = bi.player_id
               LEFT JOIN MapID m ON r.map_id = m.map_id
               LEFT JOIN Role rl_bd ON bd.armed_force_id = rl_bd.role_id
               LEFT JOIN Role rl_self ON r.role_id = rl_self.role_id
               {where}
               ORDER BY r.event_time DESC
               LIMIT 500""",
            where_args,
        ).fetchall()
        self._json([dict(r) for r in rows])

    def _api_maps(self):
        rows = self.db.conn.execute(
            "SELECT map_id, map_name FROM MapID ORDER BY map_id"
        ).fetchall()
        self._json([dict(r) for r in rows])

    def _api_assets(self, params):
        try:
            _require_open_source_feature("assets")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json(
                {"ok": False, "error": str(e), "requires_software_login": True}, 403
            )
            return
        pid = self.db.get_active_player_id()
        category_code = params.get("category", ["operator"])[0].strip() or "operator"
        grade = params.get("grade", [""])[0].strip()
        collectible_only = params.get("collectible_only", ["0"])[0].strip().lower() in {"1", "true", "yes", "on"}
        rows = self.db.get_collection_items(
            player_id=pid,
            category_code=category_code,
            grade=grade,
            collectible_only=collectible_only,
        )
        summary = self.db.get_collection_summary(player_id=pid)
        counts = self.db.get_collection_category_counts(player_id=pid)
        self._json(
            {
                "summary": summary,
                "rows": rows,
                "filters": {
                    "categories": [
                        {"code": "operator", "name": "干员", "count": counts.get("operator", 0)},
                        {"code": "gun", "name": "枪械", "count": counts.get("gun", 0)},
                        {"code": "dagger", "name": "近战", "count": counts.get("dagger", 0)},
                        {"code": "vehicle", "name": "载具", "count": counts.get("vehicle", 0)},
                        {"code": "pendant", "name": "挂饰", "count": counts.get("pendant", 0)},
                    ],
                    "grades": [
                        {"value": "", "label": "全部"},
                        {"value": "6", "label": "红"},
                        {"value": "5", "label": "传说"},
                        {"value": "4", "label": "史诗"},
                        {"value": "3", "label": "稀有"},
                        {"value": "2", "label": "普通"},
                        {"value": "1", "label": "白"},
                    ],
                },
                "catalog_source": (
                    "override"
                    if config.COLLECTIBLE_OBJECT_OVERRIDE_JSON.exists()
                    else "builtin"
                ),
            }
        )

    def _api_assets_refresh(self):
        try:
            _require_open_source_feature("assets")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json(
                {"ok": False, "error": str(e), "requires_software_login": True}, 403
            )
            return
        if not self._try_acquire_action():
            return
        try:
            catalog_result = _refresh_collectible_catalog()
            result = _refresh_collection(self.db)
            self._json({"ok": True, **result, **{
                "catalog_updated": catalog_result.get("updated", False),
                "item_count": catalog_result.get("item_count", 0),
                "catalog_source": catalog_result.get("source", "unknown"),
            }})
        except Exception as e:
            self._json({"ok": False, "error": str(e)}, 400)
        finally:
            self.action_lock.release()

    def _api_assets_catalog_refresh(self):
        try:
            _require_open_source_feature("assets")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json(
                {"ok": False, "error": str(e), "requires_software_login": True}, 403
            )
            return
        if not self._try_acquire_action():
            return
        try:
            result = _refresh_collectible_catalog()
            self._json({"ok": True, **result})
        except Exception as e:
            self._json({"ok": False, "error": str(e)}, 400)
        finally:
            self.action_lock.release()

    def _api_software_session(self):
        self._json({"ok": True, **_open_source_session_snapshot()})

    def _api_ai_config(self):
        try:
            self._json({"ok": True, "config": _read_ai_config_for_editor()})
        except Exception as e:
            self._json({"ok": False, "error": str(e)}, 400)

    def _api_ai_config_save(self):
        try:
            body = self._read_json_body()
            config_data = _save_ai_config(body)
            self._json({"ok": True, "config": config_data})
        except Exception as e:
            self._json({"ok": False, "error": str(e)}, 400)

    def _api_software_register(self):
        self._json({"ok": True, **_open_source_session_snapshot()})

    def _api_software_login(self):
        self._json({"ok": True, **_open_source_session_snapshot()})

    def _api_software_logout(self):
        self._json({"ok": True, **_open_source_session_snapshot()})

    def _api_software_bind_current(self):
        account = self.db.get_active_account()
        if not account:
            self._json({"ok": False, "error": "当前没有已选中的游戏账号"}, 400)
            return
        self._json(
            {
                "ok": True,
                "already_bound": True,
                "binding": _account_availability_status(account),
                "software_session": _open_source_session_snapshot(),
            }
        )

    def _api_accounts(self):
        accounts = self.db.get_all_accounts()
        snapshot = _open_source_session_snapshot()
        for account in accounts:
            account["account_status"] = _account_availability_status(
                account, snapshot
            )
        self._json(
            {
                "accounts": accounts,
                "active_player_id": self.db.get_active_player_id() or "",
                "logged_in": vault_store.has_cookies(),
                "software_session": snapshot,
            }
        )

    def _api_action_status(self, params):
        job_id = params.get("job_id", [""])[0]
        if not job_id:
            self._json({"ok": False, "error": "missing job_id"}, 400)
            return
        job = _get_action_job(job_id)
        if not job:
            self._json({"ok": False, "error": "任务不存在或已过期"}, 404)
            return
        self._json({"ok": True, **job})

    def _api_detail(self, params):
        room_id = params.get("room_id", [""])[0]
        if not room_id:
            self._json({"self": None, "teammates": [], "items": []})
            return

        self_row = self.db.conn.execute(
            """SELECT bd.*, m.map_name, rl.role_name
               FROM BattleDetail bd
               LEFT JOIN MapID m ON bd.map_id = m.map_id
               LEFT JOIN Role rl ON bd.armed_force_id = rl.role_id
               WHERE bd.room_id = ? AND bd.is_self = 1""",
            (room_id,),
        ).fetchone()

        if not self_row:
            any_row = self.db.conn.execute(
                "SELECT room_id FROM BattleDetail WHERE room_id = ?", (room_id,)
            ).fetchone()
            rec_row = self.db.conn.execute(
                "SELECT * FROM Record WHERE room_id = ?", (room_id,)
            ).fetchone()
            if not any_row and not rec_row:
                self._json(
                    {
                        "self": dict(rec_row) if rec_row else None,
                        "teammates": [],
                        "items": [],
                    }
                )
                return
            if not self_row and rec_row:
                self_row = rec_row
            elif not self_row:
                self_row = self.db.conn.execute(
                    """SELECT bd.*, m.map_name, rl.role_name
                       FROM BattleDetail bd
                       LEFT JOIN MapID m ON bd.map_id = m.map_id
                       LEFT JOIN Role rl ON bd.armed_force_id = rl.role_id
                       WHERE bd.room_id = ? LIMIT 1""",
                    (room_id,),
                ).fetchone()

        tm_rows = self.db.conn.execute(
            """SELECT bd.*, m.map_name, rl.role_name
               FROM BattleDetail bd
               LEFT JOIN MapID m ON bd.map_id = m.map_id
               LEFT JOIN Role rl ON bd.armed_force_id = rl.role_id
               WHERE bd.room_id = ? AND bd.is_self = 0""",
            (room_id,),
        ).fetchall()

        items_rows = self.db.conn.execute(
            """SELECT bi.*, bd.player_name
               FROM battles_items bi
               LEFT JOIN BattleDetail bd ON bi.room_id = bd.room_id AND bi.player_id = bd.player_id
               WHERE bi.room_id = ?""",
            (room_id,),
        ).fetchall()

        rule_rows, rules_by_name = self._get_battle_tag_rules()
        items_list = [dict(r) for r in items_rows]
        red_counts = defaultdict(int)
        for item in items_list:
            if _safe_int(item.get("grade")) == 6:
                key = item.get("player_id") or item.get("player_name") or ""
                red_counts[key] += _safe_int(item.get("num"), 1)

        def with_tags(row):
            if not row:
                return None
            data = dict(row)
            player_key = data.get("player_id") or data.get("player_name") or ""
            tag_names = self._evaluate_common_tags(data, red_counts.get(player_key, 0))
            data["tags"] = [
                self._make_tag_payload(name, rules_by_name) for name in tag_names
            ]
            return data

        self._json(
            {
                "tag_rules": rule_rows,
                "self": with_tags(self_row),
                "teammates": [with_tags(r) for r in tm_rows],
                "items": items_list,
            }
        )

    def _get_active_team_player_name(self, player_id: str | None = None) -> str:
        player_id = player_id or self.db.get_active_player_id()
        account = self.db.get_active_account() or {}
        account_name = str(account.get("player_name", "") or "").strip()
        if account_name:
            return account_name
        if player_id:
            row = self.db.conn.execute(
                """SELECT player_name
                   FROM BattleDetail
                   WHERE player_id = ? AND player_name IS NOT NULL AND player_name != ''
                   ORDER BY event_time DESC
                   LIMIT 1""",
                (player_id,),
            ).fetchone()
            if row and row["player_name"]:
                return str(row["player_name"])
        row = self.db.conn.execute(
            """SELECT player_name
               FROM BattleDetail
               WHERE is_self = 1 AND player_name IS NOT NULL AND player_name != ''
               ORDER BY event_time DESC
               LIMIT 1"""
        ).fetchone()
        return str(row["player_name"]) if row and row["player_name"] else ""

    def _get_scoped_record_room_ids(
        self, player_id: str | None = None, start_dt: str = "", end_dt: str = ""
    ) -> list[str]:
        where_parts = []
        args: list[str] = []
        if player_id:
            where_parts.append("player_id = ?")
            args.append(player_id)
        if start_dt:
            where_parts.append("event_time >= ?")
            args.append(start_dt)
        if end_dt:
            where_parts.append("event_time <= ?")
            args.append(end_dt)
        sql = "SELECT DISTINCT room_id FROM Record"
        if where_parts:
            sql += " WHERE " + " AND ".join(where_parts)
        rows = self.db.conn.execute(sql, args).fetchall()
        return [str(r["room_id"]) for r in rows if r["room_id"]]

    def _aggregate_other_player_record(
        self, rows: list[dict], display_name: str = "其他玩家"
    ) -> dict | None:
        if not rows:
            return None
        role_counts: dict[str, int] = defaultdict(int)
        out = {
            "player_key": TEAM_OTHER_TOKEN,
            "player_id": "",
            "player_name": display_name,
            "armed_force_id": "",
            "game_result": rows[0].get("game_result"),
            "kill_player": 0,
            "rescue": 0,
            "original_equipment_price": 0,
            "gained_price": 0,
            "profit_loss": 0,
        }
        for row in rows:
            role_id = str(row.get("armed_force_id", "") or "")
            if role_id:
                role_counts[role_id] += 1
            out["kill_player"] += int(row.get("kill_player", 0) or 0)
            out["rescue"] += int(row.get("rescue", 0) or 0)
            out["original_equipment_price"] += int(
                row.get("original_equipment_price", 0) or 0
            )
            out["gained_price"] += int(row.get("gained_price", 0) or 0)
            out["profit_loss"] += int(row.get("profit_loss", 0) or 0)
        if role_counts:
            out["armed_force_id"] = max(role_counts.items(), key=lambda item: item[1])[0]
        return out

    def _api_players(self, params):
        try:
            _require_open_source_feature("team_analysis")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        pid = self.db.get_active_player_id()
        active_name = self._get_active_team_player_name(pid)
        start_dt = params.get("start", [""])[0]
        end_dt = params.get("end", [""])[0]
        room_ids = self._get_scoped_record_room_ids(pid, start_dt, end_dt)
        if not room_ids:
            self._json(
                {
                    "players": [],
                    "active_player_name": active_name,
                    "active_player_id": pid or "",
                }
            )
            return

        placeholders = ",".join("?" for _ in room_ids)
        rows = self.db.conn.execute(
            f"""SELECT room_id, player_name, player_id
                FROM BattleDetail
                WHERE room_id IN ({placeholders})
                ORDER BY event_time DESC""",
            room_ids,
        ).fetchall()
        room_players: dict[str, list[dict]] = defaultdict(list)
        for row in rows:
            room_players[str(row["room_id"])].append(dict(row))

        filter_targets, _ = _parse_team_player_targets(params.get("with", [""])[0])
        if filter_targets:
            room_ids = [
                rid
                for rid, members in room_players.items()
                if len(_resolve_team_targets_for_room(members, filter_targets))
                == len(filter_targets)
            ]
            if not room_ids:
                self._json(
                    {
                        "players": [],
                        "active_player_name": active_name,
                        "active_player_id": pid or "",
                    }
                )
                return

        players_by_key: dict[str, dict] = {}
        for rid in room_ids:
            seen_keys: set[str] = set()
            for member in room_players.get(rid, []):
                pname = sanitize_player_name(member.get("player_name"))
                player_id = str(member.get("player_id", "") or "").strip()
                key = _team_player_key(player_id, pname)
                if not pname or key in seen_keys:
                    continue
                seen_keys.add(key)
                item = players_by_key.setdefault(
                    key,
                    {
                        "key": key,
                        "name": pname,
                        "player_name": pname,
                        "player_id": player_id,
                        "count": 0,
                        "is_self": False,
                    },
                )
                item["count"] += 1
                if player_id and player_id == str(pid or ""):
                    item["is_self"] = True
                if not player_id and active_name and pname == active_name:
                    item["is_self"] = True

        duplicate_counts: dict[str, int] = defaultdict(int)
        for item in players_by_key.values():
            duplicate_counts[item["name"]] += 1
        players = list(players_by_key.values())
        for item in players:
            item["duplicate_name"] = duplicate_counts[item["name"]] > 1
        players.sort(
            key=lambda item: (
                0 if item["is_self"] else 1,
                -int(item["count"]),
                item["name"],
                item["player_id"],
            )
        )
        self._json(
            {
                "players": players,
                "active_player_name": active_name,
                "active_player_id": pid or "",
                "active_player_key": _team_player_key(pid, active_name) if (pid or active_name) else "",
            }
        )

    def _build_team_analysis_data(
        self, players_param: str, start_dt: str = "", end_dt: str = ""
    ):
        if not players_param:
            return {"error": "请先选择玩家"}
        targets, use_other_slot = _parse_team_player_targets(players_param)
        if not targets:
            return {"error": "无效的玩家名"}
        if use_other_slot and len(targets) != 2:
            use_other_slot = False

        pid = self.db.get_active_player_id()
        active_name = self._get_active_team_player_name(pid)
        active_key = _team_player_key(pid, active_name) if (pid or active_name) else ""
        room_ids = self._get_scoped_record_room_ids(pid, start_dt, end_dt)
        if not room_ids:
            return {"error": "暂无数据"}

        placeholders = ",".join("?" for _ in room_ids)
        rows = self.db.conn.execute(
            f"""SELECT bd.*, m.map_name as bd_map_name, rl.role_name as bd_role_name
               FROM BattleDetail bd
               LEFT JOIN MapID m ON bd.map_id = m.map_id
               LEFT JOIN Role rl ON bd.armed_force_id = rl.role_id
               WHERE bd.room_id IN ({placeholders})
               ORDER BY bd.event_time ASC, bd.is_self DESC""",
            room_ids,
        ).fetchall()
        target_meta: dict[str, dict] = {
            str(t["key"]): {
                "player_key": str(t["key"]),
                "player_id": str(t.get("player_id", "") or ""),
                "player_name": str(t.get("player_name", "") or ""),
            }
            for t in targets
        }
        bd_by_room: dict[str, list[dict]] = defaultdict(list)
        for r in rows:
            rid = str(r["room_id"])
            bd = dict(r)
            bd_by_room[rid].append(bd)

        target_keys = {str(t["key"]) for t in targets}
        resolved_by_room: dict[str, dict[str, dict]] = {}
        for rid, members in bd_by_room.items():
            resolved = _resolve_team_targets_for_room(members, targets)
            resolved_by_room[rid] = resolved
            for matched_key, bd in resolved.items():
                meta = target_meta.setdefault(
                    matched_key,
                    {"player_key": matched_key, "player_id": "", "player_name": ""},
                )
                if not meta.get("player_id") and bd.get("player_id"):
                    meta["player_id"] = str(bd.get("player_id") or "")
                if not meta.get("player_name") and bd.get("player_name"):
                    meta["player_name"] = sanitize_player_name(bd.get("player_name"))
        target_rooms = [
            rid for rid, resolved in resolved_by_room.items() if target_keys.issubset(resolved.keys())
        ]
        if not target_rooms:
            return {"error": "这些玩家没有共同参与的对局"}

        tr_placeholders = ",".join("?" for _ in target_rooms)

        rec_rows = self.db.conn.execute(
            f"""SELECT r.*, m.map_name, rl.role_name
               FROM Record r
               LEFT JOIN MapID m ON r.map_id = m.map_id
               LEFT JOIN Role rl ON r.role_id = rl.role_id
               WHERE r.room_id IN ({tr_placeholders})
               ORDER BY r.event_time ASC""",
            target_rooms,
        ).fetchall()

        rec_map = {r["room_id"]: dict(r) for r in rec_rows}
        room_order = [str(r["room_id"]) for r in rec_rows]
        high_value_item_rows = self.db.conn.execute(
            f"""SELECT bi.room_id,
                       bi.item_name,
                       bi.num,
                       bi.price,
                       bi.player_id,
                       r.event_time,
                       m.map_name,
                       COALESCE(
                           bd.player_name,
                           CASE WHEN bi.player_id = r.player_id THEN r.player_name END,
                           bi.player_id
                       ) as owner_name,
                       COALESCE(
                           rl_bd.role_name,
                           CASE WHEN bi.player_id = r.player_id THEN rl_self.role_name END
                       ) as role_name
                FROM battles_items bi
                JOIN Record r ON bi.room_id = r.room_id
                LEFT JOIN BattleDetail bd ON bd.room_id = bi.room_id AND bd.player_id = bi.player_id
                LEFT JOIN MapID m ON r.map_id = m.map_id
                LEFT JOIN Role rl_bd ON bd.armed_force_id = rl_bd.role_id
                LEFT JOIN Role rl_self ON r.role_id = rl_self.role_id
                WHERE bi.room_id IN ({tr_placeholders})
                  AND CAST(COALESCE(bi.price, 0) AS INTEGER) > 1000000
                ORDER BY CAST(COALESCE(bi.price, 0) AS INTEGER) DESC, r.event_time DESC""",
            target_rooms,
        ).fetchall()

        player_order = [str(t["key"]) for t in targets]
        active_order_key = active_key if active_key in player_order else ""
        if not active_order_key and active_name:
            active_order_key = next(
                (
                    key
                    for key in player_order
                    if sanitize_player_name(target_meta.get(key, {}).get("player_name"))
                    == active_name
                ),
                "",
            )
        if active_order_key:
            player_order = [active_order_key] + [
                key for key in player_order if key != active_order_key
            ]
        if use_other_slot:
            player_order.append(TEAM_OTHER_TOKEN)

        def player_display_name(slot_key: str) -> str:
            if slot_key == TEAM_OTHER_TOKEN:
                return "其他玩家"
            meta = target_meta.get(slot_key, {})
            return sanitize_player_name(meta.get("player_name")) or str(
                meta.get("player_id") or slot_key
            )

        player_id_to_target_key = {
            str(meta.get("player_id") or ""): key
            for key, meta in target_meta.items()
            if meta.get("player_id")
        }
        high_value_items = []
        for row in high_value_item_rows:
            item = dict(row)
            owner_name = str(item.get("owner_name") or "")
            item_player_id = str(item.get("player_id", "") or "")
            owner = {
                "player_id": item_player_id,
                "player_name": owner_name,
            }
            owner_key = player_id_to_target_key.get(item_player_id) or _match_team_target(owner, targets)
            if owner_key:
                owner_name = player_display_name(owner_key)
            elif use_other_slot:
                owner_key = TEAM_OTHER_TOKEN
                owner_name = "其他玩家"
            else:
                continue
            if owner_key not in player_order:
                continue
            high_value_items.append(
                {
                    "event_time": item.get("event_time", ""),
                    "room_id": str(item.get("room_id") or ""),
                    "map_name": item.get("map_name", ""),
                    "player_key": owner_key,
                    "player_id": target_meta.get(owner_key, {}).get("player_id", ""),
                    "player_name": owner_name,
                    "role_name": item.get("role_name", "") or "",
                    "item_name": item.get("item_name", "") or "-",
                    "num": _safe_int(item.get("num"), 1),
                    "price": _safe_int(item.get("price")),
                    "total_price": _safe_int(item.get("price")) * max(_safe_int(item.get("num"), 1), 1),
                }
            )

        games = []
        all_player_stats: dict[str, dict] = {}
        total_kills = 0
        total_escaped = 0
        total_profit = 0
        total_duration = 0
        max_collection = 0
        max_profit = 0
        all_maps: list[str] = []

        for rid in room_order:
            rec = rec_map.get(rid, {})
            bds = bd_by_room.get(rid, [])
            game = {
                "room_id": rid,
                "event_time": rec.get("event_time", ""),
                "map_name": rec.get("map_name", ""),
                "duration_s": rec.get("duration_s", 0),
                "game_result": rec.get("game_result"),
                "player_records": {},
            }

            if rec.get("game_result") == 0:
                total_escaped += 1
            total_kills += rec.get("kill_player", 0) or 0
            pl = int(rec.get("profit_loss", 0) or 0)
            total_profit += pl
            dur = rec.get("duration_s", 0) or 0
            total_duration += dur
            cp = int(rec.get("gained_price", 0) or 0)
            if cp > max_collection:
                max_collection = cp
            if pl > max_profit:
                max_profit = pl
            all_maps.append(rec.get("map_name", ""))

            resolved_rows = resolved_by_room.get(rid, {})
            relevant_rows: dict[str, dict] = {}
            assigned_row_ids = {id(row) for row in resolved_rows.values()}
            for slot_key, bd in resolved_rows.items():
                row = dict(bd)
                row["player_key"] = slot_key
                relevant_rows[slot_key] = row
            if use_other_slot:
                other_rows = [
                    bd
                    for bd in bds
                    if id(bd) not in assigned_row_ids
                ]
                other_record = self._aggregate_other_player_record(other_rows)
                if other_record:
                    relevant_rows[TEAM_OTHER_TOKEN] = other_record

            for slot_name, bd in relevant_rows.items():
                display_name = player_display_name(slot_name)
                player_id = str(
                    bd.get("player_id")
                    or target_meta.get(slot_name, {}).get("player_id", "")
                    or ""
                )
                bd["player_key"] = slot_name
                bd["player_id"] = player_id
                bd["player_name"] = display_name
                if slot_name not in all_player_stats:
                    all_player_stats[slot_name] = {
                        "player_key": slot_name,
                        "player_id": player_id,
                        "player_name": display_name,
                        "games": 0,
                        "escaped": 0,
                        "kill_player": 0,
                        "rescue": 0,
                        "eq": 0,
                        "out": 0,
                        "profit": 0,
                        "roles": [],
                    }
                ps = all_player_stats[slot_name]
                ps["games"] += 1
                if bd.get("game_result") == 0:
                    ps["escaped"] += 1
                ps["kill_player"] += bd.get("kill_player", 0) or 0
                ps["rescue"] += bd.get("rescue", 0) or 0
                ps["eq"] += int(bd.get("original_equipment_price", 0) or 0)
                ps["out"] += int(bd.get("gained_price", 0) or 0)
                ps["profit"] += int(bd.get("profit_loss", 0) or 0)
                if bd.get("armed_force_id"):
                    ps["roles"].append(str(bd["armed_force_id"]))
                game["player_records"][slot_name] = bd

            games.append(game)

        n = len(games)
        from collections import Counter

        common_map = Counter(all_maps).most_common(1)[0][0] if all_maps else ""

        players_out = []
        for pname in player_order:
            if pname not in all_player_stats:
                continue
            ps = all_player_stats[pname]
            g = ps["games"]
            er = ps["escaped"] / g if g > 0 else 0
            fail = g - ps["escaped"]
            kd = ps["kill_player"] / fail if fail > 0 else float(ps["kill_player"])
            avg_eq = ps["eq"] / g if g > 0 else 0
            avg_rate_num = ps["profit"] / ps["eq"] if ps["eq"] > 0 else 0
            common_role_id = (
                Counter(ps["roles"]).most_common(1)[0][0] if ps["roles"] else ""
            )
            players_out.append(
                {
                    "player_key": ps["player_key"],
                    "player_id": ps["player_id"],
                    "player_name": ps["player_name"],
                    "games": g,
                    "evac_rate_num": er,
                    "evac_rate": f"{er * 100:.1f}%",
                    "kd": kd,
                    "avg_pk": ps["kill_player"] / g if g > 0 else 0,
                    "avg_eq": avg_eq,
                    "total_out": ps["out"],
                    "total_profit": ps["profit"],
                    "avg_rate_num": avg_rate_num,
                    "avg_rate": f"{avg_rate_num * 100:.1f}%",
                    "common_role": ROLE_MAP.get(common_role_id, common_role_id),
                }
            )

        games_out = []
        for game in games:
            prs = []
            for pname in player_order:
                bd = game["player_records"].get(pname)
                prs.append(bd)
            games_out.append(
                {
                    "room_id": game["room_id"],
                    "event_time": game["event_time"],
                    "map_name": game["map_name"],
                    "duration_s": game["duration_s"],
                    "game_result": game["game_result"],
                    "player_records": prs,
                }
            )

        if players_out:
            avg_kd = sum(p["kd"] for p in players_out) / len(players_out)
            kill_king = max(players_out, key=lambda p: p["kd"])
            profit_king = max(players_out, key=lambda p: p["total_profit"])
        else:
            avg_kd = 0
            kill_king = {}
            profit_king = {}

        return {
            "summary": {
                "total_games": n,
                "total_escaped": total_escaped,
                "evac_rate": f"{total_escaped / n * 100:.1f}%" if n else "0%",
                "total_kills": total_kills,
                "total_profit": total_profit,
                "avg_duration": total_duration / n if n else 0,
                "common_map": common_map,
                "team_kd": round(avg_kd, 2),
                "kill_king": kill_king.get("player_name", ""),
                "kill_king_kd": kill_king.get("kd", 0),
                "profit_king": profit_king.get("player_name", ""),
                "profit_king_val": profit_king.get("total_profit", 0),
            },
            "players": players_out,
            "games": games_out,
            "high_value_items": high_value_items,
        }

    def _api_team_analysis(self, params):
        try:
            _require_open_source_feature("team_analysis")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        players_param = params.get("players", [""])[0]
        start_dt = params.get("start", [""])[0]
        end_dt = params.get("end", [""])[0]
        self._json(self._build_team_analysis_data(players_param, start_dt, end_dt))

    def _api_team_report_pdf(self):
        data = self._read_json_body()
        players_param = str(data.get("players", "") or "").strip()
        players_label = str(data.get("players_label", "") or "").strip()
        start_dt = str(data.get("start", "") or "").strip()
        end_dt = str(data.get("end", "") or "").strip()
        if not players_param:
            self._json({"ok": False, "error": "请先选择玩家并生成组队分析"}, 400)
            return
        try:
            ai_cfg = _load_ai_config()
            ai_cfg["mode"] = "local"
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e)}, 400)
            return
        job = _create_team_report_job(ai_cfg, players_param, start_dt, end_dt, players_label)
        self._json({"ok": True, "job_id": job["id"], "job": job})

    def _api_team_report_status(self, params):
        job_id = params.get("job_id", [""])[0]
        if not job_id:
            self._json({"ok": False, "error": "missing job_id"}, 400)
            return
        reports = {report["id"]: report for report in _list_team_reports()}
        job = reports.get(job_id)
        if not job:
            self._json({"ok": False, "error": "报告任务不存在或已过期"}, 404)
            return
        self._json({"ok": True, **job})

    def _api_team_report_list(self):
        self._json({"ok": True, "reports": _list_team_reports()})

    def _api_team_report_cancel(self):
        data = self._read_json_body()
        job_id = str(data.get("job_id", "") or "").strip()
        if not job_id:
            self._json({"ok": False, "error": "missing job_id"}, 400)
            return
        result = _cancel_team_report(job_id)
        self._json(result, 200 if result.get("ok") else 400)

    def _api_team_report_delete(self):
        data = self._read_json_body()
        job_id = str(data.get("job_id", "") or "").strip()
        if not job_id:
            self._json({"ok": False, "error": "missing job_id"}, 400)
            return
        result = _delete_team_report(job_id)
        self._json(result, 200 if result.get("ok") else 400)

    def _api_team_report_download(self, params):
        job_id = params.get("job_id", [""])[0]
        if not job_id:
            self._json({"error": "missing job_id"}, 400)
            return
        report = _get_team_report_file_info(job_id)
        if not report:
            self._json({"error": "PDF 报告不存在或已过期，请重新生成"}, 404)
            return
        body = report["path"].read_bytes()
        filename = report["filename"]
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _api_export_team(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._json(
                {
                    "error": "导出 Excel 依赖未安装，请执行 `python -m pip install openpyxl` 后重试"
                },
                500,
            )
            return

        from .config import ROLE_MAP as RM
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length)
        data = json.loads(body)

        analysis = self._build_team_analysis_data(
            str(data.get("players", "") or ""),
            str(data.get("start", "") or ""),
            str(data.get("end", "") or ""),
        )
        if analysis.get("error"):
            self._json(analysis, 400)
            return

        summary = analysis.get("summary", {})
        players = analysis.get("players", [])
        games = analysis.get("games", [])

        if not games:
            self._json({"error": "无数据"})
            return

        GC = ["FFE0E0", "FFF0D0", "FFFFD0", "D0F0D0", "B0E8B0"]

        def grad_color(val, mn, mx):
            if mn == mx:
                return GC[2]
            r = (val - mn) / (mx - mn)
            i = min(max(int(r * 5), 0), 4)
            return GC[i]

        wb = Workbook()
        ws = wb.active
        ws.title = "组队分析"

        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold = Font(bold=True)
        red_fill = PatternFill("solid", fgColor="F5D4D4")
        green_fill = PatternFill("solid", fgColor="D4F5E0")

        col_start = 4
        for i in range(len(games)):
            ws.column_dimensions[get_column_letter(col_start + i)].width = 18
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12

        def sc(r, c, v, **kw):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = center
            cell.border = thin_border
            if "font" in kw:
                cell.font = kw["font"]
            if "fill" in kw and kw["fill"] is not None:
                cell.fill = kw["fill"]
            if "fmt" in kw:
                cell.number_format = kw["fmt"]
            return cell

        row = 1
        sc(row, 1, "汇总", font=bold)
        sc(row, 2, "汇总数据", font=bold)
        sc(row, 3, "明细", font=bold)
        for i, g in enumerate(games):
            sc(row, col_start + i, g.get("event_time", ""))

        row += 1
        sc(row, 1, "常用地图", font=bold)
        sc(row, 2, summary.get("common_map", ""))
        sc(row, 3, "地图", font=bold)
        for i, g in enumerate(games):
            sc(row, col_start + i, g.get("map_name", ""))

        row += 1
        sc(row, 1, "平均时长", font=bold)
        avg_dur = summary.get("avg_duration", 0)
        m_, s_ = divmod(int(avg_dur), 60)
        sc(row, 2, f"{m_}分{s_}秒")
        sc(row, 3, "时长", font=bold)
        durs = [g.get("duration_s", 0) or 0 for g in games]
        min_d, max_d = min(durs), max(durs)
        for i, g in enumerate(games):
            dur = g.get("duration_s", 0) or 0
            dm, ds = divmod(int(dur), 60)
            sc(
                row,
                col_start + i,
                f"{dm}分{ds}秒",
                fill=PatternFill("solid", fgColor=grad_color(dur, min_d, max_d)),
            )

        row += 1
        sc(row, 1, "撤离率", font=bold)
        sc(row, 2, summary.get("evac_rate", ""))
        sc(row, 3, "结果", font=bold)
        rcls_xlsx = {
            0: ("撤离成功", green_fill),
            1: ("撤离失败", red_fill),
            2: ("行动超时", None),
            3: ("中途退出", None),
        }
        for i, g in enumerate(games):
            gr = g.get("game_result")
            label, fill = rcls_xlsx.get(gr, (str(gr), None))
            sc(row, col_start + i, label, fill=fill)

        row += 1

        def get_pr(pi, g):
            prs = g.get("player_records", [])
            return prs[pi] if pi < len(prs) and prs[pi] else None

        def game_metric_range(g, getter):
            vals = []
            for pi in range(len(players)):
                pr = get_pr(pi, g)
                if not pr:
                    continue
                try:
                    vals.append(float(getter(pr) or 0))
                except Exception:
                    vals.append(0)
            if not vals:
                return 0, 0
            return min(vals), max(vals)

        def game_metric_fill(g, getter, value):
            mn, mx = game_metric_range(g, getter)
            return PatternFill("solid", fgColor=grad_color(value, mn, mx))

        def record_profit_rate(pr):
            eq = int(pr.get("original_equipment_price", 0) or 0)
            pl = int(pr.get("profit_loss", 0) or 0)
            return pl / eq if eq > 0 else 0

        duplicate_names = defaultdict(int)
        for p in players:
            duplicate_names[str(p.get("player_name", "") or "")] += 1

        def player_export_label(p):
            name = str(p.get("player_name", "") or "")
            player_id = str(p.get("player_id", "") or "")
            if name and player_id and duplicate_names[name] > 1:
                return f"{name} (ID {player_id[-6:]})"
            return name

        for pi, p in enumerate(players):
            sc(row, 1, f"玩家{pi + 1}", font=Font(bold=True, size=12))
            sc(row, 2, player_export_label(p))
            sc(row, 3, f"玩家{pi + 1}", font=Font(bold=True, size=12))
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                sc(row, col_start + i, player_export_label(pr) if pr else "-")
            row += 1

            sc(row, 1, "常用角色", font=bold)
            sc(row, 2, p.get("common_role", ""))
            sc(row, 3, "角色", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                role = pr.get("armed_force_id", "") if pr else ""
                sc(row, col_start + i, RM.get(str(role), str(role)) if role else "-")
            row += 1

            sc(row, 1, "撤离率", font=bold)
            sc(row, 2, p.get("evac_rate", ""))
            sc(row, 3, "结果", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                if pr:
                    gr = pr.get("game_result")
                    label, fill = rcls_xlsx.get(gr, (str(gr), None))
                    sc(row, col_start + i, label, fill=fill)
                else:
                    sc(row, col_start + i, "-")
            row += 1

            max_kd = max((pp.get("kd", 0) for pp in players), default=1) or 1

            sc(row, 1, "KD", font=bold)
            sc(
                row,
                2,
                round(p.get("kd", 0), 2),
                fmt="0.00",
                fill=PatternFill(
                    "solid", fgColor=grad_color(p.get("kd", 0), 0, max_kd)
                ),
            )
            sc(row, 3, "击杀玩家", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                if not pr:
                    sc(row, col_start + i, "-")
                    continue
                kp = pr.get("kill_player", 0) or 0
                sc(
                    row,
                    col_start + i,
                    kp,
                    fill=game_metric_fill(g, lambda pr: pr.get("kill_player", 0) or 0, kp),
                )
            row += 1

            sc(row, 1, "带入装备", font=bold)
            sc(row, 2, int(p.get("avg_eq", 0)), fmt="#,##0")
            sc(row, 3, "带入装备", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                if not pr:
                    sc(row, col_start + i, "-")
                    continue
                val = int(pr.get("original_equipment_price", 0) or 0)
                sc(
                    row,
                    col_start + i,
                    val,
                    fmt="#,##0",
                    fill=PatternFill(
                        "solid",
                        fgColor=grad_color(
                            val,
                            *game_metric_range(
                                g,
                                lambda pr: int(
                                    pr.get("original_equipment_price", 0) or 0
                                ),
                            ),
                        ),
                    ),
                )
            row += 1

            max_total_out = (
                max((pp.get("total_out", 0) for pp in players), default=1) or 1
            )
            sc(row, 1, "带出价值", font=bold)
            sc(
                row,
                2,
                int(p.get("total_out", 0)),
                fmt="#,##0",
                fill=PatternFill(
                    "solid", fgColor=grad_color(p.get("total_out", 0), 0, max_total_out)
                ),
            )
            sc(row, 3, "带出价值", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                if not pr:
                    sc(row, col_start + i, "-")
                    continue
                val = int(pr.get("gained_price", 0) or 0)
                sc(
                    row,
                    col_start + i,
                    val,
                    fmt="#,##0",
                    fill=PatternFill(
                        "solid",
                        fgColor=grad_color(
                            val,
                            *game_metric_range(
                                g, lambda pr: int(pr.get("gained_price", 0) or 0)
                            ),
                        ),
                    ),
                )
            row += 1

            min_tp = min((pp.get("total_profit", 0) for pp in players), default=0)
            max_tp = max((pp.get("total_profit", 0) for pp in players), default=0)
            sc(row, 1, "盈亏", font=bold)
            sc(
                row,
                2,
                int(p.get("total_profit", 0)),
                fmt="#,##0",
                fill=PatternFill(
                    "solid",
                    fgColor=grad_color(p.get("total_profit", 0), min_tp, max_tp),
                ),
            )
            sc(row, 3, "盈亏", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                if not pr:
                    sc(row, col_start + i, "-")
                    continue
                val = int(pr.get("profit_loss", 0) or 0)
                sc(
                    row,
                    col_start + i,
                    val,
                    fmt="#,##0",
                    fill=PatternFill(
                        "solid",
                        fgColor=grad_color(
                            val,
                            *game_metric_range(
                                g, lambda pr: int(pr.get("profit_loss", 0) or 0)
                            ),
                        ),
                    ),
                )
            row += 1

            sc(row, 1, "收益率", font=bold)
            sc(row, 2, p.get("avg_rate", ""))
            sc(row, 3, "收益率", font=bold)
            for i, g in enumerate(games):
                pr = get_pr(pi, g)
                if not pr:
                    sc(row, col_start + i, "-")
                    continue
                val = record_profit_rate(pr)
                sc(
                    row,
                    col_start + i,
                    val,
                    fmt="0.0%",
                    fill=PatternFill(
                        "solid",
                        fgColor=grad_color(
                            val, *game_metric_range(g, record_profit_rate)
                        ),
                    ),
                )
            row += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header(
            "Content-Disposition", "attachment; filename=team_analysis.xlsx"
        )
        self.end_headers()
        self.wfile.write(buf.read())

    def _api_trends(self, params):
        try:
            _require_open_source_feature("trend_analysis")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return

        pid = self.db.get_active_player_id()
        bucket = str(params.get("bucket", ["day"])[0] or "day").strip().lower()
        if bucket not in {"day", "week", "month"}:
            bucket = "day"
        range_key = str(params.get("range", ["30"])[0] or "30").strip().lower()
        if range_key not in {"7", "30", "90", "all"}:
            range_key = "30"
        start_dt = str(params.get("start", [""])[0] or "").strip()
        end_dt = str(params.get("end", [""])[0] or "").strip()
        map_id = str(params.get("map_id", [""])[0] or "").strip()

        base_parts = []
        base_args = []
        if pid:
            base_parts.append("r.player_id = ?")
            base_args.append(pid)
        if map_id:
            try:
                base_parts.append("r.map_id = ?")
                base_args.append(int(map_id))
            except ValueError:
                pass

        base_where = (" WHERE " + " AND ".join(base_parts)) if base_parts else ""
        base_range_row = self.db.conn.execute(
            f"SELECT MIN(r.event_time) as min_time, MAX(r.event_time) as max_time FROM Record r{base_where}",
            base_args,
        ).fetchone()
        max_time = str(base_range_row["max_time"] if base_range_row and base_range_row["max_time"] else "")
        min_time = str(base_range_row["min_time"] if base_range_row and base_range_row["min_time"] else "")
        if not start_dt and not end_dt and range_key != "all" and max_time:
            try:
                end_day = datetime.strptime(max_time[:10], "%Y-%m-%d")
                start_day = end_day - timedelta(days=int(range_key) - 1)
                start_dt = start_day.strftime("%Y-%m-%d 00:00:00")
                end_dt = end_day.strftime("%Y-%m-%d 23:59:59")
            except Exception:
                pass

        where_parts = list(base_parts)
        where_args = list(base_args)
        if start_dt:
            where_parts.append("r.event_time >= ?")
            where_args.append(start_dt)
        if end_dt:
            where_parts.append("r.event_time <= ?")
            where_args.append(end_dt)
        where = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""

        summary_row = self.db.conn.execute(
            f"""SELECT
                COUNT(*) as total,
                SUM(CASE WHEN r.game_result=0 THEN 1 ELSE 0 END) as escaped,
                SUM(r.kill_player) as total_player_kills,
                AVG(r.duration_s) as avg_duration,
                AVG(CAST(r.profit_loss AS REAL)) as avg_profit,
                SUM(CAST(r.profit_loss AS REAL)) as total_profit,
                MAX(CAST(r.gained_price AS INTEGER)) as max_collection,
                MAX(CAST(r.profit_loss AS INTEGER)) as max_profit
               FROM Record r{where}""",
            where_args,
        ).fetchone()
        total = _safe_int(summary_row["total"] if summary_row else 0)
        escaped = _safe_int(summary_row["escaped"] if summary_row else 0)
        total_player_kills = _safe_int(summary_row["total_player_kills"] if summary_row else 0)
        fail_count = max(total - escaped, 0)
        summary = {
            "total": total,
            "escaped": escaped,
            "evac_rate_num": round(escaped / total * 100, 1) if total else 0,
            "kd": round(total_player_kills / fail_count, 2) if fail_count > 0 else total_player_kills,
            "total_player_kills": total_player_kills,
            "total_profit": round(float(summary_row["total_profit"] or 0), 0) if summary_row else 0,
            "avg_profit": round(float(summary_row["avg_profit"] or 0), 0) if summary_row else 0,
            "avg_duration": round(float(summary_row["avg_duration"] or 0), 0) if summary_row else 0,
            "max_collection": _safe_int(summary_row["max_collection"] if summary_row else 0),
            "max_profit": _safe_int(summary_row["max_profit"] if summary_row else 0),
        }

        bucket_expr = {
            "day": "date(r.event_time)",
            "week": "strftime('%Y-W%W', r.event_time)",
            "month": "substr(r.event_time, 1, 7)",
        }[bucket]
        trend_rows = self.db.conn.execute(
            f"""SELECT
                {bucket_expr} as bucket,
                MIN(r.event_time) as first_time,
                COUNT(*) as total,
                SUM(CASE WHEN r.game_result=0 THEN 1 ELSE 0 END) as escaped,
                SUM(r.kill_player) as total_player_kills,
                AVG(r.duration_s) as avg_duration,
                AVG(CAST(r.profit_loss AS REAL)) as avg_profit,
                SUM(CAST(r.profit_loss AS REAL)) as total_profit,
                MAX(CAST(r.gained_price AS INTEGER)) as max_collection,
                MAX(CAST(r.profit_loss AS INTEGER)) as max_profit
               FROM Record r{where}
               GROUP BY bucket
               ORDER BY MIN(r.event_time) ASC""",
            where_args,
        ).fetchall()

        rows = []
        for row in trend_rows:
            item = dict(row)
            row_total = _safe_int(item.get("total"))
            row_escaped = _safe_int(item.get("escaped"))
            row_player_kills = _safe_int(item.get("total_player_kills"))
            row_fail = max(row_total - row_escaped, 0)
            rows.append(
                {
                    "bucket": item.get("bucket") or "",
                    "label": item.get("bucket") or "",
                    "total": row_total,
                    "escaped": row_escaped,
                    "evac_rate_num": round(row_escaped / row_total * 100, 1) if row_total else 0,
                    "kd": round(row_player_kills / row_fail, 2) if row_fail > 0 else row_player_kills,
                    "total_player_kills": row_player_kills,
                    "total_profit": round(float(item.get("total_profit") or 0), 0),
                    "avg_profit": round(float(item.get("avg_profit") or 0), 0),
                    "avg_duration": round(float(item.get("avg_duration") or 0), 0),
                    "max_collection": _safe_int(item.get("max_collection")),
                    "max_profit": _safe_int(item.get("max_profit")),
                }
            )

        resolved_start = (start_dt or min_time or "")[:10]
        resolved_end = (end_dt or max_time or "")[:10]
        self._json(
            {
                "ok": True,
                "bucket": bucket,
                "range": {"key": range_key, "start": resolved_start, "end": resolved_end},
                "summary": summary,
                "rows": rows,
            }
        )

    def _api_analysis(self, params):
        try:
            _require_open_source_feature("analysis")
        except PermissionError as e:
            self._json({"ok": False, "error": str(e), "requires_member": True}, 403)
            return
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        pid = self.db.get_active_player_id()
        start_dt = params.get("start", [""])[0]
        end_dt = params.get("end", [""])[0]

        where_parts = []
        where_args = []
        if pid:
            where_parts.append("r.player_id = ?")
            where_args.append(pid)
        if start_dt:
            where_parts.append("r.event_time >= ?")
            where_args.append(start_dt)
        if end_dt:
            where_parts.append("r.event_time <= ?")
            where_args.append(end_dt)

        w = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        a = where_args if where_args else []

        total_row = self.db.conn.execute(
            f"SELECT COUNT(*) as c FROM Record r{w}", a
        ).fetchone()
        total = total_row["c"] if total_row else 0
        if total == 0:
            self._json({"error": "暂无数据"})
            return

        stats = self.db.get_battle_stats(
            player_id=pid, start_dt=start_dt, end_dt=end_dt
        )
        max_coll = (
            self.db.conn.execute(
                f"SELECT MAX(CAST(gained_price AS INTEGER)) as m FROM Record r{w}",
                a,
            ).fetchone()["m"]
            or 0
        )
        max_prof = (
            self.db.conn.execute(
                f"SELECT MAX(CAST(profit_loss AS INTEGER)) as m FROM Record r{w}", a
            ).fetchone()["m"]
            or 0
        )

        escaped = stats.get("escaped", 0) or 0
        total_player_kills = stats.get("total_player_kills", 0) or 0
        fail_count = max(total - escaped, 0)
        kd = (
            round(total_player_kills / fail_count, 2)
            if fail_count > 0
            else total_player_kills
        )

        result_dist_row = self.db.conn.execute(
            f"""SELECT
                SUM(CASE WHEN game_result=0 THEN 1 ELSE 0 END) as success,
                SUM(CASE WHEN game_result=1 THEN 1 ELSE 0 END) as fail,
                SUM(CASE WHEN game_result=2 THEN 1 ELSE 0 END) as timeout,
                SUM(CASE WHEN game_result=3 THEN 1 ELSE 0 END) as leave
               FROM Record r{w}""",
            a,
        ).fetchone()

        map_stats = self.db.conn.execute(
            f"""SELECT r.map_id, m.map_name,
                COUNT(*) as count,
                SUM(CASE WHEN r.game_result=0 THEN 1 ELSE 0 END) as escaped,
                ROUND(AVG(CAST(r.game_result=0 AS REAL))*100,1) as evac_rate_num,
                ROUND(AVG(r.kill_player),1) as avg_player_kills,
                ROUND(AVG(CAST(r.profit_loss AS REAL)),0) as avg_profit,
                ROUND(AVG(r.duration_s),0) as avg_duration
               FROM Record r
               LEFT JOIN MapID m ON r.map_id = m.map_id
               {w}
               GROUP BY r.map_id
               ORDER BY evac_rate_num DESC, count DESC""",
            a,
        ).fetchall()

        role_stats = self.db.conn.execute(
            f"""SELECT r.role_id, rl.role_name,
                COUNT(*) as count,
                ROUND(AVG(CAST(r.game_result=0 AS REAL))*100,1) as evac_rate_num,
                ROUND(AVG(r.kill_player),1) as avg_player_kills,
                ROUND(AVG(CAST(r.profit_loss AS REAL)),0) as avg_profit
               FROM Record r
               LEFT JOIN Role rl ON r.role_id = rl.role_id
               {w}
               GROUP BY r.role_id
               ORDER BY evac_rate_num DESC, count DESC""",
            a,
        ).fetchall()

        top_games = self.db.conn.execute(
            f"""SELECT r.*, m.map_name, rl.role_name
               FROM Record r
               LEFT JOIN MapID m ON r.map_id = m.map_id
               LEFT JOIN Role rl ON r.role_id = rl.role_id
               {w}
               ORDER BY CAST(r.profit_loss AS INTEGER) DESC
               LIMIT 10""",
            a,
        ).fetchall()

        ms = [dict(r) for r in map_stats]
        for m in ms:
            m["evac_rate"] = f"{m['evac_rate_num']}%"
        rs = [dict(r) for r in role_stats]
        for r in rs:
            r["evac_rate"] = f"{r['evac_rate_num']}%"

        self._json(
            {
                "total": total,
                "escaped": escaped,
                "evac_rate": f"{escaped / total * 100:.1f}%" if total else "0%",
                "kd": kd,
                "total_kills": stats.get("total_kills", 0) or 0,
                "total_player_kills": total_player_kills,
                "avg_profit": stats.get("avg_profit", 0) or 0,
                "avg_duration": stats.get("avg_duration", 0) or 0,
                "total_profit": stats.get("total_profit", 0) or 0,
                "max_collection": max_coll,
                "max_profit": max_prof,
                "result_dist": dict(result_dist_row) if result_dist_row else {},
                "map_stats": ms,
                "role_stats": rs,
                "top_games": [dict(r) for r in top_games],
            }
        )

    def _api_login(self):
        try:
            _require_open_source_session()
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        if not self._try_acquire_action():
            return
        try:
            result = _run_login(self.db)
            account = result["account"]
            binding_status = _account_availability_status(account)
            self._json(
                {
                    "ok": True,
                    "active_account": account.get("player_name", "")
                    or account.get("player_id", ""),
                    "active_player_id": account.get("player_id", ""),
                    "logs": result.get("logs", []),
                    "account_status": binding_status,
                    "software_session": _open_source_session_snapshot(),
                }
            )
        except Exception as e:
            self._json({"ok": False, "error": str(e)})
        finally:
            self.action_lock.release()

    def _api_logout(self):
        if not self._try_acquire_action():
            return
        try:
            _cleanup_credentials()
            self._json({"ok": True})
        except Exception as e:
            self._json({"ok": False, "error": str(e)})
        finally:
            self.action_lock.release()

    def _api_account_select(self):
        body = self._read_json_body()
        player_id = str(body.get("player_id", "")).strip()
        if not player_id:
            self._json({"ok": False, "error": "missing player_id"}, 400)
            return
        try:
            account = self.db.get_account(player_id)
            if not account:
                self._json({"ok": False, "error": "账号不存在"}, 404)
                return
            self.db.set_active_account(player_id)
            account = self.db.get_active_account() or {}
            self._json(
                {
                    "ok": True,
                    "active_account": account.get("player_name", "")
                    or account.get("player_id", ""),
                    "active_player_id": account.get("player_id", ""),
                }
            )
        except Exception as e:
            self._json({"ok": False, "error": str(e)})

    def _api_account_delete(self):
        body = self._read_json_body()
        player_id = str(body.get("player_id", "")).strip()
        if not player_id:
            self._json({"ok": False, "error": "missing player_id"}, 400)
            return
        if not self._try_acquire_action():
            return
        try:
            result = _delete_account(self.db, player_id)
            self._json({"ok": True, **result})
        except Exception as e:
            self._json({"ok": False, "error": str(e)})
        finally:
            self.action_lock.release()

    def _api_fetch_details(self):
        try:
            _require_open_source_session()
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        if not self._try_acquire_action():
            return
        job_id = _create_action_job("fetch-details")
        _run_action_in_background(
            job_id,
            lambda db, log_sink, progress_sink: _run_fetch_missing_details(
                db, log_sink=log_sink
            ),
        )
        self._json({"ok": True, "job_id": job_id})

    def _api_clear_data(self):
        try:
            _require_open_source_session()
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        if not self._try_acquire_action():
            return
        try:
            result = _clear_data(self.db)
            self._json({"ok": True, **result})
        except Exception as e:
            self._json({"ok": False, "error": str(e)})
        finally:
            self.action_lock.release()

    def _api_fetch(self):
        body = self._read_json_body()
        queue = body.get("queue", "sol")
        count = int(body.get("count", 100) or 100)
        try:
            _require_open_source_session()
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        if not self._try_acquire_action():
            return
        job_id = _create_action_job("fetch")
        _run_action_in_background(
            job_id,
            lambda db, log_sink, progress_sink: _run_fetch(
                db, queue, count, log_sink=log_sink, progress_sink=progress_sink
            ),
        )
        self._json({"ok": True, "job_id": job_id})

    def _api_fetch_smart(self):
        body = self._read_json_body()
        queue = body.get("queue", "sol")
        count = int(body.get("count", 100) or 100)
        try:
            _require_open_source_session()
        except RuntimeError as e:
            self._json({"ok": False, "error": str(e), "requires_software_login": True}, 403)
            return
        if not self._try_acquire_action():
            return
        job_id = _create_action_job("fetch")
        _run_action_in_background(
            job_id,
            lambda db, log_sink, progress_sink: _run_fetch_smart(
                db, queue, count, log_sink=log_sink, progress_sink=progress_sink
            ),
        )
        self._json({"ok": True, "job_id": job_id})

    def _api_backup_export(self):
        if not self._try_acquire_action():
            return
        try:
            package = self.db.export_backup_package()
            self._json(
                {
                    "ok": True,
                    "filename": package["filename"],
                    "content_base64": base64.b64encode(package["content"]).decode("ascii"),
                    "manifest": package["manifest"],
                }
            )
        except Exception as exc:
            self._json({"ok": False, "error": str(exc)}, 400)
        finally:
            self.action_lock.release()

    def _api_backup_import(self):
        if not self._try_acquire_action():
            return
        try:
            body = self._read_json_body()
            content_base64 = str(body.get("content_base64", "") or "").strip()
            if not content_base64:
                raise ValueError("未选择备份文件")
            raw = base64.b64decode(content_base64.encode("ascii"))
            result = self.db.import_backup_package(raw)
            self._json({"ok": True, **result})
        except Exception as exc:
            self._json({"ok": False, "error": str(exc)}, 400)
        finally:
            self.action_lock.release()


def start_server(
    port: int = 8080,
    db: Database = None,
    fetch_callback=None,
    open_browser: bool = True,
):
    APIHandler.db = db
    APIHandler.fetch_callback = fetch_callback
    preferred_port = int(port or 8080)
    actual_port = preferred_port
    server = None
    last_error = None

    for candidate_port in [preferred_port] + [preferred_port + offset for offset in range(1, 20)]:
        try:
            server = HTTPServer(("0.0.0.0", candidate_port), APIHandler)
            actual_port = candidate_port
            break
        except OSError as exc:
            last_error = exc

    if server is None:
        raise last_error

    if open_browser:
        threading.Timer(
            0.5, lambda: webbrowser.open(f"http://127.0.0.1:{actual_port}")
        ).start()

    if actual_port != preferred_port:
        print(f"[web] 端口 {preferred_port} 不可用，已自动切换到 {actual_port}")

    print(f"[web] 服务已启动: http://127.0.0.1:{actual_port} (监听 0.0.0.0:{actual_port})")
    print("[web] 按 Ctrl+C 停止")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[web] 已停止")
    finally:
        server.server_close()

